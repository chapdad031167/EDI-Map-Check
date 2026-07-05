"""Color-coded Excel export of a validation run."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mapcheck.engine.results import RunResult, Status

_STATUS_FILLS = {
    Status.PASS: PatternFill("solid", fgColor="C6EFCE"),
    Status.FAIL: PatternFill("solid", fgColor="FFC7CE"),
    Status.WARNING: PatternFill("solid", fgColor="FFEB9C"),
    Status.NOT_TESTED: PatternFill("solid", fgColor="D9D9D9"),
}
_STATUS_FONTS = {
    Status.PASS: Font(color="006100"),
    Status.FAIL: Font(color="9C0006"),
    Status.WARNING: Font(color="9C6500"),
    Status.NOT_TESTED: Font(color="595959"),
}
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_RESULT_COLUMNS = (
    ("Status", 12),
    ("Row ID", 9),
    ("Source", 22),
    ("Target", 28),
    ("Expected", 24),
    ("Actual", 24),
    ("Category", 20),
    ("Message", 70),
)


def export_excel(result: RunResult, path: str | Path) -> Path:
    """Write the run to a workbook: Results sheet + Summary sheet."""
    path = Path(path)
    wb = Workbook()

    ws = wb.active
    ws.title = "Results"
    for idx, (header, width) in enumerate(_RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    for row_idx, finding in enumerate(result.sorted_findings(), start=2):
        values = (
            finding.status.value,
            finding.row_id,
            finding.source_ref,
            finding.target,
            finding.expected,
            finding.actual,
            finding.category.value if finding.category else None,
            finding.message,
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 8))
        status_cell = ws.cell(row=row_idx, column=1)
        status_cell.fill = _STATUS_FILLS[finding.status]
        status_cell.font = _STATUS_FONTS[finding.status]
    ws.auto_filter.ref = f"A1:H{max(len(result.findings) + 1, 2)}"

    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 70

    def put(row: int, key: str, value: object, bold: bool = False) -> None:
        key_cell = ws_sum.cell(row=row, column=1, value=key)
        key_cell.font = Font(bold=True)
        value_cell = ws_sum.cell(row=row, column=2, value=value)
        if bold:
            value_cell.font = Font(bold=True)

    put(1, "EDI MapCheck report", result.spec_name or "", bold=True)
    put(2, "Run at (UTC)", result.run_at.strftime("%Y-%m-%d %H:%M:%S"))
    put(3, "Transaction", f"{result.transaction_set or ''} {result.transaction_name or ''}".strip())
    put(4, "Spec", result.spec_path)
    put(5, "Source", result.source_path)
    put(6, "Output", result.output_path)
    put(7, "Overall result", result.overall.value, bold=True)
    row = 9
    put(row, "Status counts", "")
    for status, count in result.counts.items():
        row += 1
        ws_sum.cell(row=row, column=1, value=status.value)
        ws_sum.cell(row=row, column=2, value=count)
    row += 2
    put(row, "Root causes", "")
    for category, count in result.category_counts.items():
        row += 1
        ws_sum.cell(row=row, column=1, value=category.value)
        ws_sum.cell(row=row, column=2, value=count)

    wb.save(path)
    return path
