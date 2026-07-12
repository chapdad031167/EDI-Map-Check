"""Partner overrides: one base spec plus small per-partner delta workbooks.

A distributor runs the same map for many partners with small deviations.
Rather than fork the whole spec per partner, keep a base spec and a delta
workbook that adds, replaces, or removes rules by Row ID and shadows code
lists by name. :func:`resolve_spec` merges them — after load — into one
:class:`~mapcheck.spec.model.MappingSpec` the validator consumes unchanged.

Deltas are ordinary spec workbooks (so ``import-spec`` / ``init-spec``
author them for free); the only extra convention is a ``REMOVE`` Rule Type
that deletes the base rule with that Row ID. Every merged rule carries an
``origin`` — ``"base"`` or ``"partner:<name>"`` — so a finding reads as
"the ACME override is wrong" rather than "the base map is wrong".
"""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from mapcheck.spec.model import Direction, MappingRule, MappingSpec
from mapcheck.spec.parser import (
    _EXPECTED_CODELIST_HEADER,
    _EXPECTED_MAPPING_HEADER,
    _cell_str,
    _check_header,
    _load_code_lists,
    _load_meta,
    _parse_rule,
    _sheet,
    SpecLoadError,
    load_spec,
)
from mapcheck.spec.template import MAPPING_COLUMNS

#: Rule Type cell value marking a delta row that deletes a base rule.
REMOVE = "REMOVE"

_RULE_TYPE_COL = 4  # 0-based index of Rule Type in a Mapping row


class DeltaSpec:
    """A parsed partner delta: rules to add/replace, ids to remove, lists."""

    def __init__(
        self,
        rules: list[MappingRule],
        removals: list[str],
        code_lists: dict,
        meta: dict,
        direction: Direction,
        label: str,
    ) -> None:
        self.rules = rules
        self.removals = removals
        self.code_lists = code_lists
        self.meta = meta
        self.direction = direction
        self.label = label


def _partner_label(path: Path, meta: dict) -> str:
    name = meta.get("Partner") or path.stem
    return f"partner:{name.strip().lower().replace(' ', '_')}"


def load_delta(path: str | Path) -> DeltaSpec:
    """Load a partner delta workbook (a spec workbook with REMOVE rows)."""
    path = Path(path)
    if not path.exists():
        raise SpecLoadError([f"partner delta file not found: {path}"])

    errors: list[str] = []
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws_map = _sheet(wb, "Mapping", errors)
        ws_codes = _sheet(wb, "CodeLists", errors)
        ws_meta = _sheet(wb, "Meta", errors)
        if errors:
            raise SpecLoadError(errors)
        assert ws_map and ws_codes and ws_meta

        _check_header(ws_map, _EXPECTED_MAPPING_HEADER, "Mapping", errors)
        _check_header(ws_codes, _EXPECTED_CODELIST_HEADER, "CodeLists", errors)
        if errors:
            raise SpecLoadError(errors)

        meta = _load_meta(ws_meta)
        code_lists = _load_code_lists(ws_codes, errors)

        direction = Direction.INBOUND
        if (raw := meta.get("Direction")) is not None:
            try:
                direction = Direction(raw.lower())
            except ValueError:
                errors.append(f"Meta Direction must be 'inbound' or 'outbound', got {raw!r}")

        rules: list[MappingRule] = []
        removals: list[str] = []
        notes: list[str] = []
        n_cols = len(MAPPING_COLUMNS)
        for idx, row in enumerate(
            ws_map.iter_rows(min_row=2, max_col=n_cols, values_only=True), start=2
        ):
            cells = tuple(_cell_str(v) for v in row)
            if not any(cells):
                continue
            if (cells[_RULE_TYPE_COL] or "").upper() == REMOVE:
                row_id = cells[0]
                if not row_id:
                    errors.append(f"Mapping row {idx}: REMOVE requires a Row ID")
                else:
                    removals.append(row_id)
                continue
            rule = _parse_rule(idx, cells, code_lists, errors, direction, notes)
            if rule is not None:
                rules.append(rule)

        if errors:
            raise SpecLoadError(errors)
        return DeltaSpec(
            rules=rules, removals=removals, code_lists=code_lists,
            meta=meta, direction=direction, label=_partner_label(path, meta),
        )
    finally:
        wb.close()


def merge_specs(base: MappingSpec, delta: DeltaSpec) -> tuple[MappingSpec, list[str]]:
    """Merge a delta onto a base spec; returns ``(merged, warnings)``.

    Rules are keyed by Row ID: a delta rule replaces the base rule with the
    same id (keeping its position) or is appended when new; a removal drops
    the base rule. Code lists shadow whole-list. Every merged rule is
    stamped with its origin.
    """
    errors: list[str] = []
    if delta.direction is not base.direction:
        errors.append(
            f"partner delta direction {delta.direction.value!r} does not match "
            f"the base spec direction {base.direction.value!r}"
        )
    base_set = base.meta.get("Transaction Set")
    delta_set = delta.meta.get("Transaction Set")
    if delta_set and base_set and delta_set != base_set:
        errors.append(
            f"partner delta is for transaction set {delta_set!r} but the base "
            f"spec is {base_set!r}"
        )
    if errors:
        raise SpecLoadError(errors)

    warnings: list[str] = []
    by_id: dict[str, MappingRule] = {r.row_id: r for r in base.rules}
    order: list[str] = [r.row_id for r in base.rules]

    for row_id in delta.removals:
        if row_id in by_id:
            del by_id[row_id]
            order.remove(row_id)
        else:
            warnings.append(f"REMOVE for Row ID {row_id!r} not found in the base spec")

    for rule in delta.rules:
        stamped = replace(rule, origin=delta.label)
        if rule.row_id not in by_id:
            order.append(rule.row_id)
        by_id[rule.row_id] = stamped

    merged_rules = tuple(by_id[row_id] for row_id in order)
    merged_code_lists = {**base.code_lists, **delta.code_lists}
    merged_meta = {**base.meta}
    merged_meta["Partner"] = delta.label.split(":", 1)[-1]

    merged = MappingSpec(
        meta=merged_meta,
        rules=merged_rules,
        code_lists=merged_code_lists,
        source_path=base.source_path,
        direction=base.direction,
        load_notes=base.load_notes + tuple(warnings),
    )
    return merged, warnings


def resolve_spec(base_path: str | Path, partner_path: str | Path) -> tuple[MappingSpec, list[str]]:
    """Load a base spec and a partner delta and return the merged spec."""
    base = load_spec(base_path)
    delta = load_delta(partner_path)
    return merge_specs(base, delta)


def _rule_row(rule: MappingRule) -> dict[str, str]:
    """Serialize a rule back to the native Mapping columns."""
    notes = rule.notes or ""
    if rule.origin != "base":
        tag = f"[{rule.origin}]"
        notes = f"{notes} {tag}".strip() if notes else tag
    return {
        "Row ID": rule.row_id,
        "Source Field": rule.source_field or "",
        "Loop Context": str(rule.loop_context) if rule.loop_context else "",
        "Target Field": rule.target_field,
        "Rule Type": rule.rule_type.value,
        "Condition (Text)": rule.condition_text or "",
        "Condition (Coded)": rule.condition.raw if rule.condition else "",
        "Then": str(rule.then_outcome) if rule.then_outcome else "",
        "Else": str(rule.else_outcome) if rule.else_outcome else "",
        "Default Value": rule.default_value or "",
        "Code List Ref": rule.code_list_ref or "",
        "Data Type": rule.data_type.value if rule.data_type else "",
        "Format": rule.format or "",
        "Notes": notes,
    }


def export_spec(spec: MappingSpec, path: str | Path) -> Path:
    """Write a (merged) spec to a native template workbook.

    Each rule's origin is recorded in Notes so the effective, resolved spec
    is auditable as one sheet. Loads and round-trips through ``load_spec``.
    """
    from mapcheck.spec.template import build_template

    path = Path(path)
    direction = spec.direction.value
    wb = build_template(direction=direction)
    ws = wb["Mapping"]
    headers = [h for h, _ in MAPPING_COLUMNS]
    col = {h: i + 1 for i, h in enumerate(headers)}
    for r, rule in enumerate(spec.rules, start=2):
        for header, value in _rule_row(rule).items():
            if value:
                ws.cell(row=r, column=col[header], value=value)

    ws_codes = wb["CodeLists"]
    cr = 2
    for name, code_list in spec.code_lists.items():
        for source, target in code_list.entries.items():
            ws_codes.cell(row=cr, column=1, value=name)
            ws_codes.cell(row=cr, column=2, value=source)
            ws_codes.cell(row=cr, column=3, value=target)
            if desc := code_list.descriptions.get(source):
                ws_codes.cell(row=cr, column=4, value=desc)
            cr += 1

    ws_meta = wb["Meta"]
    existing = {ws_meta.cell(row=i, column=1).value: i for i in range(2, ws_meta.max_row + 1)}
    nxt = ws_meta.max_row + 1
    for key, value in spec.meta.items():
        if key in existing:
            ws_meta.cell(row=existing[key], column=2, value=value)
        else:
            ws_meta.cell(row=nxt, column=1, value=key)
            ws_meta.cell(row=nxt, column=2, value=value)
            nxt += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
