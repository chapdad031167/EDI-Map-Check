"""Loader for the Excel spec template.

Reads the workbook produced by :mod:`mapcheck.spec.template` (or a
hand-filled copy of it) into a validated :class:`~mapcheck.spec.model.MappingSpec`.

The loader is strict on purpose: a broken spec fails here, with Excel row
numbers, before any EDI file is touched. All problems are collected and
reported together rather than one at a time.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mapcheck.spec.conditions import ConditionSyntaxError, parse_condition, parse_outcome
from mapcheck.spec.model import (
    CodeList,
    DataType,
    LoopContext,
    MappingRule,
    MappingSpec,
    Outcome,
    OutcomeKind,
    RuleType,
    split_source_field,
)
from mapcheck.spec.template import CODELIST_COLUMNS, MAPPING_COLUMNS


class SpecLoadError(Exception):
    """Raised when a spec workbook cannot be loaded; carries all row errors."""

    def __init__(self, errors: list[str]) -> None:
        self.errors = errors
        super().__init__(
            f"{len(errors)} problem(s) in mapping spec:\n  - " + "\n  - ".join(errors)
        )


def _cell_str(value: Any) -> str | None:
    """Normalize a cell to a stripped string, or None when empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _sheet(wb: Any, name: str, errors: list[str]) -> Worksheet | None:
    if name not in wb.sheetnames:
        errors.append(f"missing required sheet {name!r}")
        return None
    return wb[name]


def _load_meta(ws: Worksheet) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = (_cell_str(v) for v in row)
        if key and value:
            meta[key] = value
    return meta


def _load_code_lists(ws: Worksheet, errors: list[str]) -> dict[str, CodeList]:
    entries: dict[str, dict[str, str]] = {}
    descriptions: dict[str, dict[str, str]] = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        name, source, target, desc = (_cell_str(v) for v in row)
        if not any((name, source, target)):
            continue
        if not (name and source and target is not None):
            errors.append(
                f"CodeLists row {idx}: List Name, Source Value and Target Value are all required"
            )
            continue
        table = entries.setdefault(name, {})
        if source in table:
            errors.append(
                f"CodeLists row {idx}: duplicate source value {source!r} in list {name!r}"
            )
            continue
        table[source] = target
        if desc:
            descriptions.setdefault(name, {})[source] = desc
    return {
        name: CodeList(name=name, entries=table, descriptions=descriptions.get(name, {}))
        for name, table in entries.items()
    }


_EXPECTED_MAPPING_HEADER = tuple(header for header, _ in MAPPING_COLUMNS)
_EXPECTED_CODELIST_HEADER = tuple(header for header, _ in CODELIST_COLUMNS)


def _check_header(
    ws: Worksheet, expected: tuple[str, ...], sheet_name: str, errors: list[str]
) -> bool:
    actual = tuple(
        _cell_str(c) or "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    )[: len(expected)]
    if actual != expected:
        errors.append(
            f"{sheet_name} sheet header mismatch: expected {list(expected)}, got {list(actual)} "
            "(was this file created from the MapCheck template?)"
        )
        return False
    return True


def _parse_rule(
    idx: int,
    cells: tuple[str | None, ...],
    code_lists: dict[str, CodeList],
    errors: list[str],
) -> MappingRule | None:
    (
        row_id,
        source_field,
        loop_context_raw,
        target_field,
        rule_type_raw,
        condition_text,
        condition_coded,
        then_raw,
        else_raw,
        default_value,
        code_list_ref,
        data_type_raw,
        format_raw,
        notes,
    ) = cells

    def err(message: str) -> None:
        errors.append(f"Mapping row {idx}: {message}")

    ok = True
    if not row_id:
        err("Row ID is required")
        ok = False
    if not target_field:
        err("Target Field is required")
        ok = False

    rule_type: RuleType | None = None
    if not rule_type_raw:
        err("Rule Type is required")
        ok = False
    else:
        try:
            rule_type = RuleType(rule_type_raw.upper())
        except ValueError:
            err(
                f"unknown rule type {rule_type_raw!r} "
                f"(expected one of {', '.join(rt.value for rt in RuleType)})"
            )
            ok = False

    loop_context: LoopContext | None = None
    if loop_context_raw:
        try:
            loop_context = LoopContext.parse(loop_context_raw)
        except ValueError as exc:
            err(str(exc))
            ok = False

    if source_field and rule_type is not RuleType.LOOP_COUNT:
        try:
            split_source_field(source_field)
        except ValueError as exc:
            err(str(exc))
            ok = False

    condition = None
    then_outcome = else_outcome = None
    if condition_coded:
        try:
            condition = parse_condition(condition_coded)
        except ConditionSyntaxError as exc:
            err(f"Condition (Coded): {exc}")
            ok = False
    if then_raw:
        try:
            then_outcome = parse_outcome(then_raw)
        except ConditionSyntaxError as exc:
            err(f"Then: {exc}")
            ok = False
    if else_raw:
        try:
            else_outcome = parse_outcome(else_raw)
        except ConditionSyntaxError as exc:
            err(f"Else: {exc}")
            ok = False

    data_type: DataType | None = None
    if data_type_raw:
        try:
            data_type = DataType(data_type_raw.lower())
        except ValueError:
            err(
                f"unknown data type {data_type_raw!r} "
                f"(expected one of {', '.join(dt.value for dt in DataType)})"
            )
            ok = False

    # Cross-field requirements per rule type.
    if rule_type is RuleType.DIRECT and not source_field:
        err("DIRECT rule requires a Source Field")
        ok = False
    if rule_type is RuleType.CODE_LIST:
        if not source_field:
            err("CODE_LIST rule requires a Source Field")
            ok = False
        if not code_list_ref:
            err("CODE_LIST rule requires a Code List Ref")
            ok = False
        elif code_list_ref not in code_lists:
            err(
                f"Code List Ref {code_list_ref!r} not found on the CodeLists sheet "
                f"(available: {', '.join(sorted(code_lists)) or 'none'})"
            )
            ok = False
    if rule_type is RuleType.CONSTANT:
        if default_value is None:
            err("CONSTANT rule requires a Default Value")
            ok = False
        if source_field:
            err("CONSTANT rule must not have a Source Field")
            ok = False
    if rule_type is RuleType.CONDITIONAL:
        if condition is None:
            err("CONDITIONAL rule requires a Condition (Coded)")
            ok = False
        if then_outcome is None:
            err("CONDITIONAL rule requires a Then outcome")
            ok = False
        needs_source = OutcomeKind.SOURCE in {
            o.kind for o in (then_outcome, else_outcome) if o is not None
        }
        if needs_source and not source_field:
            err("CONDITIONAL rule with a SOURCE outcome requires a Source Field")
            ok = False
    elif condition is not None:
        err(f"Condition (Coded) is only allowed on CONDITIONAL rules, not {rule_type_raw}")
        ok = False
    if rule_type is RuleType.LOOP_COUNT:
        if not source_field:
            err("LOOP_COUNT rule requires a Source Field naming the loop (e.g. PO1)")
            ok = False
        elif not source_field.isalnum():
            err(f"LOOP_COUNT source must be a bare loop id like 'PO1', got {source_field!r}")
            ok = False

    if rule_type is RuleType.CONDITIONAL and then_outcome is not None and else_outcome is None:
        else_outcome = Outcome(kind=OutcomeKind.SKIP)

    per_line = target_field is not None and "[]" in target_field
    if per_line and (loop_context is None or loop_context.qualifier is not None):
        if rule_type is not RuleType.LOOP_COUNT:
            err(
                f"per-line target {target_field!r} requires a repeating Loop Context "
                "(bare loop id like 'PO1')"
            )
            ok = False
    if not per_line and loop_context is not None and loop_context.qualifier is None:
        err(
            f"repeating loop context {loop_context_raw!r} requires a per-line target "
            "(lines[] notation) — or qualify the occurrence, e.g. 'N1[ST]'"
        )
        ok = False

    if not ok:
        return None
    assert row_id and target_field and rule_type
    return MappingRule(
        row_id=row_id,
        sheet_row=idx,
        target_field=target_field,
        rule_type=rule_type,
        source_field=source_field.upper() if source_field else None,
        loop_context=loop_context,
        condition_text=condition_text,
        condition=condition,
        then_outcome=then_outcome,
        else_outcome=else_outcome,
        default_value=default_value,
        code_list_ref=code_list_ref,
        data_type=data_type,
        format=format_raw,
        notes=notes,
    )


def load_spec(path: str | Path) -> MappingSpec:
    """Load and validate a spec workbook.

    Raises :class:`SpecLoadError` (with every problem found, referenced by
    Excel row number) if the workbook is not a valid MapCheck spec.
    """
    path = Path(path)
    if not path.exists():
        raise SpecLoadError([f"spec file not found: {path}"])

    errors: list[str] = []
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws_map = _sheet(wb, "Mapping", errors)
        ws_codes = _sheet(wb, "CodeLists", errors)
        ws_meta = _sheet(wb, "Meta", errors)
        if errors:
            raise SpecLoadError(errors)
        assert ws_map and ws_codes and ws_meta

        _check_header(ws_map, _EXPECTED_MAPPING_HEADER, "Mapping", errors)
        _check_header(ws_codes, _EXPECTED_CODELIST_HEADER, "CodeLists", errors)
        if errors:
            raise SpecLoadError(errors)

        meta = _load_meta(ws_meta)
        code_lists = _load_code_lists(ws_codes, errors)

        rules: list[MappingRule] = []
        seen_ids: dict[str, int] = {}
        n_cols = len(MAPPING_COLUMNS)
        for idx, row in enumerate(
            ws_map.iter_rows(min_row=2, max_col=n_cols, values_only=True), start=2
        ):
            cells = tuple(_cell_str(v) for v in row)
            if not any(cells):
                continue
            rule = _parse_rule(idx, cells, code_lists, errors)
            if rule is None:
                continue
            if rule.row_id in seen_ids:
                errors.append(
                    f"Mapping row {idx}: duplicate Row ID {rule.row_id!r} "
                    f"(first used on row {seen_ids[rule.row_id]})"
                )
                continue
            seen_ids[rule.row_id] = idx
            rules.append(rule)

        if not rules and not errors:
            errors.append("Mapping sheet contains no rules")
        if errors:
            raise SpecLoadError(errors)

        return MappingSpec(
            meta=meta,
            rules=tuple(rules),
            code_lists=code_lists,
            source_path=str(path),
        )
    finally:
        wb.close()
