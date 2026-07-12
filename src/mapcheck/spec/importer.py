"""Import an arbitrary partner mapping document into a MapCheck spec.

Real mapping specs live in partner Excel workbooks and CSVs with every
imaginable column layout, prose conditions, and inconsistent rule
expression. This module turns one of those into the native MapCheck
template — auto-classifying what the evidence supports and **flagging,
never silently guessing**, what it can't, so the human finishes a short
worklist instead of retyping the whole spec.

Three stages (design 004):

* **read** any ``.xlsx`` / ``.csv`` into uniform ``{our-column: value}``
  rows, locating the header row and mapping their headers to ours,
* **classify** each row's rule type from the evidence in its cells,
* **write** the native template with NEEDS REVIEW rows flagged, plus a
  worklist of everything that needs a human.

Only the spec template/loader is touched; the engine is untouched.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from mapcheck.spec.conditions import ConditionSyntaxError, parse_condition
from mapcheck.spec.model import RuleType
from mapcheck.spec.template import MAPPING_COLUMNS, build_template

# The native Mapping columns we classify into (subset that carries meaning
# for import; Row ID and Notes are generated/annotated).
SOURCE_FIELD = "Source Field"
LOOP_CONTEXT = "Loop Context"
TARGET_FIELD = "Target Field"
RULE_TYPE = "Rule Type"
CONDITION_TEXT = "Condition (Text)"
CONDITION_CODED = "Condition (Coded)"
THEN = "Then"
ELSE = "Else"
DEFAULT_VALUE = "Default Value"
CODE_LIST_REF = "Code List Ref"
DATA_TYPE = "Data Type"
FORMAT = "Format"

#: our column -> normalized header fragments that identify it in a partner doc
_SYNONYMS: dict[str, tuple[str, ...]] = {
    SOURCE_FIELD: ("source field", "source", "x12 element", "x12", "element",
                   "segment element", "edi element", "source element", "from"),
    LOOP_CONTEXT: ("loop context", "loop", "context", "segment loop", "occurrence"),
    TARGET_FIELD: ("target field", "target", "destination", "erp field", "idoc field",
                   "field name", "internal field", "to", "output field", "field",
                   "destination field"),
    RULE_TYPE: ("rule type", "mapping type", "map type", "rule kind", "rule"),
    CONDITION_TEXT: ("condition", "logic", "rule logic", "condition text",
                     "notes logic", "mapping logic", "comments logic",
                     "business rule", "mapping rule", "instructions", "instruction"),
    CONDITION_CODED: ("condition coded", "coded condition", "condition code"),
    THEN: ("then", "if true", "true value"),
    ELSE: ("else", "if false", "false value", "otherwise"),
    DEFAULT_VALUE: ("default value", "default", "constant", "hardcode", "hard code",
                    "literal", "fixed value", "hardcoded"),
    CODE_LIST_REF: ("code list ref", "code list", "codelist", "lookup", "xref",
                    "cross reference", "translation", "translate", "value map"),
    DATA_TYPE: ("data type", "datatype", "type of data", "type"),
    FORMAT: ("format", "picture", "data format", "output format"),
}

#: partner Rule Type cell text -> our RuleType
_RULE_TYPE_WORDS: dict[str, RuleType] = {
    "direct": RuleType.DIRECT, "move": RuleType.DIRECT, "copy": RuleType.DIRECT,
    "passthrough": RuleType.DIRECT, "pass through": RuleType.DIRECT,
    "constant": RuleType.CONSTANT, "hardcode": RuleType.CONSTANT,
    "hardcoded": RuleType.CONSTANT, "default": RuleType.CONSTANT,
    "literal": RuleType.CONSTANT, "fixed": RuleType.CONSTANT,
    "conditional": RuleType.CONDITIONAL, "condition": RuleType.CONDITIONAL,
    "if": RuleType.CONDITIONAL, "if/then": RuleType.CONDITIONAL,
    "code_list": RuleType.CODE_LIST, "code list": RuleType.CODE_LIST,
    "codelist": RuleType.CODE_LIST, "xref": RuleType.CODE_LIST,
    "lookup": RuleType.CODE_LIST, "translate": RuleType.CODE_LIST,
    "translation": RuleType.CODE_LIST,
    "loop_count": RuleType.LOOP_COUNT, "loop count": RuleType.LOOP_COUNT,
    "count": RuleType.LOOP_COUNT,
}


class ImportError_(Exception):
    """Raised when a mapping document can't be read or column-mapped."""


def _norm(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").strip().lower()).strip()


# --------------------------------------------------------------------------
# Stage 1: read any tabular file into rows
# --------------------------------------------------------------------------


@dataclass
class Table:
    headers: list[str]
    rows: list[dict[str, str]]  # keyed by header (raw header text)


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_header(cells: list[str]) -> bool:
    """A header row is mostly short, non-empty, non-numeric labels."""
    filled = [c for c in cells if c]
    if len(filled) < 2:
        return False
    shortish = [c for c in filled if len(c) <= 40 and not _is_number(c)]
    return len(shortish) >= max(2, len(filled) * 0.6)


def _is_number(text: str) -> bool:
    try:
        float(text.replace(",", ""))
        return True
    except ValueError:
        return False


def _rows_from_grid(grid: list[list[str]]) -> Table:
    """Locate the header row (first plausible one) and build keyed rows."""
    header_index = next(
        (i for i, row in enumerate(grid) if _looks_like_header(row)), None
    )
    if header_index is None:
        raise ImportError_("could not locate a header row in the document")
    raw_headers = grid[header_index]
    # de-duplicate blank/repeat headers so keys stay unique
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        name = h or f"column_{idx + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        headers.append(name)
    rows = []
    for row in grid[header_index + 1:]:
        if not any(c for c in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    return Table(headers=headers, rows=rows)


def read_table(path: str | Path) -> Table:
    """Read an .xlsx or .csv mapping document into a :class:`Table`."""
    path = Path(path)
    if not path.exists():
        raise ImportError_(f"mapping document not found: {path}")
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            grid = [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    elif suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as fh:
            grid = [[_cell(c) for c in row] for row in csv.reader(fh)]
    else:
        raise ImportError_(f"unsupported document type {suffix!r} (expected .xlsx or .csv)")
    if not grid:
        raise ImportError_(f"{path}: document is empty")
    return _rows_from_grid(grid)


# --------------------------------------------------------------------------
# Stage 2: map their headers to our columns
# --------------------------------------------------------------------------


def resolve_columns(
    headers: list[str], overrides: dict[str, str] | None = None
) -> dict[str, str]:
    """Map each MapCheck column to one of ``headers``.

    ``overrides`` maps our column name to their header (exact) or an Excel
    column letter. Returns ``{our_column: their_header}`` for every column
    that resolved. Raises when the required Target Field can't be found.
    """
    overrides = overrides or {}
    norm_headers = {_norm(h): h for h in headers}
    letters = {get_col_letter(i): h for i, h in enumerate(headers)}

    mapping: dict[str, str] = {}
    claimed: set[str] = set()
    # explicit overrides first — they always win and reserve their header
    for our_col, token in overrides.items():
        if token in headers:
            resolved = token
        elif token.upper() in letters:
            resolved = letters[token.upper()]
        elif _norm(token) in norm_headers:
            resolved = norm_headers[_norm(token)]
        else:
            raise ImportError_(
                f"--map for {our_col!r} names {token!r}, which is not a header "
                "or column letter in the document"
            )
        mapping[our_col] = resolved
        claimed.add(resolved)
    # auto-detect the rest by quality-tiered global assignment: exact
    # header matches are claimed before word matches, which beat substrings.
    # So "Business Rule" (exact -> Condition) wins over "rule" -> Rule Type.
    candidates: list[tuple[int, str, str]] = []  # (quality, our_col, header)
    for our_col in _SYNONYMS:
        if our_col in mapping:
            continue
        for norm_h, raw_h in norm_headers.items():
            q = _match_quality(_SYNONYMS[our_col], norm_h)
            if q:
                candidates.append((q, our_col, raw_h))
    candidates.sort(key=lambda c: -c[0])
    for _q, our_col, raw_h in candidates:
        if our_col in mapping or raw_h in claimed:
            continue
        mapping[our_col] = raw_h
        claimed.add(raw_h)

    if TARGET_FIELD not in mapping:
        raise ImportError_(
            "could not identify the Target Field column — pass it explicitly, e.g. "
            "--map \"Target Field=<their header or column letter>\""
        )
    return mapping


def get_col_letter(index: int) -> str:
    """0-based column index -> Excel letter (A, B, ..., Z, AA, ...)."""
    letter = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letter = chr(65 + rem) + letter
    return letter


def _match_quality(fragments: tuple[str, ...], norm_h: str) -> int:
    """3 = header equals a fragment, 2 = fragment is a whole word, 1 = substring."""
    words = norm_h.split()
    if any(norm_h == frag for frag in fragments):
        return 3
    if any(frag in words for frag in fragments):
        return 2
    if any(frag in norm_h for frag in fragments):
        return 1
    return 0


# --------------------------------------------------------------------------
# Stage 3: classify each row
# --------------------------------------------------------------------------


@dataclass
class ImportedRow:
    row_id: str
    values: dict[str, str]  # our column -> value (Rule Type filled when known)
    rule_type: RuleType | None
    needs_review: bool
    reason: str


def _get(mapping: dict[str, str], raw: dict[str, str], our_col: str) -> str:
    header = mapping.get(our_col)
    return raw.get(header, "").strip() if header else ""


_SIMPLE_COND = re.compile(
    r"""(?:when|if|only\s+when|map\s+when)?\s*
        ([A-Z][A-Z0-9]{1,2}\d{2})\s*
        (=|==|!=|<>|\bis\b|\bequals\b|\bnot\b)\s*
        '?([A-Za-z0-9]+)'?""",
    re.IGNORECASE | re.VERBOSE,
)
_EXISTS_COND = re.compile(
    r"(?:when|if)?\s*([A-Z][A-Z0-9]{1,2}\d{2})\s+(?:is\s+)?(?:present|sent|exists|populated)",
    re.IGNORECASE,
)


def derive_condition(text: str) -> str | None:
    """Derive a coded condition from prose, or None when not translatable.

    Only patterns the existing grammar accepts are returned; anything else
    stays a human job (the raw prose is preserved in Condition (Text)).
    """
    raw = text.strip()
    if not raw:
        return None
    # EXISTS first — "REF02 is present" must not read as "= 'present'"
    if m := _EXISTS_COND.search(raw):
        coded = f"EXISTS({m.group(1).upper()})"
    elif m := _SIMPLE_COND.search(raw):
        field_, op, value = m.group(1).upper(), m.group(2).lower(), m.group(3)
        coded_op = "!=" if op in ("!=", "<>", "not") else "="
        coded = f"{field_} {coded_op} '{value}'"
    else:
        return None
    try:
        parse_condition(coded)
    except ConditionSyntaxError:
        return None
    return coded


def classify_row(mapping: dict[str, str], raw: dict[str, str], row_id: str) -> ImportedRow:
    """Classify one source row into a rule type, flagging when unsure."""
    source = _get(mapping, raw, SOURCE_FIELD)
    target = _get(mapping, raw, TARGET_FIELD)
    default = _get(mapping, raw, DEFAULT_VALUE)
    code_list = _get(mapping, raw, CODE_LIST_REF)
    cond_text = _get(mapping, raw, CONDITION_TEXT)
    explicit = _get(mapping, raw, RULE_TYPE)

    values: dict[str, str] = {}
    for col in (SOURCE_FIELD, LOOP_CONTEXT, TARGET_FIELD, DEFAULT_VALUE,
                CODE_LIST_REF, DATA_TYPE, FORMAT, THEN, ELSE):
        v = _get(mapping, raw, col)
        if v:
            values[col] = v
    if cond_text:
        values[CONDITION_TEXT] = cond_text

    def done(rt: RuleType | None, reason: str, review: bool) -> ImportedRow:
        if rt is not None:
            values[RULE_TYPE] = rt.value
        return ImportedRow(row_id, values, rt, review, reason)

    # explicit, recognizable rule type wins
    rt = _RULE_TYPE_WORDS.get(_norm(explicit)) if explicit else None
    reason_prefix = f"explicit type {explicit!r}" if rt else None

    if rt is None:
        # infer from evidence, conservatively
        if code_list:
            rt, reason_prefix = RuleType.CODE_LIST, f"inferred CODE_LIST: lookup {code_list!r}"
        elif default and not source:
            rt, reason_prefix = RuleType.CONSTANT, f"inferred CONSTANT: default {default!r}, no source"
        elif cond_text:
            rt, reason_prefix = RuleType.CONDITIONAL, "inferred CONDITIONAL: condition present"
        elif source:
            rt, reason_prefix = RuleType.DIRECT, "inferred DIRECT: source field, no other evidence"
        else:
            return done(None, "type unclear: no source field, default, lookup, or condition", True)

    # per-type completion + review flags
    if rt is RuleType.CODE_LIST:
        if not source:
            return done(rt, f"{reason_prefix}, but no source field to translate", True)
        if not code_list:
            return done(rt, f"{reason_prefix}, but no code list named", True)
        return done(rt, reason_prefix, False)

    if rt is RuleType.CONSTANT:
        if not default:
            return done(rt, f"{reason_prefix}, but no constant value given", True)
        values.pop(SOURCE_FIELD, None)  # CONSTANT must not carry a source field
        return done(rt, reason_prefix, False)

    if rt is RuleType.CONDITIONAL:
        coded = derive_condition(cond_text)
        if coded is None:
            return done(rt, f"{reason_prefix}: condition prose needs coding by hand", True)
        values[CONDITION_CODED] = coded
        values.setdefault(THEN, "SOURCE")
        return done(rt, f"{reason_prefix}: coded as {coded!r}", False)

    if rt is RuleType.LOOP_COUNT:
        if not source:
            return done(rt, f"{reason_prefix}, but no loop named in source", True)
        return done(rt, reason_prefix, False)

    # DIRECT
    if not source:
        return done(rt, f"{reason_prefix}, but no source field", True)
    return done(rt, reason_prefix, False)


# --------------------------------------------------------------------------
# CodeLists import
# --------------------------------------------------------------------------


@dataclass
class ImportResult:
    rows: list[ImportedRow]
    column_mapping: dict[str, str]
    code_lists: dict[str, list[tuple[str, str, str]]] = field(default_factory=dict)
    meta: dict[str, str] = field(default_factory=dict)
    #: headers that carried data but matched no MapCheck column (dropped —
    #: surfaced so nothing is lost silently)
    unmapped_columns: list[str] = field(default_factory=list)

    @property
    def review_rows(self) -> list[ImportedRow]:
        return [r for r in self.rows if r.needs_review]

    @property
    def auto_classified_ratio(self) -> float:
        if not self.rows:
            return 1.0
        return 1.0 - len(self.review_rows) / len(self.rows)


def import_mapping(
    path: str | Path,
    overrides: dict[str, str] | None = None,
    meta: dict[str, str] | None = None,
    code_lists_path: str | Path | None = None,
) -> ImportResult:
    """Read, column-map, and classify a partner mapping document."""
    table = read_table(path)
    mapping = resolve_columns(table.headers, overrides)
    rows = [
        classify_row(mapping, raw, row_id=f"IMP-{i:03d}")
        for i, raw in enumerate(table.rows, start=1)
    ]
    # any header with data that mapped to none of our columns is dropped —
    # surface it so the loss is never silent
    mapped_headers = set(mapping.values())
    unmapped = [
        h for h in table.headers
        if h not in mapped_headers and any(r.get(h, "").strip() for r in table.rows)
    ]
    code_lists: dict[str, list[tuple[str, str, str]]] = {}
    if code_lists_path is not None:
        code_lists = import_code_lists(code_lists_path)
    return ImportResult(
        rows=rows, column_mapping=mapping, code_lists=code_lists,
        meta=meta or {}, unmapped_columns=unmapped,
    )


def import_code_lists(path: str | Path) -> dict[str, list[tuple[str, str, str]]]:
    """Import a lookup table (list-name / source / target / description)."""
    table = read_table(path)
    mapping = {
        "List Name": _best_synonym_free(table.headers, ("list name", "list", "table", "name")),
        "Source Value": _best_synonym_free(table.headers, ("source value", "source", "from", "x12", "code")),
        "Target Value": _best_synonym_free(table.headers, ("target value", "target", "to", "value", "mapped")),
        "Description": _best_synonym_free(table.headers, ("description", "desc", "notes")),
    }
    if not (mapping["List Name"] and mapping["Source Value"] and mapping["Target Value"]):
        raise ImportError_(
            "code list document needs List Name, Source Value, and Target Value columns"
        )
    out: dict[str, list[tuple[str, str, str]]] = {}
    for raw in table.rows:
        name = raw.get(mapping["List Name"], "").strip()
        src = raw.get(mapping["Source Value"], "").strip()
        tgt = raw.get(mapping["Target Value"], "").strip()
        desc = raw.get(mapping["Description"], "").strip() if mapping["Description"] else ""
        if name and src and tgt:
            out.setdefault(name, []).append((src, tgt, desc))
    return out


def _best_synonym_free(headers: list[str], fragments: tuple[str, ...]) -> str | None:
    norm_headers = {_norm(h): h for h in headers}
    for frag in fragments:
        if frag in norm_headers:
            return norm_headers[frag]
    for norm_h, raw_h in norm_headers.items():
        if any(frag in norm_h for frag in fragments):
            return raw_h
    return None


# --------------------------------------------------------------------------
# Stage 3 (cont.): write the native workbook + worklist
# --------------------------------------------------------------------------

_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # soft yellow
_MAPPING_HEADERS = [header for header, _ in MAPPING_COLUMNS]


def write_workbook(result: ImportResult, path: str | Path) -> Path:
    """Write the imported spec to a native MapCheck workbook.

    Rows that need review keep everything the importer could determine, get
    a soft-yellow fill and a Notes reason, and (when the type itself is
    unknown) a blank Rule Type — the workbook is a *draft* that will not
    ``load_spec`` until those rows are resolved, which is the honest state.
    """
    path = Path(path)
    direction = result.meta.get("Direction", "inbound")
    wb = build_template(direction=direction if direction in ("inbound", "outbound") else "inbound")
    ws = wb["Mapping"]
    col_index = {header: i + 1 for i, header in enumerate(_MAPPING_HEADERS)}

    for r, row in enumerate(result.rows, start=2):
        cells: dict[str, str] = {"Row ID": row.row_id, **row.values}
        note = row.reason
        if row.needs_review:
            note = f"NEEDS REVIEW — {note}"
        cells["Notes"] = note
        for header, value in cells.items():
            if header in col_index and value:
                ws.cell(row=r, column=col_index[header], value=value)
        if row.needs_review:
            for c in range(1, len(_MAPPING_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = _REVIEW_FILL

    # CodeLists
    ws_codes = wb["CodeLists"]
    cr = 2
    for name, entries in result.code_lists.items():
        for src, tgt, desc in entries:
            ws_codes.cell(row=cr, column=1, value=name)
            ws_codes.cell(row=cr, column=2, value=src)
            ws_codes.cell(row=cr, column=3, value=tgt)
            if desc:
                ws_codes.cell(row=cr, column=4, value=desc)
            cr += 1

    # Meta
    ws_meta = wb["Meta"]
    existing = {ws_meta.cell(row=i, column=1).value: i for i in range(2, ws_meta.max_row + 1)}
    next_row = ws_meta.max_row + 1
    for key, value in result.meta.items():
        if key in existing:
            ws_meta.cell(row=existing[key], column=2, value=value)
        else:
            ws_meta.cell(row=next_row, column=1, value=key)
            ws_meta.cell(row=next_row, column=2, value=value)
            next_row += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def format_worklist(result: ImportResult) -> str:
    """Human-readable worklist of every row that needs review."""
    lines = [
        "MapCheck spec import — review worklist",
        "",
        f"Rows imported:     {len(result.rows)}",
        f"Auto-classified:   {len(result.rows) - len(result.review_rows)} "
        f"({result.auto_classified_ratio:.0%})",
        f"Needs review:      {len(result.review_rows)}",
        "",
    ]
    mapped = ", ".join(f"{k} <- {v!r}" for k, v in result.column_mapping.items())
    lines.append(f"Column mapping: {mapped}")
    if result.unmapped_columns:
        cols = ", ".join(repr(c) for c in result.unmapped_columns)
        lines.append(
            f"IGNORED columns (had data, matched no MapCheck column): {cols}"
        )
        lines.append(
            "  → if any of these carry mapping meaning, re-run with "
            "--map \"<Our Column>=<their header>\"."
        )
    lines.append("")
    if not result.review_rows:
        lines.append("No rows need review — the imported spec is ready to load.")
        return "\n".join(lines) + "\n"
    lines.append("Rows needing a human decision:")
    for row in result.review_rows:
        target = row.values.get(TARGET_FIELD, "(no target)")
        lines.append(f"  {row.row_id}  {target:<28} {row.reason}")
    lines.append("")
    lines.append(
        "Open the workbook, resolve the yellow-filled rows (fill in Rule Type "
        "and any missing cells), then validate as usual."
    )
    return "\n".join(lines) + "\n"


def write_worklist(result: ImportResult, path: str | Path) -> Path:
    path = Path(path)
    path.write_text(format_worklist(result), encoding="utf-8")
    return path
