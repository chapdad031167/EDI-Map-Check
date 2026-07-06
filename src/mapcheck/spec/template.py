"""Generator for the EDI MapCheck Excel spec template.

The template is produced by code (not a checked-in binary) so the format
is reproducible and diffs live in one place. Workbook layout:

* **Mapping** — one row per mapping rule (columns below).
* **CodeLists** — long-format lookup tables: List Name / Source Value /
  Target Value / Description.
* **Meta** — key/value pairs (Transaction Set, X12 Version, Spec Name, ...).
* **Instructions** — the addressing and condition mini-grammar, for humans.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from mapcheck.spec.model import DataType, RuleType

#: Mapping sheet columns, in order: (header, width).
MAPPING_COLUMNS: tuple[tuple[str, int], ...] = (
    ("Row ID", 10),
    ("Source Field", 14),
    ("Loop Context", 14),
    ("Target Field", 24),
    ("Rule Type", 14),
    ("Condition (Text)", 34),
    ("Condition (Coded)", 26),
    ("Then", 16),
    ("Else", 12),
    ("Default Value", 14),
    ("Code List Ref", 14),
    ("Data Type", 11),
    ("Format", 14),
    ("Notes", 40),
)

CODELIST_COLUMNS: tuple[tuple[str, int], ...] = (
    ("List Name", 18),
    ("Source Value", 14),
    ("Target Value", 18),
    ("Description", 44),
)

META_KEYS: tuple[str, ...] = (
    "Transaction Set",
    "X12 Version",
    "Spec Name",
    "Author",
    "Date",
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_INSTRUCTIONS = """EDI MapCheck — spec template quick reference

SOURCE FIELD    Plain segment+element notation: BEG03, N104, PO102.
LOOP CONTEXT    Which occurrence the source field comes from:
                  (blank)      header/summary area, first occurrence
                  PO1          each occurrence of the PO1 loop (per-line
                               rule; target must use lines[] notation)
                  N1[ST]       the N1 loop whose N101 = ST
                  REF[DP]      the REF segment whose REF01 = DP
                  LIN>QTY[QA]  per-line on the LIN loop, but the source
                               field reads the QTY whose QTY01 = QA (for
                               repeating qualified segments in a loop)
                The [qualifier] always matches element 01 of that segment.
TARGET FIELD    Dot path into the output: order.po_number, ship_to.name,
                lines[].qty, summary.line_count.
RULE TYPE       DIRECT       source value lands in target (after type/format
                             conversion)
                CONDITIONAL  Condition (Coded) decides between Then and Else
                CODE_LIST    source value translated via the named code list
                CONSTANT     hardcoded value from Default Value (no source)
                LOOP_COUNT   target = occurrence count of the loop named in
                             Source Field (e.g. PO1)
CONDITION       (Text) is documentation for humans. (Coded) is what the
                engine runs:  N101 = 'ST'
                              BEG02 != 'SA'
                              PO103 IN ('EA', 'CA')
                              EXISTS(REF02)
                Join with AND. Values always in single quotes.
THEN / ELSE     SOURCE (map the source value), SKIP (field must be absent),
                BLANK (present but empty), or a 'quoted literal'.
                Else defaults to SKIP when left empty.
DEFAULT VALUE   CONSTANT rules: the hardcoded value. Other rules: fallback
                used when the source element is empty/absent.
DATA TYPE       string | integer | decimal | date | time
FORMAT          date:    strftime pattern of the OUTPUT value, e.g. %Y-%m-%d
                         (X12 source dates are read as CCYYMMDD/YYMMDD)
                decimal: places:2 (output has 2 decimal places) and/or
                         implied:2 (source is X12 N2 — implied 2 decimals)
                string:  len:1..30 (allowed output length range)
                Combine with ; e.g.  implied:2;places:2
CODE LISTS      Long format on the CodeLists sheet; Code List Ref names the
                List Name value.
"""


def _write_header(ws: Worksheet, columns: tuple[tuple[str, int], ...]) -> None:
    for idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def build_template() -> Workbook:
    """Build the blank spec template as an in-memory workbook."""
    wb = Workbook()

    ws_map = wb.active
    ws_map.title = "Mapping"
    _write_header(ws_map, MAPPING_COLUMNS)

    rule_types = ",".join(rt.value for rt in RuleType)
    dv_rule = DataValidation(
        type="list", formula1=f'"{rule_types}"', allow_blank=True, showDropDown=False
    )
    ws_map.add_data_validation(dv_rule)
    dv_rule.add("E2:E500")

    data_types = ",".join(dt.value for dt in DataType)
    dv_type = DataValidation(
        type="list", formula1=f'"{data_types}"', allow_blank=True, showDropDown=False
    )
    ws_map.add_data_validation(dv_type)
    dv_type.add("L2:L500")

    ws_codes = wb.create_sheet("CodeLists")
    _write_header(ws_codes, CODELIST_COLUMNS)

    ws_meta = wb.create_sheet("Meta")
    _write_header(ws_meta, (("Key", 20), ("Value", 40)))
    for i, key in enumerate(META_KEYS, start=2):
        ws_meta.cell(row=i, column=1, value=key)
    ws_meta.cell(row=2, column=2, value="850")
    ws_meta.cell(row=3, column=2, value="004010")

    ws_help = wb.create_sheet("Instructions")
    ws_help.column_dimensions["A"].width = 100
    for i, line in enumerate(_INSTRUCTIONS.splitlines(), start=1):
        cell = ws_help.cell(row=i, column=1, value=line)
        cell.font = Font(name="Consolas", size=10, bold=(i == 1))

    return wb


def create_template(path: str | Path) -> Path:
    """Write the blank spec template to ``path`` and return it."""
    path = Path(path)
    build_template().save(path)
    return path
