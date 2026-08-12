"""draft-spec: definition-driven draft mapping specs (Design 012).

Three layers, deterministic by construction:

1. **Structural walk** — two pure functions enumerate the sides:
   :func:`walk_source_elements` (X12 elements with loop context, from the
   transaction definition's element dictionary) and
   :func:`walk_target_paths` (concrete target paths from the output
   definition's ``draft`` block). Ordering follows definition file order.
2. **Crosswalk** — maintainer-authored YAML of known pairings
   (``mapcheck/crosswalks/<transaction>_<target>.yaml``), validated at
   load: a malformed entry is a :class:`CrosswalkError` naming the file
   and entry index, never a silent skip. Repeated files merge by target
   path — a later file's rule replaces an earlier one.
3. **TODO emission** — every required target path with no crosswalk hit
   becomes a ``TODO`` row (the loader reports those NOT TESTED with a
   warning); source elements no rule references land on an Unmapped
   Source sheet for human triage.

The same definitions plus the same crosswalks always produce identical
sheet content; golden tests compare parsed content, not raw bytes (xlsx
embeds save timestamps).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from dataclasses import replace as dataclass_replace
from pathlib import Path

import yaml
from openpyxl.styles import PatternFill

from mapcheck import __version__
from mapcheck.guides.profile import GuideProfile
from mapcheck.output.idoc import DraftTarget, OutputFormatDef
from mapcheck.spec.conditions import ConditionSyntaxError, parse_condition, parse_outcome
from mapcheck.spec.formats import FormatSpecError, parse_format
from mapcheck.spec.model import LoopContext, RuleType, split_source_field
from mapcheck.spec.template import build_template
from mapcheck.transactions.schema import TransactionDefinition

#: Display form of the guide's normalized usage values.
_USAGE_DISPLAY = {"must_use": "Must use", "used": "Used", "not_used": "Not used"}

#: Directory of bundled starter crosswalks, one file per
#: (transaction, target) pair: ``850_orders05.yaml``.
CROSSWALKS_DIR = Path(__file__).parent.parent / "crosswalks"

#: Soft amber fill for TODO rows (same convention as import-spec's
#: NEEDS-REVIEW rows).
_TODO_FILL = PatternFill("solid", fgColor="FFF2CC")

#: Crosswalk rule fields accepted by the schema (typo protection).
_RULE_FIELDS = {
    "source", "target", "rule", "context", "code_list",
    "condition", "then", "else", "default", "format", "note",
}


class CrosswalkError(Exception):
    """A crosswalk file failed validation. The message names the file and
    entry index of every problem."""


class DraftSpecError(Exception):
    """draft-spec cannot run for this (transaction, target) pair."""


# --------------------------------------------------------------------------
# Crosswalk schema and loader
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class CrosswalkRule:
    """One validated crosswalk pairing."""

    target: str
    rule: RuleType
    source: str | None = None
    context: str | None = None
    code_list: str | None = None
    condition: str | None = None
    then: str | None = None
    else_: str | None = None
    default: str | None = None
    format: str | None = None
    note: str = ""
    origin: str = ""  # "file[index]", for later error context


@dataclass(frozen=True)
class CodeListEntry:
    source: str
    target: str
    description: str = ""


@dataclass
class Crosswalk:
    """Merged crosswalk content. ``rules`` is keyed by target path — the
    merge key: with repeated files, a later rule replaces an earlier one
    for the same target path; code lists replace whole lists by name."""

    rules: dict[str, CrosswalkRule] = field(default_factory=dict)
    code_lists: dict[str, tuple[CodeListEntry, ...]] = field(default_factory=dict)
    files: tuple[str, ...] = ()


def _validate_rule(node: object, where: str, errors: list[str]) -> CrosswalkRule | None:
    if not isinstance(node, dict):
        errors.append(f"{where}: expected a mapping")
        return None
    unknown = set(node) - _RULE_FIELDS
    if unknown:
        errors.append(
            f"{where}: unknown field(s) {', '.join(sorted(map(str, unknown)))} "
            f"(expected only: {', '.join(sorted(_RULE_FIELDS))})"
        )
        return None

    def text(key: str) -> str | None:
        value = node.get(key)
        if value is None:
            return None
        return str(value)

    target = text("target")
    if not target:
        errors.append(f"{where}: 'target' is required")
        return None

    rule_raw = text("rule")
    if not rule_raw:
        errors.append(f"{where}: 'rule' is required")
        return None
    try:
        rule = RuleType(rule_raw.upper())
    except ValueError:
        errors.append(
            f"{where}: unknown rule {rule_raw!r} (expected one of "
            f"{', '.join(rt.value for rt in RuleType if rt is not RuleType.TODO)})"
        )
        return None
    if rule is RuleType.TODO:
        errors.append(
            f"{where}: rule TODO is not a crosswalk pairing — leave the target "
            "out of the crosswalk and draft-spec emits the TODO row itself"
        )
        return None

    source = text("source")
    context = text("context")
    condition = text("condition")
    then = text("then")
    else_ = text("else")

    if rule in (RuleType.DIRECT, RuleType.CODE_LIST, RuleType.LOOP_COUNT) and not source:
        errors.append(f"{where}: rule {rule.value} requires 'source'")
        return None
    if rule is RuleType.CODE_LIST and not text("code_list"):
        errors.append(f"{where}: rule CODE_LIST requires 'code_list'")
        return None
    if rule is RuleType.CONSTANT and text("default") is None:
        errors.append(f"{where}: rule CONSTANT requires 'default'")
        return None
    if rule is RuleType.CONDITIONAL:
        if not condition or not then:
            errors.append(f"{where}: rule CONDITIONAL requires 'condition' and 'then'")
            return None

    if source and rule is not RuleType.LOOP_COUNT:
        try:
            split_source_field(source)
        except ValueError as exc:
            errors.append(f"{where}: source: {exc}")
            return None
    if context:
        try:
            LoopContext.parse(context)
        except ValueError as exc:
            errors.append(f"{where}: context: {exc}")
            return None
    if condition:
        try:
            parse_condition(condition)
        except ConditionSyntaxError as exc:
            errors.append(f"{where}: condition: {exc}")
            return None
    for outcome_name, outcome_raw in (("then", then), ("else", else_)):
        if outcome_raw:
            try:
                parse_outcome(outcome_raw)
            except ConditionSyntaxError as exc:
                errors.append(f"{where}: {outcome_name}: {exc}")
                return None
    if fmt := text("format"):
        try:
            parse_format(fmt)
        except FormatSpecError as exc:
            errors.append(f"{where}: format: {exc}")
            return None

    return CrosswalkRule(
        target=target,
        rule=rule,
        source=source.upper() if source else None,
        context=context,
        code_list=text("code_list"),
        condition=condition,
        then=then,
        else_=else_,
        default=text("default"),
        format=text("format"),
        note=text("note") or "",
        origin=where,
    )


def _validate_code_lists(
    node: object, source_name: str, errors: list[str]
) -> dict[str, tuple[CodeListEntry, ...]]:
    lists: dict[str, tuple[CodeListEntry, ...]] = {}
    if node is None:
        return lists
    if not isinstance(node, dict):
        errors.append(f"{source_name}: code_lists: expected a mapping of list name -> entries")
        return lists
    for name, entries in node.items():
        where = f"{source_name}: code_lists.{name}"
        if not isinstance(entries, list):
            errors.append(f"{where}: expected a list of entries")
            continue
        parsed: list[CodeListEntry] = []
        for index, entry in enumerate(entries):
            if (
                not isinstance(entry, dict)
                or entry.get("source") is None
                or entry.get("target") is None
            ):
                errors.append(
                    f"{where}[{index}]: expected a mapping with 'source' and 'target'"
                )
                continue
            parsed.append(
                CodeListEntry(
                    source=str(entry["source"]),
                    target=str(entry["target"]),
                    description=str(entry.get("description", "") or ""),
                )
            )
        lists[str(name)] = tuple(parsed)
    return lists


def load_crosswalk(path: str | Path) -> Crosswalk:
    """Load and validate one crosswalk file.

    Raises :class:`CrosswalkError` listing every problem, each naming the
    file and entry index.
    """
    path = Path(path)
    if not path.exists():
        raise CrosswalkError(f"crosswalk file not found: {path}")
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
    except yaml.YAMLError as exc:
        raise CrosswalkError(f"{path}: invalid YAML: {exc}") from exc

    if isinstance(data, list):  # bare list = rules-only shorthand
        rules_node, code_lists_node = data, None
    elif isinstance(data, dict):
        rules_node = data.get("rules", [])
        code_lists_node = data.get("code_lists")
        unknown = set(data) - {"rules", "code_lists"}
        if unknown:
            raise CrosswalkError(
                f"{path}: unknown top-level key(s) {', '.join(sorted(map(str, unknown)))} "
                "(expected 'rules' and optionally 'code_lists')"
            )
    else:
        raise CrosswalkError(
            f"{path}: expected a list of rules, or a mapping with 'rules'"
        )
    if not isinstance(rules_node, list):
        raise CrosswalkError(f"{path}: 'rules' must be a list")

    errors: list[str] = []
    rules: dict[str, CrosswalkRule] = {}
    for index, node in enumerate(rules_node):
        parsed = _validate_rule(node, f"{path}: entry [{index}]", errors)
        if parsed is None:
            continue
        if parsed.target in rules:
            errors.append(
                f"{path}: entry [{index}]: duplicate target {parsed.target!r} within "
                "one file — a target path may appear once per crosswalk (use a "
                "second --crosswalk file to override)"
            )
            continue
        rules[parsed.target] = parsed

    code_lists = _validate_code_lists(code_lists_node, str(path), errors)

    # CODE_LIST rules must reference a list this file (or a later merge)
    # can satisfy; per-file check happens after merge in load_crosswalks.
    if errors:
        raise CrosswalkError(
            f"{len(errors)} problem(s) in crosswalk {path}:\n  - " + "\n  - ".join(errors)
        )
    return Crosswalk(rules=rules, code_lists=code_lists, files=(str(path),))


def load_crosswalks(paths: list[str | Path]) -> Crosswalk:
    """Load and merge crosswalk files in order.

    Merge key is the target path: a later file's rule replaces an earlier
    file's rule for the same target. Code lists replace whole lists by
    name. After the merge, every CODE_LIST rule must resolve its list.
    """
    merged = Crosswalk()
    files: list[str] = []
    for path in paths:
        one = load_crosswalk(path)
        merged.rules.update(one.rules)
        merged.code_lists.update(one.code_lists)
        files.extend(one.files)
    merged.files = tuple(files)

    missing = [
        f"{rule.origin}: CODE_LIST references unknown code list {rule.code_list!r}"
        for rule in merged.rules.values()
        if rule.rule is RuleType.CODE_LIST and rule.code_list not in merged.code_lists
    ]
    if missing:
        raise CrosswalkError(
            f"{len(missing)} problem(s) after crosswalk merge:\n  - " + "\n  - ".join(missing)
        )
    return merged


def bundled_crosswalks(transaction: str, target: str) -> list[Path]:
    """The bundled starter crosswalk for a pair, when one ships."""
    candidate = CROSSWALKS_DIR / f"{transaction}_{target}.yaml"
    return [candidate] if candidate.exists() else []


# --------------------------------------------------------------------------
# Structural walk: two pure, deterministic functions
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class SourceElement:
    """One enumerable X12 source element."""

    field: str  # e.g. "BEG03"
    context: str  # "" for area segments, the loop id ("PO1") for loop members
    name: str  # human name from the element dictionary
    partner_usage: str = ""  # "Must use"/"Used" from a guide, "" without one


def walk_source_elements(definition: TransactionDefinition) -> tuple[SourceElement, ...]:
    """Enumerate source elements with loop context, in definition order.

    Areas in order; within an area, flat segments then loops (depth-first),
    the loop trigger before its member segments. Each segment enumerates
    once, at its first declared location; segments without an element
    dictionary entry don't enumerate at all.
    """
    elements: list[SourceElement] = []
    seen: set[str] = set()

    def emit(seg_id: str, context: str) -> None:
        if seg_id in seen:
            return
        names = definition.elements.get(seg_id)
        if not names:
            return
        seen.add(seg_id)
        for index, name in enumerate(names, start=1):
            elements.append(
                SourceElement(field=f"{seg_id}{index:02d}", context=context, name=name)
            )

    def walk_loop(loop) -> None:
        emit(loop.trigger, loop.id)
        for seg_id in loop.segments:
            emit(seg_id, loop.id)
        for child in loop.loops:
            walk_loop(child)

    for area in definition.areas:
        for seg_id in area.segments:
            emit(seg_id, "")
        for loop in area.loops:
            walk_loop(loop)
    return tuple(elements)


def walk_target_paths(output_def: OutputFormatDef) -> tuple[DraftTarget, ...]:
    """Enumerate concrete target paths, in output-definition order."""
    if not output_def.draft_targets:
        raise DraftSpecError(
            f"output definition {output_def.format!r} declares no draft targets — "
            "add a 'draft:' block to its definition file to enable draft-spec"
        )
    return output_def.draft_targets


# --------------------------------------------------------------------------
# Assembly
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class DraftRow:
    """One Mapping-sheet row of the draft."""

    row_id: str
    target: str
    rule: str
    source: str = ""
    context: str = ""
    condition: str = ""
    then: str = ""
    else_: str = ""
    default: str = ""
    code_list: str = ""
    data_type: str = ""
    format: str = ""
    notes: str = ""
    filled: bool = False
    required: bool = True


@dataclass(frozen=True)
class DraftResult:
    """Everything one draft-spec run produced."""

    transaction: str
    transaction_version: str
    target_format: str
    rows: tuple[DraftRow, ...]
    unmapped_source: tuple[SourceElement, ...]
    code_lists: dict[str, tuple[CodeListEntry, ...]]
    crosswalk_files: tuple[str, ...]
    filled_required: int
    total_required: int
    #: Guide flavoring (Design 014): set when a profile shaped this draft.
    guided: bool = False
    guide_partner: str = ""
    guide_source: str = ""
    #: Decisions the guide surfaced that no rule can settle mechanically
    #: (partner codes with no crosswalk translation, mapped-but-not-used
    #: elements) — written to the Guide Review sheet, never dropped.
    guide_review: tuple[str, ...] = ()

    @property
    def prefill(self) -> float:
        if self.total_required == 0:
            return 1.0
        return self.filled_required / self.total_required

    @property
    def todo_count(self) -> int:
        return sum(1 for row in self.rows if not row.filled)


def assemble_draft(
    definition: TransactionDefinition,
    output_def: OutputFormatDef,
    crosswalk: Crosswalk,
    include_optional_todos: bool = False,
    guide: GuideProfile | None = None,
) -> DraftResult:
    """Join the two walks through the crosswalk.

    A required target with a crosswalk hit becomes a filled row; a required
    target with no hit becomes a TODO row. Optional targets appear when
    crosswalked (or, with ``include_optional_todos``, as TODO rows too).
    Source elements never referenced by a placed rule go on the Unmapped
    Source list. prefill = filled required rows / total required rows.

    With a ``guide`` profile (Design 014) the draft is partner-flavored:
    guide notes flow into filled rows' Notes, ``not_used`` elements leave
    Unmapped Source (the partner said no — they are not triage), remaining
    unmapped elements show the partner's usage, and CODE_LIST code lists
    are filtered to the partner's code subset. Partner codes with no
    crosswalk translation, and mapped elements the guide marks not-used,
    land in ``guide_review`` — surfaced, never resolved by guesswork.
    """
    rows: list[DraftRow] = []
    referenced: set[tuple[str, int]] = set()
    filled_required = 0
    total_required = 0
    guide_elements = guide.element_usage() if guide is not None else {}
    guide_review: list[str] = []

    def mark(field_ref: str | None) -> None:
        if not field_ref:
            return
        try:
            seg_id, element = split_source_field(field_ref)
        except ValueError:
            return
        referenced.add((seg_id, element))

    def guide_element(field_ref: str | None):
        if not field_ref:
            return None
        try:
            key = split_source_field(field_ref)
        except ValueError:
            return None
        return guide_elements.get(key)

    for target in walk_target_paths(output_def):
        if target.required:
            total_required += 1
        hit = crosswalk.rules.get(target.path)
        if hit is not None:
            if target.required:
                filled_required += 1
            notes = hit.note or target.name
            source_el = guide_element(hit.source)
            if source_el is not None:
                if source_el.notes:
                    suffix = f"Partner note: {'; '.join(source_el.notes)}"
                    notes = f"{notes} — {suffix}" if notes else suffix
                if source_el.usage == "not_used":
                    guide_review.append(
                        f"{hit.source} is mapped to {target.path} but the guide "
                        "marks it Not used — confirm the mapping with the partner"
                    )
            rows.append(
                DraftRow(
                    row_id=f"D-{len(rows) + 1:03d}",
                    target=target.path,
                    rule=hit.rule.value,
                    source=hit.source or "",
                    context=hit.context or "",
                    condition=hit.condition or "",
                    then=hit.then or "",
                    else_=hit.else_ or "",
                    default=hit.default or "",
                    code_list=hit.code_list or "",
                    data_type=target.data_type or "",
                    format=hit.format if hit.format is not None else (target.format or ""),
                    notes=notes,
                    filled=True,
                    required=target.required,
                )
            )
            if hit.rule is not RuleType.LOOP_COUNT:
                mark(hit.source)
            if hit.condition:
                for predicate in parse_condition(hit.condition).predicates:
                    mark(predicate.field)
            if hit.context:
                loop_context = LoopContext.parse(hit.context)
                if loop_context.qualifier is not None:
                    referenced.add((loop_context.segment_id, 1))
        elif target.required or include_optional_todos:
            rows.append(
                DraftRow(
                    row_id=f"D-{len(rows) + 1:03d}",
                    target=target.path,
                    rule=RuleType.TODO.value,
                    data_type=target.data_type or "",
                    notes=(
                        f"TODO: {target.name}" if target.name else "TODO: decide the mapping"
                    ),
                    filled=False,
                    required=target.required,
                )
            )

    code_lists = dict(crosswalk.code_lists)
    if guide is not None:
        # Partner code subsets: rows in draft order, codes in guide order.
        subsets: dict[str, dict[str, str]] = {}
        for row in rows:
            if row.rule != RuleType.CODE_LIST.value or not row.code_list:
                continue
            el = guide_element(row.source)
            if el is None or not el.codes:
                continue
            bucket = subsets.setdefault(row.code_list, {})
            for code in el.codes:
                bucket.setdefault(code.code, code.name)
        for list_name, partner_codes in subsets.items():
            entries = code_lists.get(list_name, ())
            known = {entry.source for entry in entries}
            code_lists[list_name] = tuple(
                entry for entry in entries if entry.source in partner_codes
            )
            for code, code_name in partner_codes.items():
                if code not in known:
                    label = f"{code} ({code_name})" if code_name else code
                    guide_review.append(
                        f"code list {list_name}: partner code {label} has no "
                        "crosswalk translation — decide its target value"
                    )

    unmapped: list[SourceElement] = []
    for element in walk_source_elements(definition):
        key = split_source_field(element.field)
        if key in referenced:
            continue
        el = guide_elements.get(key)
        if el is not None and el.usage == "not_used":
            continue  # the partner said no — not triage
        usage = _USAGE_DISPLAY.get(el.usage, "") if el is not None else ""
        unmapped.append(
            element if not usage else dataclass_replace(element, partner_usage=usage)
        )

    return DraftResult(
        transaction=definition.set_code,
        transaction_version=definition.version or "",
        target_format=output_def.format,
        rows=tuple(rows),
        unmapped_source=tuple(unmapped),
        code_lists=code_lists,
        crosswalk_files=crosswalk.files,
        filled_required=filled_required,
        total_required=total_required,
        guided=guide is not None,
        guide_partner=guide.partner if guide is not None else "",
        guide_source=guide.source if guide is not None else "",
        guide_review=tuple(guide_review),
    )


# --------------------------------------------------------------------------
# Emission through the existing template writer
# --------------------------------------------------------------------------


def write_draft(result: DraftResult, path: str | Path) -> Path:
    """Write the draft workbook. Content is deterministic for identical
    inputs; only xlsx-embedded save timestamps vary between runs."""
    path = Path(path)
    wb = build_template(direction="inbound")

    ws = wb["Mapping"]
    for r, row in enumerate(result.rows, start=2):
        values = (
            row.row_id, row.source, row.context, row.target, row.rule,
            "", row.condition, row.then, row.else_, row.default,
            row.code_list, row.data_type, row.format, row.notes,
        )
        for c, value in enumerate(values, start=1):
            if value != "":
                ws.cell(row=r, column=c, value=value)
        if not row.filled:
            for c in range(1, len(values) + 1):
                ws.cell(row=r, column=c).fill = _TODO_FILL

    ws_codes = wb["CodeLists"]
    r = 2
    for name, entries in result.code_lists.items():
        for entry in entries:
            ws_codes.cell(row=r, column=1, value=name)
            ws_codes.cell(row=r, column=2, value=entry.source)
            ws_codes.cell(row=r, column=3, value=entry.target)
            if entry.description:
                ws_codes.cell(row=r, column=4, value=entry.description)
            r += 1

    ws_meta = wb["Meta"]
    meta_values = {
        "Transaction Set": result.transaction,
        "X12 Version": result.transaction_version,
        "Direction": "inbound",
        "Spec Name": f"Draft {result.transaction} to {result.target_format} (draft-spec)",
        "Author": "mapcheck draft-spec",
    }
    written: set[str] = set()
    for r in range(2, ws_meta.max_row + 1):
        key = ws_meta.cell(row=r, column=1).value
        if key in meta_values:
            ws_meta.cell(row=r, column=2, value=meta_values[key])
            written.add(key)
    extra = {
        "Target Format": result.target_format,
        "Crosswalks": ", ".join(Path(f).name for f in result.crosswalk_files) or "(none)",
        "Tool Version": __version__,
        "Prefill": f"{result.prefill:.2f}",
    }
    if result.guided:
        guide_label = result.guide_source or "(unnamed guide)"
        if result.guide_partner:
            guide_label = f"{result.guide_partner} — {guide_label}"
        extra["Guide"] = guide_label
    next_row = ws_meta.max_row + 1
    for key, value in extra.items():
        ws_meta.cell(row=next_row, column=1, value=key)
        ws_meta.cell(row=next_row, column=2, value=value)
        next_row += 1

    unmapped_headers = [("Source Field", 14), ("Loop Context", 14), ("Element Name", 44)]
    if result.guided:
        unmapped_headers.append(("Partner Usage", 14))
    ws_unmapped = wb.create_sheet("Unmapped Source")
    for c, (header, width) in enumerate(unmapped_headers, start=1):
        ws_unmapped.cell(row=1, column=c, value=header)
        ws_unmapped.column_dimensions[chr(ord("A") + c - 1)].width = width
    ws_unmapped.freeze_panes = "A2"
    for r, element in enumerate(result.unmapped_source, start=2):
        ws_unmapped.cell(row=r, column=1, value=element.field)
        if element.context:
            ws_unmapped.cell(row=r, column=2, value=element.context)
        ws_unmapped.cell(row=r, column=3, value=element.name)
        if result.guided and element.partner_usage:
            ws_unmapped.cell(row=r, column=4, value=element.partner_usage)

    if result.guide_review:
        ws_review = wb.create_sheet("Guide Review")
        ws_review.cell(row=1, column=1, value="Decision Needed")
        ws_review.column_dimensions["A"].width = 100
        ws_review.freeze_panes = "A2"
        for r, note in enumerate(result.guide_review, start=2):
            ws_review.cell(row=r, column=1, value=note)

    wb.save(path)
    return path


def generate_draft(
    definition: TransactionDefinition,
    output_def: OutputFormatDef,
    crosswalk_paths: list[str | Path],
    output_path: str | Path,
    include_optional_todos: bool = False,
    guide: GuideProfile | None = None,
) -> DraftResult:
    """Load crosswalks, assemble, and write — the CLI/UI entry point."""
    crosswalk = load_crosswalks(crosswalk_paths) if crosswalk_paths else Crosswalk()
    result = assemble_draft(
        definition, output_def, crosswalk,
        include_optional_todos=include_optional_todos, guide=guide,
    )
    write_draft(result, output_path)
    return result
