#!/usr/bin/env python3
"""Regenerate the synthetic example artifacts under examples/.

Everything here is synthetic, built only from the public X12 850 structure
(segment/element definitions from X12.org). No proprietary data.

Usage:
    python scripts/generate_examples.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from mapcheck.spec.template import build_template  # noqa: E402

EXAMPLES = REPO_ROOT / "examples"


# --------------------------------------------------------------------------
# Reference mapping spec: synthetic retail PO, 850 v004010
# --------------------------------------------------------------------------

# Row ID, Source Field, Loop Context, Target Field, Rule Type,
# Condition (Text), Condition (Coded), Then, Else, Default Value,
# Code List Ref, Data Type, Format, Notes
SPEC_ROWS: list[tuple[str, ...]] = [
    ("M-001", "BEG01", "", "order.tx_purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", "Transaction set purpose code."),
    ("M-002", "BEG02", "", "order.po_type", "CODE_LIST", "", "", "", "",
     "", "PO_TYPE", "string", "", "PO type code."),
    ("M-003", "BEG02", "", "order.drop_ship_flag", "CONDITIONAL",
     "If the PO type is drop ship, set the flag; otherwise N.",
     "BEG02 = 'DS'", "'Y'", "'N'", "", "", "string", "len:1..1", ""),
    ("M-004", "BEG03", "", "order.po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", "Customer PO number."),
    ("M-005", "BEG05", "", "order.po_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "PO date, CCYYMMDD in source."),
    ("M-006", "CUR02", "", "order.currency", "CONDITIONAL",
     "Map the currency only when a buying-party currency code is sent.",
     "CUR01 = 'BY'", "SOURCE", "SKIP", "", "", "string", "len:3..3", ""),
    ("M-007", "REF02", "REF[DP]", "order.department", "DIRECT", "", "", "", "",
     "", "", "string", "", "Department number."),
    ("M-008", "REF02", "REF[IA]", "order.vendor_number", "DIRECT", "", "", "", "",
     "", "", "string", "", "Internal vendor number."),
    ("M-009", "REF02", "REF[PD]", "order.promo_code", "CONDITIONAL",
     "Map the promotion/deal number only when the trading partner sends one.",
     "EXISTS(REF02)", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("M-010", "DTM02", "DTM[002]", "order.requested_delivery_date", "DIRECT",
     "", "", "", "", "", "", "date", "%Y-%m-%d", "DTM 002 = requested delivery."),
    ("M-011", "DTM02", "DTM[001]", "order.cancel_after_date", "DIRECT",
     "", "", "", "", "", "", "date", "%Y-%m-%d",
     "DTM 001 = cancel after. Optional for this partner."),
    ("M-012", "", "", "order.record_type", "CONSTANT", "", "", "", "",
     "PO_INBOUND", "", "string", "", "Hardcoded record type for downstream routing."),
    ("M-013", "SAC05", "SAC[A]", "order.allowance_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "implied:2;places:2",
     "SAC05 is X12 N2: implied two decimals. 2500 in source = 25.00 out."),
    ("M-014", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("M-015", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the store number only when the ID qualifier is 92 (assigned by buyer).",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("M-016", "N301", "N1[ST]", "ship_to.address1", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("M-017", "N401", "N1[ST]", "ship_to.city", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("M-018", "N402", "N1[ST]", "ship_to.state", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", ""),
    ("M-019", "N403", "N1[ST]", "ship_to.zip", "DIRECT", "", "", "", "",
     "", "", "string", "len:5..10", ""),
    ("M-020", "N102", "N1[BT]", "bill_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("M-021", "N104", "N1[BT]", "bill_to.id", "CONDITIONAL",
     "Map the bill-to ID only when the ID qualifier is 92 (assigned by buyer).",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("M-022", "PO101", "PO1", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Line sequence number."),
    ("M-023", "PO102", "PO1", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Quantity ordered."),
    ("M-024", "PO103", "PO1", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", "Unit of measure, translated to internal values."),
    ("M-025", "PO104", "PO1", "lines[].unit_price", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "PO104 is X12 R: explicit decimal."),
    ("M-026", "PO107", "PO1", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "PO106 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("M-027", "PID05", "PO1", "lines[].description", "CONDITIONAL",
     "Map the free-form description when the PID is free-form (PID01 = F).",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..80", ""),
    ("M-028", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", "CTT line count as transmitted."),
    ("M-029", "PO1", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "",
     "Actual PO1 loop occurrences must also equal the output line count."),
    ("M-030", "AMT02", "AMT[TT]", "summary.total_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "Total transaction amount."),
    ("M-031", "SAC02", "SAC[A]", "order.allowance_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:4..4", "Allowance/charge code (SAC01 = A)."),
    ("M-032", "BEG04", "", "order.release_number", "DIRECT", "", "", "", "",
     "NONE", "", "string", "", "Release number; this partner never sends one, default NONE."),
]

# List Name, Source Value, Target Value, Description
CODE_LIST_ROWS: list[tuple[str, str, str, str]] = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "01", "CANCELLATION", "Cancellation"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
    ("PO_TYPE", "SA", "STANDALONE", "Stand-alone order"),
    ("PO_TYPE", "NE", "NEW_ORDER", "New order"),
    ("PO_TYPE", "DS", "DROP_SHIP", "Drop ship order"),
    ("PO_TYPE", "RL", "RELEASE", "Release against blanket order"),
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
    ("UOM", "DZ", "DOZEN", "Dozen"),
]

SPEC_META = {
    "Transaction Set": "850",
    "X12 Version": "004010",
    "Spec Name": "Synthetic retail PO - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-05",
}


def _fill_rows(ws: Worksheet, rows: list[tuple[str, ...]]) -> None:
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            if value != "":
                ws.cell(row=r, column=c, value=value)


def generate_spec(
    path: Path,
    rows: list[tuple[str, ...]],
    code_lists: list[tuple[str, str, str, str]],
    meta: dict[str, str],
) -> None:
    """Write a reference mapping spec workbook."""
    wb = build_template()
    _fill_rows(wb["Mapping"], rows)
    _fill_rows(wb["CodeLists"], code_lists)
    ws_meta = wb["Meta"]
    written: set = set()
    for r in range(2, ws_meta.max_row + 1):
        key = ws_meta.cell(row=r, column=1).value
        if key in meta:
            ws_meta.cell(row=r, column=2, value=meta[key])
            written.add(key)
    # append any meta keys the template doesn't predefine (e.g. Pairing Key)
    next_row = ws_meta.max_row + 1
    for key, value in meta.items():
        if key not in written:
            ws_meta.cell(row=next_row, column=1, value=key)
            ws_meta.cell(row=next_row, column=2, value=value)
            next_row += 1
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Synthetic X12 850 source files
# --------------------------------------------------------------------------


def build_850(
    business_segments: list[str],
    control_number: str,
    se_count_offset: int = 0,
) -> str:
    """Wrap business segments in a full ISA/GS/ST...SE/GE/IEA interchange.

    ``se_count_offset`` deliberately corrupts the SE segment count so pyx12's
    control checks have something to catch in the defective file.
    """
    icn = control_number.zfill(9)
    st = f"ST*850*{control_number.zfill(4)}"
    # SE count = business segments + ST + SE
    se = f"SE*{len(business_segments) + 2 + se_count_offset}*{control_number.zfill(4)}"
    segments = [
        "ISA*00*          *00*          *ZZ*MAPCHECKSND    *ZZ*MAPCHECKRCV    "
        f"*260615*1200*U*00401*{icn}*0*T*>",
        f"GS*PO*MAPCHECKSND*MAPCHECKRCV*20260615*1200*{int(control_number)}*X*004010",
        st,
        *business_segments,
        se,
        f"GE*1*{int(control_number)}",
        f"IEA*1*{icn}",
    ]
    return "~\n".join(segments) + "~\n"


BASELINE_850 = [
    "BEG*00*SA*PO4400021**20260615",
    "CUR*BY*USD",
    "REF*DP*045",
    "REF*IA*VEND8821",
    "REF*PD*SUMMER26",
    "DTM*002*20260701",
    "DTM*001*20260801",
    "SAC*A*C310***2500",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "N3*4501 CASCADE AVE",
    "N4*BOULDER*CO*80301",
    "N1*BT*ALPINE OUTFITTERS CORPORATE*92*0001",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "PO1*2*6*CA*24**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "PO1*3*5*DZ*12**UP*614141007351",
    "PID*F****GRANOLA BARS VARIETY",
    "CTT*3",
    "AMT*TT*306",
]

# Sparse but valid: exercises SKIP conditionals, the BEG04 default, and
# NOT TESTED statuses for segments this partner simply does not send.
MINIMAL_850 = [
    "BEG*00*NE*PO7700003**20260620",
    "N1*ST*RIVERBEND MARKET*92*0007",
    "PO1*1*100*EA*0.99**UP*614141007352",
    "PID*F****NO2 PENCILS 10PK",
    "CTT*1",
]

# Deliberately defective source. Planted defects:
#   1. BEG05 is not a valid date (month 13)
#   2. PO103 on line 2 uses UOM code 'XX', absent from the UOM code list
#   3. CTT01 says 5 lines but only 3 PO1 loops are present
#   4. ITD payment terms segment is not referenced by the spec (unmapped)
#   5. The N1[BT] loop is missing entirely
#   6. SE segment count is off by two (control-level defect for pyx12)
DEFECTS_850 = [
    "BEG*00*SA*PO4400022**20261301",
    "CUR*BY*USD",
    "REF*DP*045",
    "REF*IA*VEND8821",
    "REF*PD*SUMMER26",
    "DTM*002*20260701",
    "DTM*001*20260801",
    "ITD*01*3*2**30**60",
    "SAC*A*C310***2500",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "N3*4501 CASCADE AVE",
    "N4*BOULDER*CO*80301",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "PO1*2*6*XX*24**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "PO1*3*5*DZ*12**UP*614141007351",
    "PID*F****GRANOLA BARS VARIETY",
    "CTT*5",
    "AMT*TT*306",
]


def generate_source_files() -> None:
    source_dir = EXAMPLES / "source"
    source_dir.mkdir(parents=True, exist_ok=True)
    files = {
        "850_baseline.edi": build_850(BASELINE_850, "1"),
        "850_minimal.edi": build_850(MINIMAL_850, "2"),
        "850_defects.edi": build_850(DEFECTS_850, "3", se_count_offset=2),
    }
    for name, content in files.items():
        path = source_dir / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Translated output files (the artifact under test)
# --------------------------------------------------------------------------

BASELINE_LINES = [
    {"line_no": 1, "qty": 12, "uom": "EACH", "unit_price": 8.5,
     "upc": "614141007349", "description": "TRAIL MIX 12OZ"},
    {"line_no": 2, "qty": 6, "uom": "CASE", "unit_price": 24.0,
     "upc": "614141007350", "description": "SPRING WATER 24PK"},
    {"line_no": 3, "qty": 5, "uom": "DOZEN", "unit_price": 12.0,
     "upc": "614141007351", "description": "GRANOLA BARS VARIETY"},
]

BASELINE_OUTPUT: dict = {
    "order": {
        "tx_purpose": "ORIGINAL",
        "po_type": "STANDALONE",
        "drop_ship_flag": "N",
        "po_number": "PO4400021",
        "release_number": "NONE",
        "po_date": "2026-06-15",
        "currency": "USD",
        "department": "045",
        "vendor_number": "VEND8821",
        "promo_code": "SUMMER26",
        "requested_delivery_date": "2026-07-01",
        "cancel_after_date": "2026-08-01",
        "record_type": "PO_INBOUND",
        "allowance_code": "C310",
        "allowance_amount": 25.0,
    },
    "ship_to": {
        "name": "ALPINE OUTFITTERS STORE 118",
        "id": "0118",
        "address1": "4501 CASCADE AVE",
        "city": "BOULDER",
        "state": "CO",
        "zip": "80301",
    },
    "bill_to": {"name": "ALPINE OUTFITTERS CORPORATE", "id": "0001"},
    "lines": BASELINE_LINES,
    "summary": {"line_count": 3, "total_amount": 306.0},
}

MINIMAL_OUTPUT: dict = {
    "order": {
        "tx_purpose": "ORIGINAL",
        "po_type": "NEW_ORDER",
        "drop_ship_flag": "N",
        "po_number": "PO7700003",
        "release_number": "NONE",
        "po_date": "2026-06-20",
        "record_type": "PO_INBOUND",
    },
    "ship_to": {"name": "RIVERBEND MARKET", "id": "0007"},
    "lines": [
        {"line_no": 1, "qty": 100, "uom": "EACH", "unit_price": 0.99,
         "upc": "614141007352", "description": "NO2 PENCILS 10PK"},
    ],
    "summary": {"line_count": 1},
}


def _defective_baseline_output() -> dict:
    """Baseline output with deliberately planted mapping defects.

    Planted defects (validated against 850_baseline.edi):
      1. po_number transposed                     -> value_mismatch
      2. po_date in US format                     -> format
      3. drop_ship_flag Y though BEG02 is SA      -> condition_logic
      4. line 1 uom left untranslated ('EA')      -> code_translation
      5. record_type constant missing             -> constant_default
      6. allowance_amount implied decimal missed  -> value_mismatch (100x)
      7. third line dropped                       -> count_mismatch + missing_output
      8. extra field warehouse_code               -> unmapped_target warning
      9. ship_to.state four characters            -> format (len:2..2)
     10. release_number default not applied       -> missing_output
     11. line 2 qty carried as a JSON string      -> format warning
    """
    out = json.loads(json.dumps(BASELINE_OUTPUT))  # deep copy
    out["order"]["po_number"] = "PO4400012"
    out["order"]["po_date"] = "06/15/2026"
    out["order"]["drop_ship_flag"] = "Y"
    out["lines"][0]["uom"] = "EA"
    del out["order"]["record_type"]
    out["order"]["allowance_amount"] = 2500
    del out["lines"][2]
    out["order"]["warehouse_code"] = "WH9"
    out["ship_to"]["state"] = "COLO"
    del out["order"]["release_number"]
    out["lines"][1]["qty"] = "6"
    return out


def _naive_bad_source_output() -> dict:
    """What a naive translator emits from 850_defects.edi.

    It slices the invalid date into shape, passes the unknown UOM code
    through, trusts the (wrong) CTT count, and still emits a hardcoded
    bill_to that the source no longer sends.
    """
    out = json.loads(json.dumps(BASELINE_OUTPUT))
    out["order"]["po_number"] = "PO4400022"
    out["order"]["po_date"] = "2026-13-01"
    out["lines"][1]["uom"] = "XX"
    out["summary"]["line_count"] = 5
    return out


def _flat_value(value: object) -> str:
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _to_flat(data: dict) -> str:
    """Render a canonical output dict in the keyed flat reference format."""
    lines = ["# EDI MapCheck keyed flat output"]
    order = "|".join(f"{k}={_flat_value(v)}" for k, v in data["order"].items())
    lines.append(f"H|{order}")
    for role in (k for k in data if k not in ("order", "lines", "summary")):
        fields = "|".join(f"{k}={_flat_value(v)}" for k, v in data[role].items())
        lines.append(f"A|role={role}|{fields}")
    for line in data.get("lines", []):
        fields = "|".join(f"{k}={_flat_value(v)}" for k, v in line.items())
        lines.append(f"D|{fields}")
    summary = "|".join(f"{k}={_flat_value(v)}" for k, v in data["summary"].items())
    lines.append(f"S|{summary}")
    return "\n".join(lines) + "\n"


def generate_output_files() -> None:
    out_dir = EXAMPLES / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    json_files = {
        "po_baseline.json": BASELINE_OUTPUT,
        "po_minimal.json": MINIMAL_OUTPUT,
        "po_baseline_defects.json": _defective_baseline_output(),
        "po_from_bad_source.json": _naive_bad_source_output(),
    }
    for name, data in json_files.items():
        path = out_dir / name
        path.write_text(json.dumps(data, indent=2) + "\n")
        print(f"wrote {path.relative_to(REPO_ROOT)}")
    flat_path = out_dir / "po_baseline.flat"
    flat_path.write_text(_to_flat(BASELINE_OUTPUT))
    print(f"wrote {flat_path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# 856 Ship Notice/Manifest: reference spec, sources, and outputs
# --------------------------------------------------------------------------

SPEC856_ROWS: list[tuple[str, ...]] = [
    ("A-001", "BSN01", "", "shipment.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", "Transaction set purpose."),
    ("A-002", "BSN02", "", "shipment.asn_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..30", "Shipment identification."),
    ("A-003", "BSN03", "", "shipment.asn_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("A-004", "BSN04", "", "shipment.asn_time", "DIRECT", "", "", "", "",
     "", "", "time", "%H:%M", ""),
    ("A-005", "BSN05", "", "shipment.structure_code", "CODE_LIST", "", "", "", "",
     "", "ASN_STRUCTURE", "string", "", "Hierarchical structure code."),
    ("A-006", "DTM02", "DTM[011]", "shipment.ship_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "DTM 011 = shipped."),
    ("A-007", "TD101", "HL[S]", "shipment.packaging_code", "CONDITIONAL",
     "Map the packaging code when a shipment-level TD1 is sent.",
     "EXISTS(TD101)", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("A-008", "TD102", "HL[S]", "shipment.carton_count", "CONDITIONAL",
     "Map the lading quantity when a shipment-level TD1 is sent.",
     "EXISTS(TD102)", "SOURCE", "SKIP", "", "", "integer", "", ""),
    ("A-009", "TD503", "HL[S]", "shipment.scac", "CONDITIONAL",
     "Map the carrier code only when the ID qualifier says SCAC (2).",
     "TD502 = '2'", "SOURCE", "SKIP", "", "", "string", "len:2..4", ""),
    ("A-010", "TD504", "HL[S]", "shipment.ship_mode", "CODE_LIST", "", "", "", "",
     "", "SHIP_METHOD", "string", "", "Transportation method."),
    ("A-011", "REF02", "HL[S]", "shipment.bol_number", "CONDITIONAL",
     "Map the bill of lading number from the shipment-level REF.",
     "REF01 = 'BM'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("A-012", "", "", "shipment.record_type", "CONSTANT", "", "", "", "",
     "ASN_INBOUND", "", "string", "", "Hardcoded routing constant."),
    ("A-013", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", "Party loop nested in the shipment HL."),
    ("A-014", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the location number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("A-015", "N301", "N1[ST]", "ship_to.address1", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("A-016", "N401", "N1[ST]", "ship_to.city", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("A-017", "N402", "N1[ST]", "ship_to.state", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", ""),
    ("A-018", "N403", "N1[ST]", "ship_to.zip", "DIRECT", "", "", "", "",
     "", "", "string", "len:5..10", ""),
    ("A-019", "PRF01", "HL[I]", "lines[].po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22",
     "Read from the item's ancestor order (O) level."),
    ("A-020", "SN101", "HL[I]", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("A-021", "SN102", "HL[I]", "lines[].qty_shipped", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Units shipped for the item."),
    ("A-022", "SN103", "HL[I]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("A-023", "LIN03", "HL[I]", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "LIN02 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("A-024", "PID05", "HL[I]", "lines[].description", "CONDITIONAL",
     "Map the free-form description when the PID is free-form.",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..80", ""),
    ("A-025", "MAN02", "HL[I]", "lines[].sscc", "CONDITIONAL",
     "Map the carton SSCC (from the ancestor pack level) when marked GM.",
     "MAN01 = 'GM'", "SOURCE", "SKIP", "", "", "string", "len:18..20",
     "Standard-carton ASNs have no pack level; the field is simply absent."),
    ("A-026", "CTT01", "", "summary.hl_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", "HL count as transmitted."),
    ("A-027", "HL", "", "summary.hl_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", "Actual HL loops must also match the output."),
    ("A-028", "CTT02", "", "summary.total_units", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Hash total of SN1 quantities."),
]

CODE_LIST_856_ROWS: list[tuple[str, str, str, str]] = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "01", "CANCELLATION", "Cancellation"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
    ("ASN_STRUCTURE", "0001", "SHIPMENT_ORDER_PACK_ITEM", "Pick and pack"),
    ("ASN_STRUCTURE", "0002", "SHIPMENT_ORDER_ITEM", "Standard carton"),
    ("ASN_STRUCTURE", "0004", "SHIPMENT_ORDER_TARE_PACK_ITEM", "Palletized"),
    ("SHIP_METHOD", "M", "MOTOR", "Motor (common carrier)"),
    ("SHIP_METHOD", "A", "AIR", "Air"),
    ("SHIP_METHOD", "R", "RAIL", "Rail"),
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
    ("UOM", "DZ", "DOZEN", "Dozen"),
]

SPEC856_META = {
    "Transaction Set": "856",
    "X12 Version": "004010",
    "Spec Name": "Synthetic retail ASN - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-05",
}

PICKPACK_856 = [
    "BSN*00*ASN20260708001*20260708*1015*0001",
    "DTM*011*20260708",
    "HL*1**S",
    "TD1*CTN25*2",
    "TD5**2*RDWY*M",
    "REF*BM*BOL8802214",
    "N1*ST*ALPINE OUTFITTERS DC 12*92*0012",
    "N3*880 GRANITE PARKWAY",
    "N4*DENVER*CO*80201",
    "HL*2*1*O",
    "PRF*PO4400021",
    "HL*3*2*P",
    "MAN*GM*00061414100734900012",
    "HL*4*3*I",
    "LIN**UP*614141007349",
    "SN1*1*12*EA",
    "PID*F****TRAIL MIX 12OZ",
    "HL*5*3*I",
    "LIN**UP*614141007350",
    "SN1*2*24*EA",
    "PID*F****SPRING WATER 24PK",
    "HL*6*2*P",
    "MAN*GM*00061414100734900029",
    "HL*7*6*I",
    "LIN**UP*614141007351",
    "SN1*3*60*EA",
    "PID*F****GRANOLA BARS VARIETY",
    "CTT*7*96",
]

STANDARD_856 = [
    "BSN*00*ASN20260712002*20260712*0830*0002",
    "DTM*011*20260712",
    "HL*1**S",
    "TD5**2*UPSN*M",
    "REF*BM*BOL8802290",
    "N1*ST*RIVERBEND MARKET*92*0007",
    "N3*15 HARBOR ROW",
    "N4*PORTLAND*OR*97201",
    "HL*2*1*O",
    "PRF*PO7700003",
    "HL*3*2*I",
    "LIN**UP*614141007352",
    "SN1*1*100*EA",
    "PID*F****NO2 PENCILS 10PK",
    "HL*4*2*I",
    "LIN**UP*614141007353",
    "SN1*2*40*EA",
    "PID*F****MARKERS 8CT",
    "CTT*4*140",
]

# Deliberately defective ASN. Planted defects:
#   1. HL*4 is an item directly under the shipment (illegal nesting)
#   2. HL*6 points at parent id 9, which does not exist (orphan)
#   3. HL*7 uses unknown level code X (and X is no valid child of O)
#   4. CTT01 says 9 HL loops; the file has 7
#   5. CTT02 hash total 999 vs actual SN1 sum of 96
#   6. TD1 claims 5 cartons; only one pack (P) loop exists
#   7. PKG segment is not in the 856 definition (unmapped structure)
#   8. HL*7 carries a REF no spec rule references (unmapped source data)
DEFECTS_856 = [
    "BSN*00*ASN20260715003*20260715*1400*0001",
    "DTM*011*20260715",
    "HL*1**S",
    "TD1*CTN25*5",
    "TD5**2*RDWY*M",
    "REF*BM*BOL8802333",
    "PKG*F*01",
    "N1*ST*ALPINE OUTFITTERS DC 12*92*0012",
    "N3*880 GRANITE PARKWAY",
    "N4*DENVER*CO*80201",
    "HL*2*1*O",
    "PRF*PO4400022",
    "HL*3*2*P",
    "MAN*GM*00061414100734900036",
    "HL*4*1*I",
    "LIN**UP*614141007349",
    "SN1*1*12*EA",
    "PID*F****TRAIL MIX 12OZ",
    "HL*5*3*I",
    "LIN**UP*614141007350",
    "SN1*2*24*EA",
    "PID*F****SPRING WATER 24PK",
    "HL*6*9*I",
    "LIN**UP*614141007351",
    "SN1*3*60*EA",
    "PID*F****GRANOLA BARS VARIETY",
    "HL*7*2*X",
    "REF*ZZ*MYSTERY",
    "CTT*9*999",
]


def _asn_line(
    line_no: int, qty: int, upc: str, description: str,
    po_number: str | None = None, sscc: str | None = None,
) -> dict:
    line: dict = {}
    if po_number is not None:
        line["po_number"] = po_number
    line.update(
        {"line_no": line_no, "qty_shipped": qty, "uom": "EACH",
         "upc": upc, "description": description}
    )
    if sscc is not None:
        line["sscc"] = sscc
    return line


PICKPACK_OUTPUT: dict = {
    "shipment": {
        "purpose": "ORIGINAL",
        "asn_number": "ASN20260708001",
        "asn_date": "2026-07-08",
        "asn_time": "10:15",
        "structure_code": "SHIPMENT_ORDER_PACK_ITEM",
        "ship_date": "2026-07-08",
        "packaging_code": "CTN25",
        "carton_count": 2,
        "scac": "RDWY",
        "ship_mode": "MOTOR",
        "bol_number": "BOL8802214",
        "record_type": "ASN_INBOUND",
    },
    "ship_to": {
        "name": "ALPINE OUTFITTERS DC 12", "id": "0012",
        "address1": "880 GRANITE PARKWAY", "city": "DENVER",
        "state": "CO", "zip": "80201",
    },
    "lines": [
        _asn_line(1, 12, "614141007349", "TRAIL MIX 12OZ",
                  po_number="PO4400021", sscc="00061414100734900012"),
        _asn_line(2, 24, "614141007350", "SPRING WATER 24PK",
                  po_number="PO4400021", sscc="00061414100734900012"),
        _asn_line(3, 60, "614141007351", "GRANOLA BARS VARIETY",
                  po_number="PO4400021", sscc="00061414100734900029"),
    ],
    "summary": {"hl_count": 7, "total_units": 96},
}

STANDARD_OUTPUT: dict = {
    "shipment": {
        "purpose": "ORIGINAL",
        "asn_number": "ASN20260712002",
        "asn_date": "2026-07-12",
        "asn_time": "08:30",
        "structure_code": "SHIPMENT_ORDER_ITEM",
        "ship_date": "2026-07-12",
        "scac": "UPSN",
        "ship_mode": "MOTOR",
        "bol_number": "BOL8802290",
        "record_type": "ASN_INBOUND",
    },
    "ship_to": {
        "name": "RIVERBEND MARKET", "id": "0007",
        "address1": "15 HARBOR ROW", "city": "PORTLAND",
        "state": "OR", "zip": "97201",
    },
    "lines": [
        _asn_line(1, 100, "614141007352", "NO2 PENCILS 10PK", po_number="PO7700003"),
        _asn_line(2, 40, "614141007353", "MARKERS 8CT", po_number="PO7700003"),
    ],
    "summary": {"hl_count": 4, "total_units": 140},
}

# What a naive translator emits from the defective ASN: it trusts CTT
# outright and carries whatever parent context each item actually has.
DEFECTS_856_OUTPUT: dict = {
    "shipment": {
        "purpose": "ORIGINAL",
        "asn_number": "ASN20260715003",
        "asn_date": "2026-07-15",
        "asn_time": "14:00",
        "structure_code": "SHIPMENT_ORDER_PACK_ITEM",
        "ship_date": "2026-07-15",
        "packaging_code": "CTN25",
        "carton_count": 5,
        "scac": "RDWY",
        "ship_mode": "MOTOR",
        "bol_number": "BOL8802333",
        "record_type": "ASN_INBOUND",
    },
    "ship_to": {
        "name": "ALPINE OUTFITTERS DC 12", "id": "0012",
        "address1": "880 GRANITE PARKWAY", "city": "DENVER",
        "state": "CO", "zip": "80201",
    },
    "lines": [
        _asn_line(1, 12, "614141007349", "TRAIL MIX 12OZ"),
        _asn_line(2, 24, "614141007350", "SPRING WATER 24PK",
                  po_number="PO4400022", sscc="00061414100734900036"),
        _asn_line(3, 60, "614141007351", "GRANOLA BARS VARIETY"),
    ],
    "summary": {"hl_count": 9, "total_units": 999},
}


def generate_856_files() -> None:
    source_dir = EXAMPLES / "source"
    out_dir = EXAMPLES / "output"
    sources = {
        "856_pickpack.edi": build_856(PICKPACK_856, "11"),
        "856_standard.edi": build_856(STANDARD_856, "12"),
        "856_defects.edi": build_856(DEFECTS_856, "13"),
    }
    for name, content in sources.items():
        path = source_dir / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")
    outputs = {
        "asn_pickpack.json": PICKPACK_OUTPUT,
        "asn_standard.json": STANDARD_OUTPUT,
        "asn_defects.json": DEFECTS_856_OUTPUT,
    }
    for name, data in outputs.items():
        path = out_dir / name
        path.write_text(json.dumps(data, indent=2) + "\n")
        print(f"wrote {path.relative_to(REPO_ROOT)}")


def build_856(business_segments: list[str], control_number: str) -> str:
    """Wrap 856 business segments in a full interchange (GS01 = SH)."""
    icn = control_number.zfill(9)
    segments = [
        "ISA*00*          *00*          *ZZ*MAPCHECKSND    *ZZ*MAPCHECKRCV    "
        f"*260708*1200*U*00401*{icn}*0*T*>",
        f"GS*SH*MAPCHECKSND*MAPCHECKRCV*20260708*1200*{int(control_number)}*X*004010",
        f"ST*856*{control_number.zfill(4)}",
        *business_segments,
        f"SE*{len(business_segments) + 2}*{control_number.zfill(4)}",
        f"GE*1*{int(control_number)}",
        f"IEA*1*{icn}",
    ]
    return "~\n".join(segments) + "~\n"


# --------------------------------------------------------------------------
# Shared interchange builder for the order-cycle sets
# --------------------------------------------------------------------------


def build_interchange(
    business_segments: list[str], control_number: str, st_code: str, gs_code: str
) -> str:
    """Wrap business segments in a full ISA/GS/ST...SE/GE/IEA interchange."""
    icn = control_number.zfill(9)
    segments = [
        "ISA*00*          *00*          *ZZ*MAPCHECKSND    *ZZ*MAPCHECKRCV    "
        f"*260715*1200*U*00401*{icn}*0*T*>",
        f"GS*{gs_code}*MAPCHECKSND*MAPCHECKRCV*20260715*1200*{int(control_number)}*X*004010",
        f"ST*{st_code}*{control_number.zfill(4)}",
        *business_segments,
        f"SE*{len(business_segments) + 2}*{control_number.zfill(4)}",
        f"GE*1*{int(control_number)}",
        f"IEA*1*{icn}",
    ]
    return "~\n".join(segments) + "~\n"


def build_multi_interchange(
    transactions: list[list[str]], control_number: str, st_code: str, gs_code: str
) -> str:
    """Wrap several ST/SE transactions in one ISA/GS...GE/IEA interchange."""
    icn = control_number.zfill(9)
    segments = [
        "ISA*00*          *00*          *ZZ*MAPCHECKSND    *ZZ*MAPCHECKRCV    "
        f"*260715*1200*U*00401*{icn}*0*T*>",
        f"GS*{gs_code}*MAPCHECKSND*MAPCHECKRCV*20260715*1200*{int(control_number)}*X*004010",
    ]
    for index, business_segments in enumerate(transactions, start=1):
        stc = str(index).zfill(4)
        segments.append(f"ST*{st_code}*{stc}")
        segments.extend(business_segments)
        segments.append(f"SE*{len(business_segments) + 2}*{stc}")
    segments.append(f"GE*{len(transactions)}*{int(control_number)}")
    segments.append(f"IEA*1*{icn}")
    return "~\n".join(segments) + "~\n"


# --------------------------------------------------------------------------
# 855 Purchase Order Acknowledgment
# --------------------------------------------------------------------------

SPEC855_ROWS: list[tuple[str, ...]] = [
    ("K-001", "BAK01", "", "order.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("K-002", "BAK02", "", "order.ack_type", "CODE_LIST", "", "", "", "",
     "", "ACK_TYPE", "string", "", "Acknowledgment type."),
    ("K-003", "BAK03", "", "order.po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", "PO number being acknowledged."),
    ("K-004", "BAK04", "", "order.po_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("K-005", "BAK08", "", "order.vendor_order_number", "CONDITIONAL",
     "Map the vendor's order number when they send one.",
     "EXISTS(BAK08)", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("K-006", "", "", "order.record_type", "CONSTANT", "", "", "", "",
     "POA_INBOUND", "", "string", "", ""),
    ("K-007", "N102", "N1[SE]", "vendor.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", "Selling party."),
    ("K-008", "N104", "N1[SE]", "vendor.id", "CONDITIONAL",
     "Map the vendor number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("K-009", "PO101", "PO1", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("K-010", "PO102", "PO1", "lines[].qty_ordered", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Ordered quantity as echoed back."),
    ("K-011", "PO103", "PO1", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("K-012", "PO104", "PO1", "lines[].unit_price", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", ""),
    ("K-013", "PO107", "PO1", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "PO106 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("K-014", "ACK01", "PO1", "lines[].line_status", "CODE_LIST", "", "", "", "",
     "", "ACK_STATUS", "string", "",
     "First ACK's status; split lines reconcile via the definition rollup."),
    ("K-015", "ACK02", "PO1", "lines[].qty_acknowledged", "DIRECT", "", "", "", "",
     "", "", "integer", "", "First ACK's quantity."),
    ("K-016", "ACK03", "PO1", "lines[].ack_uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("K-017", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("K-018", "PO1", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_855_ROWS: list[tuple[str, str, str, str]] = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "01", "CANCELLATION", "Cancellation"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
    ("ACK_TYPE", "AC", "ACKNOWLEDGE", "Acknowledge, no detail"),
    ("ACK_TYPE", "AD", "ACK_WITH_DETAIL", "Acknowledge with detail, no change"),
    ("ACK_TYPE", "AK", "ACK_WITH_CHANGE", "Acknowledge with detail and change"),
    ("ACK_TYPE", "RD", "REJECT_WITH_DETAIL", "Reject with detail"),
    ("ACK_TYPE", "RJ", "REJECTED", "Rejected, no detail"),
    ("ACK_STATUS", "IA", "ACCEPTED", "Item accepted"),
    ("ACK_STATUS", "IB", "BACKORDERED", "Item backordered"),
    ("ACK_STATUS", "IQ", "QTY_CHANGED", "Item accepted, quantity changed"),
    ("ACK_STATUS", "IR", "REJECTED", "Item rejected"),
    ("ACK_STATUS", "DR", "DATE_RESCHEDULED", "Item accepted, date rescheduled"),
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
    ("UOM", "DZ", "DOZEN", "Dozen"),
]

SPEC855_META = {
    "Transaction Set": "855",
    "X12 Version": "004010",
    "Spec Name": "Synthetic retail POA - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-05",
}

BASELINE_855 = [
    "BAK*00*AD*PO4400021*20260615****VN2088841",
    "N1*SE*SUMMIT WHOLESALE FOODS*92*7731",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "ACK*IA*12*EA",
    "PO1*2*6*CA*24**UP*614141007350",
    "ACK*IA*4*CA",
    "ACK*IR*2*CA",
    "PO1*3*5*DZ*12**UP*614141007351",
    "ACK*IA*5*DZ",
    "CTT*3",
]

# Deliberately defective 855. Planted defects:
#   1. Line 2 orders 6 but the single ACK accounts for only 4 (2 vanish)
#   2. Line 3's ACK uses unknown status code ZZ
#   3. CTT01 says 5 lines; the file has 3
#   4. BAK04 is not a valid date (day 35)
#   5. ITD terms segment carries data no spec rule references
DEFECTS_855 = [
    "BAK*00*AD*PO4400022*20260635****VN2088842",
    "ITD*01*3*2**30**60",
    "N1*SE*SUMMIT WHOLESALE FOODS*92*7731",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "ACK*IA*12*EA",
    "PO1*2*6*CA*24**UP*614141007350",
    "ACK*IA*4*CA",
    "PO1*3*5*DZ*12**UP*614141007351",
    "ACK*ZZ*5*DZ",
    "CTT*5",
]


def _poa_line(
    line_no: int, qty_ordered: int, uom: str, unit_price: float, upc: str,
    line_status: str, qty_acknowledged: int,
) -> dict:
    return {
        "line_no": line_no, "qty_ordered": qty_ordered, "uom": uom,
        "unit_price": unit_price, "upc": upc, "line_status": line_status,
        "qty_acknowledged": qty_acknowledged, "ack_uom": uom,
    }


BASELINE_855_OUTPUT: dict = {
    "order": {
        "purpose": "ORIGINAL",
        "ack_type": "ACK_WITH_DETAIL",
        "po_number": "PO4400021",
        "po_date": "2026-06-15",
        "vendor_order_number": "VN2088841",
        "record_type": "POA_INBOUND",
    },
    "vendor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _poa_line(1, 12, "EACH", 8.5, "614141007349", "ACCEPTED", 12),
        _poa_line(2, 6, "CASE", 24.0, "614141007350", "ACCEPTED", 4),
        _poa_line(3, 5, "DOZEN", 12.0, "614141007351", "ACCEPTED", 5),
    ],
    "summary": {"line_count": 3},
}

DEFECTS_855_OUTPUT: dict = {
    "order": {
        "purpose": "ORIGINAL",
        "ack_type": "ACK_WITH_DETAIL",
        "po_number": "PO4400022",
        "po_date": "2026-06-35",
        "vendor_order_number": "VN2088842",
        "record_type": "POA_INBOUND",
    },
    "vendor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _poa_line(1, 12, "EACH", 8.5, "614141007349", "ACCEPTED", 12),
        _poa_line(2, 6, "CASE", 24.0, "614141007350", "ACCEPTED", 4),
        _poa_line(3, 5, "DOZEN", 12.0, "614141007351", "ZZ", 5),
    ],
    "summary": {"line_count": 5},
}


# --------------------------------------------------------------------------
# 810 Invoice
# --------------------------------------------------------------------------

SPEC810_ROWS: list[tuple[str, ...]] = [
    ("V-001", "BIG01", "", "invoice.invoice_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("V-002", "BIG02", "", "invoice.invoice_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("V-003", "BIG03", "", "invoice.po_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("V-004", "BIG04", "", "invoice.po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("V-005", "", "", "invoice.record_type", "CONSTANT", "", "", "", "",
     "INVOICE_INBOUND", "", "string", "", ""),
    ("V-006", "N102", "N1[RE]", "remit_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("V-007", "N104", "N1[RE]", "remit_to.id", "CONDITIONAL",
     "Map the remit-to ID only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("V-008", "IT101", "IT1", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("V-009", "IT102", "IT1", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Quantity invoiced."),
    ("V-010", "IT103", "IT1", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("V-011", "IT104", "IT1", "lines[].unit_price", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", ""),
    ("V-012", "IT107", "IT1", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "IT106 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("V-013", "PID05", "IT1", "lines[].description", "CONDITIONAL",
     "Map the free-form description when the PID is free-form.",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..80", ""),
    ("V-014", "TDS01", "", "summary.invoice_total", "DIRECT", "", "", "", "",
     "", "", "decimal", "implied:2;places:2", "TDS01 is X12 N2: implied 2 decimals."),
    ("V-015", "SAC02", "SAC[C]", "summary.charge_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:4..4", "Summary-level charge (SAC01 = C)."),
    ("V-016", "SAC05", "SAC[C]", "summary.charge_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "implied:2;places:2", ""),
    ("V-017", "SAC02", "SAC[A]", "summary.allowance_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:4..4", "Summary-level allowance (SAC01 = A)."),
    ("V-018", "SAC05", "SAC[A]", "summary.allowance_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "implied:2;places:2", ""),
    ("V-019", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("V-020", "IT1", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_810_ROWS: list[tuple[str, str, str, str]] = [
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
    ("UOM", "DZ", "DOZEN", "Dozen"),
]

SPEC810_META = {
    "Transaction Set": "810",
    "X12 Version": "004010",
    "Spec Name": "Synthetic retail invoice - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-05",
}

# Invoice math: 12x8.50 + 6x24.00 + 5x12.00 = 306.00
#               + 15.00 freight charge - 8.00 allowance = 313.00
BASELINE_810 = [
    "BIG*20260620*INV88214*20260615*PO4400021",
    "N1*RE*SUMMIT WHOLESALE FOODS*92*7731",
    "IT1*1*12*EA*8.5**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "IT1*2*6*CA*24**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "IT1*3*5*DZ*12**UP*614141007351",
    "PID*F****GRANOLA BARS VARIETY",
    "TDS*31300",
    "SAC*C*D240***1500",
    "SAC*A*C310***800",
    "CTT*3",
]

# Deliberately defective 810. Planted defects:
#   1. TDS says 321.00 — exactly the total with the 8.00 allowance missed
#   2. Line 2 uses unknown UOM code XX
#   3. CTT01 says 4 lines; the file has 3
#   4. BIG01 is not a valid date (day 35)
#   5. ITD terms segment carries data no spec rule references
DEFECTS_810 = [
    "BIG*20260635*INV88215*20260615*PO4400022",
    "ITD*01*3*2**30**60",
    "N1*RE*SUMMIT WHOLESALE FOODS*92*7731",
    "IT1*1*12*EA*8.5**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "IT1*2*6*XX*24**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "IT1*3*5*DZ*12**UP*614141007351",
    "PID*F****GRANOLA BARS VARIETY",
    "TDS*32100",
    "SAC*C*D240***1500",
    "SAC*A*C310***800",
    "CTT*4",
]


def _invoice_line(
    line_no: int, qty: int, uom: str, unit_price: float, upc: str, description: str
) -> dict:
    return {
        "line_no": line_no, "qty": qty, "uom": uom, "unit_price": unit_price,
        "upc": upc, "description": description,
    }


BASELINE_810_OUTPUT: dict = {
    "invoice": {
        "invoice_date": "2026-06-20",
        "invoice_number": "INV88214",
        "po_date": "2026-06-15",
        "po_number": "PO4400021",
        "record_type": "INVOICE_INBOUND",
    },
    "remit_to": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _invoice_line(1, 12, "EACH", 8.5, "614141007349", "TRAIL MIX 12OZ"),
        _invoice_line(2, 6, "CASE", 24.0, "614141007350", "SPRING WATER 24PK"),
        _invoice_line(3, 5, "DOZEN", 12.0, "614141007351", "GRANOLA BARS VARIETY"),
    ],
    "summary": {
        "invoice_total": 313.0,
        "charge_code": "D240",
        "charge_amount": 15.0,
        "allowance_code": "C310",
        "allowance_amount": 8.0,
        "line_count": 3,
    },
}

DEFECTS_810_OUTPUT: dict = {
    "invoice": {
        "invoice_date": "2026-06-35",
        "invoice_number": "INV88215",
        "po_date": "2026-06-15",
        "po_number": "PO4400022",
        "record_type": "INVOICE_INBOUND",
    },
    "remit_to": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _invoice_line(1, 12, "EACH", 8.5, "614141007349", "TRAIL MIX 12OZ"),
        _invoice_line(2, 6, "XX", 24.0, "614141007350", "SPRING WATER 24PK"),
        _invoice_line(3, 5, "DOZEN", 12.0, "614141007351", "GRANOLA BARS VARIETY"),
    ],
    "summary": {
        "invoice_total": 321.0,
        "charge_code": "D240",
        "charge_amount": 15.0,
        "allowance_code": "C310",
        "allowance_amount": 8.0,
        "line_count": 4,
    },
}


def generate_order_cycle_files() -> None:
    source_dir = EXAMPLES / "source"
    out_dir = EXAMPLES / "output"
    sources = {
        "855_baseline.edi": build_interchange(BASELINE_855, "21", "855", "PR"),
        "855_defects.edi": build_interchange(DEFECTS_855, "22", "855", "PR"),
        "810_baseline.edi": build_interchange(BASELINE_810, "23", "810", "IN"),
        "810_defects.edi": build_interchange(DEFECTS_810, "24", "810", "IN"),
    }
    for name, content in sources.items():
        path = source_dir / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")
    outputs = {
        "poa_baseline.json": BASELINE_855_OUTPUT,
        "poa_defects.json": DEFECTS_855_OUTPUT,
        "invoice_baseline.json": BASELINE_810_OUTPUT,
        "invoice_defects.json": DEFECTS_810_OUTPUT,
    }
    for name, data in outputs.items():
        path = out_dir / name
        path.write_text(json.dumps(data, indent=2) + "\n")
        print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Warehouse suite: 940 / 945 / 943 / 944 / 947
# --------------------------------------------------------------------------

_WH_UOM = [
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
]
_WH_SHIP_METHOD = [
    ("SHIP_METHOD", "M", "MOTOR", "Motor (common carrier)"),
    ("SHIP_METHOD", "A", "AIR", "Air"),
    ("SHIP_METHOD", "R", "RAIL", "Rail"),
]

SPEC940_ROWS: list[tuple[str, ...]] = [
    ("O-001", "W0501", "", "order.order_status", "CODE_LIST", "", "", "", "",
     "", "ORDER_STATUS", "string", "", ""),
    ("O-002", "W0502", "", "order.order_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", "Depositor order number."),
    ("O-003", "W0503", "", "order.po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("O-004", "G6202", "G62[10]", "order.requested_ship_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "G62 qualifier 10 = requested ship."),
    ("O-005", "W6601", "", "order.freight_terms", "CODE_LIST", "", "", "", "",
     "", "FREIGHT_TERMS", "string", "", ""),
    ("O-006", "W6602", "", "order.ship_method", "CODE_LIST", "", "", "", "",
     "", "SHIP_METHOD", "string", "", ""),
    ("O-007", "", "", "order.record_type", "CONSTANT", "", "", "", "",
     "SHIP_ORDER", "", "string", "", ""),
    ("O-008", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("O-009", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the store number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("O-010", "N301", "N1[ST]", "ship_to.address1", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("O-011", "N401", "N1[ST]", "ship_to.city", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("O-012", "N402", "N1[ST]", "ship_to.state", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", ""),
    ("O-013", "N403", "N1[ST]", "ship_to.zip", "DIRECT", "", "", "", "",
     "", "", "string", "len:5..10", ""),
    ("O-014", "N102", "N1[WH]", "warehouse.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("O-015", "N104", "N1[WH]", "warehouse.id", "CONDITIONAL",
     "Map the warehouse code only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("O-016", "W0101", "W01", "lines[].qty_ordered", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("O-017", "W0102", "W01", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("O-018", "W0103", "W01", "lines[].upc", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("O-019", "W0105", "W01", "lines[].vendor_sku", "CONDITIONAL",
     "Map the vendor SKU only when the ID qualifier is VN.",
     "W0104 = 'VN'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("O-020", "G6901", "W01", "lines[].description", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..80", ""),
    ("O-021", "W01", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_940_ROWS = [
    ("ORDER_STATUS", "N", "NEW", "Original shipping order"),
    ("ORDER_STATUS", "R", "REPLACE", "Replacement order"),
    ("ORDER_STATUS", "F", "CONFIRMATION", "Confirmation"),
    ("FREIGHT_TERMS", "PP", "PREPAID", "Prepaid by seller"),
    ("FREIGHT_TERMS", "CC", "COLLECT", "Collect"),
    *_WH_SHIP_METHOD,
    *_WH_UOM,
]

BASELINE_940 = [
    "W05*N*ORD8801*PO4400021",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "N3*4501 CASCADE AVE",
    "N4*BOULDER*CO*80301",
    "N1*WH*CASCADE 3PL DENVER*92*W22",
    "G62*10*20260722",
    "W66*PP*M",
    "W01*24*EA*614141007349*VN*SKU-1001",
    "G69*TRAIL MIX 12OZ",
    "W01*60*EA*614141007351*VN*SKU-1003",
    "G69*GRANOLA BARS VARIETY",
]

# Defects: unknown order status X, invalid requested-ship date (day 50),
# unknown UOM ZZ on line 1, unreferenced W09 temperature data.
DEFECTS_940 = [
    "W05*X*ORD8802*PO4400022",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "N3*4501 CASCADE AVE",
    "N4*BOULDER*CO*80301",
    "N1*WH*CASCADE 3PL DENVER*92*W22",
    "G62*10*20260750",
    "W66*PP*M",
    "W09*CZ*34*FA",
    "W01*24*ZZ*614141007349*VN*SKU-1001",
    "G69*TRAIL MIX 12OZ",
    "W01*60*EA*614141007351*VN*SKU-1003",
    "G69*GRANOLA BARS VARIETY",
]

BASELINE_940_OUTPUT: dict = {
    "order": {
        "order_status": "NEW", "order_number": "ORD8801", "po_number": "PO4400021",
        "requested_ship_date": "2026-07-22", "freight_terms": "PREPAID",
        "ship_method": "MOTOR", "record_type": "SHIP_ORDER",
    },
    "ship_to": {"name": "ALPINE OUTFITTERS STORE 118", "id": "0118",
                "address1": "4501 CASCADE AVE", "city": "BOULDER",
                "state": "CO", "zip": "80301"},
    "warehouse": {"name": "CASCADE 3PL DENVER", "id": "W22"},
    "lines": [
        {"qty_ordered": 24, "uom": "EACH", "upc": "614141007349",
         "vendor_sku": "SKU-1001", "description": "TRAIL MIX 12OZ"},
        {"qty_ordered": 60, "uom": "EACH", "upc": "614141007351",
         "vendor_sku": "SKU-1003", "description": "GRANOLA BARS VARIETY"},
    ],
    "summary": {"line_count": 2},
}

DEFECTS_940_OUTPUT: dict = json.loads(json.dumps(BASELINE_940_OUTPUT))
DEFECTS_940_OUTPUT["order"].update(
    {"order_status": "X", "order_number": "ORD8802", "po_number": "PO4400022",
     "requested_ship_date": "2026-07-50"}
)
DEFECTS_940_OUTPUT["lines"][0]["uom"] = "ZZ"

SPEC945_ROWS: list[tuple[str, ...]] = [
    ("S-001", "W0601", "", "advice.reporting_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..2", "Passed through; partner-specific values."),
    ("S-002", "W0602", "", "advice.order_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("S-003", "W0603", "", "advice.ship_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("S-004", "W2701", "", "advice.ship_method", "CODE_LIST", "", "", "", "",
     "", "SHIP_METHOD", "string", "", ""),
    ("S-005", "W2702", "", "advice.scac", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..4", ""),
    ("S-006", "", "", "advice.record_type", "CONSTANT", "", "", "", "",
     "SHIP_ADVICE", "", "string", "", ""),
    ("S-007", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("S-008", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the store number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("S-009", "LX01", "LX", "lines[].seq", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("S-010", "W1201", "LX", "lines[].ship_status", "CODE_LIST", "", "", "", "",
     "", "SHIP_STATUS", "string", "", "Assumed codes G/P; amend per guide."),
    ("S-011", "W1202", "LX", "lines[].qty_ordered", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("S-012", "W1203", "LX", "lines[].qty_shipped", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("S-013", "W1204", "LX", "lines[].qty_difference", "CONDITIONAL",
     "Map the quantity difference only when the warehouse declares one.",
     "EXISTS(W1204)", "SOURCE", "SKIP", "", "", "integer", "", ""),
    ("S-014", "W1205", "LX", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("S-015", "W1206", "LX", "lines[].upc", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("S-016", "G6901", "LX", "lines[].description", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..80", ""),
    ("S-017", "W0301", "", "summary.total_shipped", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("S-018", "LX", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
    ("S-019", "W0302", "", "summary.total_weight", "DIRECT", "", "", "", "",
     "", "", "decimal", "", ""),
    ("S-020", "W0303", "", "summary.weight_uom", "CODE_LIST", "", "", "", "",
     "", "WEIGHT_UOM", "string", "", ""),
]

CODE_LIST_945_ROWS = [
    ("SHIP_STATUS", "G", "COMPLETE", "Shipped complete (assumed usage)"),
    ("SHIP_STATUS", "P", "PARTIAL", "Shipped partial (assumed usage)"),
    ("WEIGHT_UOM", "LB", "POUNDS", "Pounds"),
    ("WEIGHT_UOM", "KG", "KILOGRAMS", "Kilograms"),
    *_WH_SHIP_METHOD,
    *_WH_UOM,
]

BASELINE_945 = [
    "W06*F*ORD8801*20260723",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "W27*M*RDWY",
    "LX*1",
    "W12*G*24*24**EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "LX*2",
    "W12*P*60*55*5*EA*614141007351",
    "G69*GRANOLA BARS VARIETY",
    "W03*79*310*LB",
]

# Defects: invalid ship date (day 50), line 1 short-ships 4 units WITHOUT
# declaring a difference, unknown UOM ZZ on line 2, W03 total that matches
# neither the shipped sum nor the declared differences, unreferenced W10.
DEFECTS_945 = [
    "W06*F*ORD8802*20260750",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "W27*M*RDWY",
    "W10*DOCK 7",
    "LX*1",
    "W12*G*24*20**EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "LX*2",
    "W12*P*60*55*5*ZZ*614141007351",
    "G69*GRANOLA BARS VARIETY",
    "W03*80*310*LB",
]

BASELINE_945_OUTPUT: dict = {
    "advice": {"reporting_code": "F", "order_number": "ORD8801",
               "ship_date": "2026-07-23", "ship_method": "MOTOR",
               "scac": "RDWY", "record_type": "SHIP_ADVICE"},
    "ship_to": {"name": "ALPINE OUTFITTERS STORE 118", "id": "0118"},
    "lines": [
        {"seq": 1, "ship_status": "COMPLETE", "qty_ordered": 24, "qty_shipped": 24,
         "uom": "EACH", "upc": "614141007349", "description": "TRAIL MIX 12OZ"},
        {"seq": 2, "ship_status": "PARTIAL", "qty_ordered": 60, "qty_shipped": 55,
         "qty_difference": 5, "uom": "EACH", "upc": "614141007351",
         "description": "GRANOLA BARS VARIETY"},
    ],
    "summary": {"total_shipped": 79, "line_count": 2, "total_weight": 310,
                "weight_uom": "POUNDS"},
}

DEFECTS_945_OUTPUT: dict = json.loads(json.dumps(BASELINE_945_OUTPUT))
DEFECTS_945_OUTPUT["advice"].update({"order_number": "ORD8802", "ship_date": "2026-07-50"})
DEFECTS_945_OUTPUT["lines"][0]["qty_shipped"] = 20
DEFECTS_945_OUTPUT["lines"][1]["uom"] = "ZZ"
DEFECTS_945_OUTPUT["summary"]["total_shipped"] = 80

_XFER_LINES = [
    {"qty": 40, "uom": "EACH", "upc": "614141007349", "description": "TRAIL MIX 12OZ"},
    {"qty": 24, "uom": "EACH", "upc": "614141007350", "description": "SPRING WATER 24PK"},
]

SPEC943_ROWS: list[tuple[str, ...]] = [
    ("T-001", "W0601", "", "transfer.reporting_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..2", ""),
    ("T-002", "W0602", "", "transfer.transfer_order", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("T-003", "W0603", "", "transfer.ship_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("T-004", "W2701", "", "transfer.ship_method", "CODE_LIST", "", "", "", "",
     "", "SHIP_METHOD", "string", "", ""),
    ("T-005", "W2702", "", "transfer.scac", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..4", ""),
    ("T-006", "", "", "transfer.record_type", "CONSTANT", "", "", "", "",
     "XFER_SHIP", "", "string", "", ""),
    ("T-007", "N102", "N1[WH]", "warehouse.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", "Receiving warehouse."),
    ("T-008", "N104", "N1[WH]", "warehouse.id", "CONDITIONAL",
     "Map the warehouse code only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("T-009", "W0401", "W04", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("T-010", "W0402", "W04", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("T-011", "W0403", "W04", "lines[].upc", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("T-012", "G6901", "W04", "lines[].description", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..80", ""),
    ("T-013", "W0301", "", "summary.total_units", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("T-014", "W04", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_943_ROWS = [*_WH_SHIP_METHOD, *_WH_UOM]

BASELINE_943 = [
    "W06*F*XFER7701*20260724",
    "N1*WH*CASCADE 3PL RENO*92*W31",
    "W27*M*RDWY",
    "W04*40*EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W04*24*EA*614141007350",
    "G69*SPRING WATER 24PK",
    "W03*64",
]

# Defects: W03 total lies (70 vs 64 actual), unknown UOM, invalid ship
# date, unreferenced G61 contact data.
DEFECTS_943 = [
    "W06*F*XFER7702*20260750",
    "G61*IC*JANE DOE",
    "N1*WH*CASCADE 3PL RENO*92*W31",
    "W27*M*RDWY",
    "W04*40*ZZ*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W04*24*EA*614141007350",
    "G69*SPRING WATER 24PK",
    "W03*70",
]

BASELINE_943_OUTPUT: dict = {
    "transfer": {"reporting_code": "F", "transfer_order": "XFER7701",
                 "ship_date": "2026-07-24", "ship_method": "MOTOR",
                 "scac": "RDWY", "record_type": "XFER_SHIP"},
    "warehouse": {"name": "CASCADE 3PL RENO", "id": "W31"},
    "lines": [dict(line) for line in _XFER_LINES],
    "summary": {"total_units": 64, "line_count": 2},
}

DEFECTS_943_OUTPUT: dict = json.loads(json.dumps(BASELINE_943_OUTPUT))
DEFECTS_943_OUTPUT["transfer"].update(
    {"transfer_order": "XFER7702", "ship_date": "2026-07-50"}
)
DEFECTS_943_OUTPUT["lines"][0]["uom"] = "ZZ"
DEFECTS_943_OUTPUT["summary"]["total_units"] = 70

SPEC944_ROWS: list[tuple[str, ...]] = [
    ("R-001", "W1701", "", "receipt.reporting_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..2", ""),
    ("R-002", "W1702", "", "receipt.receipt_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("R-003", "W1703", "", "receipt.receipt_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("R-004", "", "", "receipt.record_type", "CONSTANT", "", "", "", "",
     "XFER_RECEIPT", "", "string", "", ""),
    ("R-005", "N102", "N1[WH]", "warehouse.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("R-006", "N104", "N1[WH]", "warehouse.id", "CONDITIONAL",
     "Map the warehouse code only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("R-007", "W0701", "W07", "lines[].qty_received", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("R-008", "W0702", "W07", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("R-009", "W0703", "W07", "lines[].upc", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("R-010", "G6901", "W07", "lines[].description", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..80", ""),
    ("R-011", "W1401", "", "summary.total_received", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("R-012", "W07", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_944_ROWS = [*_WH_UOM]

BASELINE_944 = [
    "W17*F*20260726*REC5501",
    "N1*WH*CASCADE 3PL RENO*92*W31",
    "W07*40*EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W07*24*EA*614141007350",
    "G69*SPRING WATER 24PK",
    "W14*64",
]

# Defects: W14 total lies (60 vs 64 actual), unknown UOM, invalid receipt
# date, unreferenced G61 contact data.
DEFECTS_944 = [
    "W17*F*20260750*REC5502",
    "G61*IC*JANE DOE",
    "N1*WH*CASCADE 3PL RENO*92*W31",
    "W07*40*EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W07*24*ZZ*614141007350",
    "G69*SPRING WATER 24PK",
    "W14*60",
]

BASELINE_944_OUTPUT: dict = {
    "receipt": {"reporting_code": "F", "receipt_date": "2026-07-26",
                "receipt_number": "REC5501", "record_type": "XFER_RECEIPT"},
    "warehouse": {"name": "CASCADE 3PL RENO", "id": "W31"},
    "lines": [
        {"qty_received": 40, "uom": "EACH", "upc": "614141007349",
         "description": "TRAIL MIX 12OZ"},
        {"qty_received": 24, "uom": "EACH", "upc": "614141007350",
         "description": "SPRING WATER 24PK"},
    ],
    "summary": {"total_received": 64, "line_count": 2},
}

DEFECTS_944_OUTPUT: dict = json.loads(json.dumps(BASELINE_944_OUTPUT))
DEFECTS_944_OUTPUT["receipt"].update(
    {"receipt_date": "2026-07-50", "receipt_number": "REC5502"}
)
DEFECTS_944_OUTPUT["lines"][1]["uom"] = "ZZ"
DEFECTS_944_OUTPUT["summary"]["total_received"] = 60

SPEC947_ROWS: list[tuple[str, ...]] = [
    ("J-001", "W1501", "", "adjustment.adjustment_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("J-002", "W1502", "", "adjustment.adjustment_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("J-003", "", "", "adjustment.record_type", "CONSTANT", "", "", "", "",
     "INV_ADJUST", "", "string", "", ""),
    ("J-004", "N102", "N1[WH]", "warehouse.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("J-005", "N104", "N1[WH]", "warehouse.id", "CONDITIONAL",
     "Map the warehouse code only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("J-006", "W1901", "W19", "lines[].reason", "CODE_LIST", "", "", "", "",
     "", "ADJ_REASON", "string", "", "The adjustment-reason showcase."),
    ("J-007", "W1902", "W19", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Credit/debit quantity; negative = loss."),
    ("J-008", "W1903", "W19", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("J-009", "W1904", "W19", "lines[].upc", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("J-010", "G6901", "W19", "lines[].description", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..80", ""),
    ("J-011", "W19", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_947_ROWS = [
    ("ADJ_REASON", "AD", "CYCLE_COUNT_ADJ", "Cycle-count adjustment"),
    ("ADJ_REASON", "CC", "CYCLE_COUNT", "Cycle count"),
    ("ADJ_REASON", "DA", "DAMAGED", "Damaged"),
    ("ADJ_REASON", "EX", "EXPIRED", "Expired product"),
    ("ADJ_REASON", "RE", "RETURN_TO_STOCK", "Return to stock"),
    ("ADJ_REASON", "SH", "SHRINKAGE", "Shrinkage/theft"),
    ("ADJ_REASON", "TR", "TRANSFER", "Transfer"),
    *_WH_UOM,
]

BASELINE_947 = [
    "W15*20260725*ADJ2201",
    "N1*WH*CASCADE 3PL DENVER*92*W22",
    "W19*DA*-5*EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W19*CC*12*EA*614141007350",
    "G69*SPRING WATER 24PK",
    "W19*EX*-3*EA*614141007351",
    "G69*GRANOLA BARS VARIETY",
]

# Defects: unknown adjustment reason XX, invalid adjustment date,
# unreferenced G61 contact data.
DEFECTS_947 = [
    "W15*20260750*ADJ2202",
    "G61*IC*JANE DOE",
    "N1*WH*CASCADE 3PL DENVER*92*W22",
    "W19*XX*-5*EA*614141007349",
    "G69*TRAIL MIX 12OZ",
    "W19*CC*12*EA*614141007350",
    "G69*SPRING WATER 24PK",
]

BASELINE_947_OUTPUT: dict = {
    "adjustment": {"adjustment_date": "2026-07-25", "adjustment_number": "ADJ2201",
                   "record_type": "INV_ADJUST"},
    "warehouse": {"name": "CASCADE 3PL DENVER", "id": "W22"},
    "lines": [
        {"reason": "DAMAGED", "qty": -5, "uom": "EACH",
         "upc": "614141007349", "description": "TRAIL MIX 12OZ"},
        {"reason": "CYCLE_COUNT", "qty": 12, "uom": "EACH",
         "upc": "614141007350", "description": "SPRING WATER 24PK"},
        {"reason": "EXPIRED", "qty": -3, "uom": "EACH",
         "upc": "614141007351", "description": "GRANOLA BARS VARIETY"},
    ],
    "summary": {"line_count": 3},
}

DEFECTS_947_OUTPUT: dict = {
    "adjustment": {"adjustment_date": "2026-07-50", "adjustment_number": "ADJ2202",
                   "record_type": "INV_ADJUST"},
    "warehouse": {"name": "CASCADE 3PL DENVER", "id": "W22"},
    "lines": [
        {"reason": "XX", "qty": -5, "uom": "EACH",
         "upc": "614141007349", "description": "TRAIL MIX 12OZ"},
        {"reason": "CYCLE_COUNT", "qty": 12, "uom": "EACH",
         "upc": "614141007350", "description": "SPRING WATER 24PK"},
    ],
    "summary": {"line_count": 2},
}

#: (set code, GS code, spec rows, code lists, spec name,
#:  [(source file, control#, segments), ...], {output file: data})
WAREHOUSE_SETS = [
    ("940", "OW", SPEC940_ROWS, CODE_LIST_940_ROWS,
     "Synthetic 3PL ship order - reference example",
     [("940_baseline.edi", "31", BASELINE_940), ("940_defects.edi", "32", DEFECTS_940)],
     {"shiporder_baseline.json": BASELINE_940_OUTPUT,
      "shiporder_defects.json": DEFECTS_940_OUTPUT}),
    ("945", "SW", SPEC945_ROWS, CODE_LIST_945_ROWS,
     "Synthetic 3PL ship advice - reference example",
     [("945_baseline.edi", "33", BASELINE_945), ("945_defects.edi", "34", DEFECTS_945)],
     {"shipadvice_baseline.json": BASELINE_945_OUTPUT,
      "shipadvice_defects.json": DEFECTS_945_OUTPUT}),
    ("943", "AR", SPEC943_ROWS, CODE_LIST_943_ROWS,
     "Synthetic stock transfer shipment - reference example",
     [("943_baseline.edi", "35", BASELINE_943), ("943_defects.edi", "36", DEFECTS_943)],
     {"xfership_baseline.json": BASELINE_943_OUTPUT,
      "xfership_defects.json": DEFECTS_943_OUTPUT}),
    ("944", "RE", SPEC944_ROWS, CODE_LIST_944_ROWS,
     "Synthetic stock transfer receipt - reference example",
     [("944_baseline.edi", "37", BASELINE_944), ("944_defects.edi", "38", DEFECTS_944)],
     {"xferreceipt_baseline.json": BASELINE_944_OUTPUT,
      "xferreceipt_defects.json": DEFECTS_944_OUTPUT}),
    ("947", "AW", SPEC947_ROWS, CODE_LIST_947_ROWS,
     "Synthetic inventory adjustment - reference example",
     [("947_baseline.edi", "39", BASELINE_947), ("947_defects.edi", "40", DEFECTS_947)],
     {"invadjust_baseline.json": BASELINE_947_OUTPUT,
      "invadjust_defects.json": DEFECTS_947_OUTPUT}),
]


def generate_warehouse_files() -> None:
    for set_code, gs_code, rows, code_lists, spec_name, sources, outputs in WAREHOUSE_SETS:
        meta = {
            "Transaction Set": set_code,
            "X12 Version": "004010",
            "Spec Name": spec_name,
            "Author": "EDI MapCheck project",
            "Date": "2026-07-05",
        }
        generate_spec(
            EXAMPLES / "specs" / f"{set_code}_reference_spec.xlsx",
            rows, code_lists, meta,
        )
        for name, control, segments in sources:
            path = EXAMPLES / "source" / name
            path.write_text(build_interchange(segments, control, set_code, gs_code))
            print(f"wrote {path.relative_to(REPO_ROOT)}")
        for name, data in outputs.items():
            path = EXAMPLES / "output" / name
            path.write_text(json.dumps(data, indent=2) + "\n")
            print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Inventory & product movement: 846 / 812 / 867
# --------------------------------------------------------------------------

SPEC846_ROWS: list[tuple[str, ...]] = [
    ("Q-001", "BIA01", "", "inquiry.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("Q-002", "BIA02", "", "inquiry.report_type", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", "Passed through; partner-specific values."),
    ("Q-003", "BIA03", "", "inquiry.reference_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("Q-004", "BIA04", "", "inquiry.report_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("Q-005", "", "", "inquiry.record_type", "CONSTANT", "", "", "", "",
     "INV_ADVICE", "", "string", "", ""),
    ("Q-006", "N102", "N1[SE]", "seller.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("Q-007", "N104", "N1[SE]", "seller.id", "CONDITIONAL",
     "Map the seller number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("Q-008", "LIN01", "LIN", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("Q-009", "LIN03", "LIN", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "LIN02 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("Q-010", "PID05", "LIN", "lines[].description", "CONDITIONAL",
     "Map the free-form description when the PID is free-form.",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..80", ""),
    ("Q-011", "QTY02", "LIN>QTY[QA]", "lines[].qty_available", "DIRECT",
     "", "", "", "", "", "", "integer", "",
     "Path context: the QTY whose qualifier is QA (available)."),
    ("Q-012", "QTY03", "LIN>QTY[QA]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", "UOM sent on the available-quantity segment."),
    ("Q-013", "QTY02", "LIN>QTY[QO]", "lines[].qty_on_order", "DIRECT",
     "", "", "", "", "", "", "integer", "", "QO = on order."),
    ("Q-014", "QTY02", "LIN>QTY[QC]", "lines[].qty_committed", "DIRECT",
     "", "", "", "", "", "", "integer", "", "QC = committed."),
    ("Q-015", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("Q-016", "LIN", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_846_ROWS = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
    *_WH_UOM,
]

BASELINE_846 = [
    "BIA*00*DD*INV20260801*20260801",
    "N1*SE*SUMMIT WHOLESALE FOODS*92*7731",
    "LIN*1*UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "QTY*QA*500*EA",
    "QTY*QO*200",
    "QTY*QC*50",
    "LIN*2*UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "QTY*QA*1200*EA",
    "QTY*QO*0",
    "QTY*QC*75",
    "CTT*2",
]

# Defects: invalid report date, unknown UOM on the available bucket, a
# mystery ZZ quantity bucket no rule references, a line missing its
# committed quantity, and a CTT count lie.
DEFECTS_846 = [
    "BIA*00*DD*INV20260802*20260835",
    "N1*SE*SUMMIT WHOLESALE FOODS*92*7731",
    "LIN*1*UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "QTY*QA*500*ZZ",
    "QTY*QO*200",
    "QTY*ZZ*10",
    "LIN*2*UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "QTY*QA*1200*EA",
    "QTY*QO*0",
    "QTY*QC*75",
    "CTT*5",
]


def _inv_line(line_no: int, upc: str, description: str, qa: int, qo: int, qc: int | None) -> dict:
    line = {"line_no": line_no, "upc": upc, "description": description,
            "qty_available": qa, "uom": "EACH", "qty_on_order": qo}
    if qc is not None:
        line["qty_committed"] = qc
    return line


BASELINE_846_OUTPUT: dict = {
    "inquiry": {"purpose": "ORIGINAL", "report_type": "DD",
                "reference_number": "INV20260801", "report_date": "2026-08-01",
                "record_type": "INV_ADVICE"},
    "seller": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _inv_line(1, "614141007349", "TRAIL MIX 12OZ", 500, 200, 50),
        _inv_line(2, "614141007350", "SPRING WATER 24PK", 1200, 0, 75),
    ],
    "summary": {"line_count": 2},
}

DEFECTS_846_OUTPUT: dict = {
    "inquiry": {"purpose": "ORIGINAL", "report_type": "DD",
                "reference_number": "INV20260802", "report_date": "2026-08-35",
                "record_type": "INV_ADVICE"},
    "seller": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        _inv_line(1, "614141007349", "TRAIL MIX 12OZ", 500, 200, None),
        _inv_line(2, "614141007350", "SPRING WATER 24PK", 1200, 0, 75),
    ],
    "summary": {"line_count": 5},
}

SPEC812_ROWS: list[tuple[str, ...]] = [
    ("C-001", "BCD01", "", "adjustment.adjustment_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("C-002", "BCD02", "", "adjustment.adjustment_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", ""),
    ("C-003", "BCD03", "", "adjustment.handling_code", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..2", "Passed through; partner-specific values."),
    ("C-004", "BCD04", "", "adjustment.total_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "Signed net total (credits negative)."),
    ("C-005", "BCD05", "", "adjustment.cd_flag", "CODE_LIST", "", "", "", "",
     "", "CD_FLAG", "string", "", ""),
    ("C-006", "", "", "adjustment.record_type", "CONSTANT", "", "", "", "",
     "ADJ_INBOUND", "", "string", "", ""),
    ("C-007", "N102", "N1[VN]", "vendor.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("C-008", "N104", "N1[VN]", "vendor.id", "CONDITIONAL",
     "Map the vendor number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("C-009", "CDD01", "CDD", "lines[].reason", "CODE_LIST", "", "", "", "",
     "", "ADJ_REASON_812", "string", "", ""),
    ("C-010", "CDD02", "CDD", "lines[].cd_flag", "CODE_LIST", "", "", "", "",
     "", "CD_FLAG", "string", "", ""),
    ("C-011", "CDD03", "CDD", "lines[].line_id", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("C-012", "CDD04", "CDD", "lines[].amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "Signed (credits negative, debits positive)."),
    ("C-013", "CDD", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_812_ROWS = [
    ("CD_FLAG", "C", "CREDIT", "Credit"),
    ("CD_FLAG", "D", "DEBIT", "Debit"),
    ("ADJ_REASON_812", "01", "PRICING_ERROR", "Pricing error"),
    ("ADJ_REASON_812", "06", "QUANTITY_CONTESTED", "Quantity contested"),
    ("ADJ_REASON_812", "59", "ITEM_NOT_RECEIVED", "Item not received"),
]

# Signed math: -25.00 credit + 40.00 debit = 15.00 net debit.
BASELINE_812 = [
    "BCD*20260805*CDA9001*I*15.00*D",
    "N1*VN*SUMMIT WHOLESALE FOODS*92*7731",
    "CDD*01*C*1*-25.00",
    "CDD*06*D*2*40.00",
]

# Defects: the credit line was keyed positive while the header still
# carries the intended net total, so the recon gap (65.00 vs 15.00) is
# exactly twice the mis-signed credit — the classic backwards-key
# signature; unknown reason code 99; invalid date; unreferenced ITD.
DEFECTS_812 = [
    "BCD*20260835*CDA9002*I*15.00*D",
    "ITD*01*3*2**30**60",
    "N1*VN*SUMMIT WHOLESALE FOODS*92*7731",
    "CDD*99*C*1*25.00",
    "CDD*06*D*2*40.00",
]

BASELINE_812_OUTPUT: dict = {
    "adjustment": {"adjustment_date": "2026-08-05", "adjustment_number": "CDA9001",
                   "handling_code": "I", "total_amount": 15.0, "cd_flag": "DEBIT",
                   "record_type": "ADJ_INBOUND"},
    "vendor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        {"reason": "PRICING_ERROR", "cd_flag": "CREDIT", "line_id": "1", "amount": -25.0},
        {"reason": "QUANTITY_CONTESTED", "cd_flag": "DEBIT", "line_id": "2", "amount": 40.0},
    ],
    "summary": {"line_count": 2},
}

DEFECTS_812_OUTPUT: dict = {
    "adjustment": {"adjustment_date": "2026-08-35", "adjustment_number": "CDA9002",
                   "handling_code": "I", "total_amount": 15.0, "cd_flag": "DEBIT",
                   "record_type": "ADJ_INBOUND"},
    "vendor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        {"reason": "99", "cd_flag": "CREDIT", "line_id": "1", "amount": 25.0},
        {"reason": "QUANTITY_CONTESTED", "cd_flag": "DEBIT", "line_id": "2", "amount": 40.0},
    ],
    "summary": {"line_count": 2},
}

SPEC867_ROWS: list[tuple[str, ...]] = [
    ("P-001", "BPT01", "", "report.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("P-002", "BPT02", "", "report.report_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("P-003", "BPT03", "", "report.report_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("P-004", "BPT04", "", "report.report_type", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", "Passed through; partner-specific values."),
    ("P-005", "QTY02", "QTY[TO]", "report.total_quantity", "CONDITIONAL",
     "Map the header total quantity when the partner sends one.",
     "EXISTS(QTY02)", "SOURCE", "SKIP", "", "", "integer", "", ""),
    ("P-006", "", "", "report.record_type", "CONSTANT", "", "", "", "",
     "XFER_REPORT", "", "string", "", ""),
    ("P-007", "N102", "N1[DS]", "distributor.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("P-008", "N104", "N1[DS]", "distributor.id", "CONDITIONAL",
     "Map the distributor number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("P-009", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("P-010", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the ship-to number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("P-011", "PTD01", "PTD", "lines[].transfer_type", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", "Passed through; partner-specific values."),
    ("P-012", "LIN03", "PTD", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "LIN02 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("P-013", "PID05", "PTD", "lines[].description", "CONDITIONAL",
     "Map the free-form description when the PID is free-form.",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..80", ""),
    ("P-014", "QTY02", "PTD>QTY[38]", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Assumed quantity qualifier 38; amend per guide."),
    ("P-015", "QTY03", "PTD>QTY[38]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("P-016", "CTP03", "PTD", "lines[].resale_price", "CONDITIONAL",
     "Map the resale price when the price qualifier is RES.",
     "CTP02 = 'RES'", "SOURCE", "SKIP", "", "", "decimal", "places:2", ""),
    ("P-017", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("P-018", "PTD", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_867_ROWS = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
    *_WH_UOM,
]

BASELINE_867 = [
    "BPT*00*RPT20260810*20260810*DD",
    "QTY*TO*175",
    "N1*DS*SUMMIT WHOLESALE FOODS*92*7731",
    "N1*ST*RIVERBEND MARKET*92*0007",
    "PTD*BD",
    "LIN**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "QTY*38*100*EA",
    "CTP**RES*8.99",
    "PTD*BD",
    "LIN**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "QTY*38*75*EA",
    "CTP**RES*12.49",
    "CTT*2",
]

# Defects: header total quantity lies (200 vs 175 line sum), CTT lies,
# unknown UOM on line 2, invalid report date, an unreferenced header REF.
DEFECTS_867 = [
    "BPT*00*RPT20260811*20260835*DD",
    "QTY*TO*200",
    "REF*ZZ*MYSTERY",
    "N1*DS*SUMMIT WHOLESALE FOODS*92*7731",
    "N1*ST*RIVERBEND MARKET*92*0007",
    "PTD*BD",
    "LIN**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "QTY*38*100*EA",
    "CTP**RES*8.99",
    "PTD*BD",
    "LIN**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "QTY*38*75*ZZ",
    "CTP**RES*12.49",
    "CTT*3",
]


def _xfer_report_line(upc: str, description: str, qty: int, price: float) -> dict:
    return {"transfer_type": "BD", "upc": upc, "description": description,
            "qty": qty, "uom": "EACH", "resale_price": price}


BASELINE_867_OUTPUT: dict = {
    "report": {"purpose": "ORIGINAL", "report_number": "RPT20260810",
               "report_date": "2026-08-10", "report_type": "DD",
               "total_quantity": 175, "record_type": "XFER_REPORT"},
    "distributor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "ship_to": {"name": "RIVERBEND MARKET", "id": "0007"},
    "lines": [
        _xfer_report_line("614141007349", "TRAIL MIX 12OZ", 100, 8.99),
        _xfer_report_line("614141007350", "SPRING WATER 24PK", 75, 12.49),
    ],
    "summary": {"line_count": 2},
}

DEFECTS_867_OUTPUT: dict = json.loads(json.dumps(BASELINE_867_OUTPUT))
DEFECTS_867_OUTPUT["report"].update(
    {"report_number": "RPT20260811", "report_date": "2026-08-35", "total_quantity": 200}
)
DEFECTS_867_OUTPUT["lines"][1]["uom"] = "ZZ"
DEFECTS_867_OUTPUT["summary"]["line_count"] = 3

INVENTORY_SETS = [
    ("846", "IB", SPEC846_ROWS, CODE_LIST_846_ROWS,
     "Synthetic inventory advice - reference example",
     [("846_baseline.edi", "41", BASELINE_846), ("846_defects.edi", "42", DEFECTS_846)],
     {"invinquiry_baseline.json": BASELINE_846_OUTPUT,
      "invinquiry_defects.json": DEFECTS_846_OUTPUT}),
    ("812", "CD", SPEC812_ROWS, CODE_LIST_812_ROWS,
     "Synthetic credit/debit adjustment - reference example",
     [("812_baseline.edi", "43", BASELINE_812), ("812_defects.edi", "44", DEFECTS_812)],
     {"creditdebit_baseline.json": BASELINE_812_OUTPUT,
      "creditdebit_defects.json": DEFECTS_812_OUTPUT}),
    ("867", "PT", SPEC867_ROWS, CODE_LIST_867_ROWS,
     "Synthetic product transfer report - reference example",
     [("867_baseline.edi", "45", BASELINE_867), ("867_defects.edi", "46", DEFECTS_867)],
     {"xferreport_baseline.json": BASELINE_867_OUTPUT,
      "xferreport_defects.json": DEFECTS_867_OUTPUT}),
]


def generate_inventory_files() -> None:
    for set_code, gs_code, rows, code_lists, spec_name, sources, outputs in INVENTORY_SETS:
        meta = {
            "Transaction Set": set_code,
            "X12 Version": "004010",
            "Spec Name": spec_name,
            "Author": "EDI MapCheck project",
            "Date": "2026-07-06",
        }
        generate_spec(
            EXAMPLES / "specs" / f"{set_code}_reference_spec.xlsx",
            rows, code_lists, meta,
        )
        for name, control, segments in sources:
            path = EXAMPLES / "source" / name
            path.write_text(build_interchange(segments, control, set_code, gs_code))
            print(f"wrote {path.relative_to(REPO_ROOT)}")
        for name, data in outputs.items():
            path = EXAMPLES / "output" / name
            path.write_text(json.dumps(data, indent=2) + "\n")
            print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Pharma contract & chargeback set: 844 / 845 / 849 / 854
# --------------------------------------------------------------------------

_PHARMA_PURPOSE = [
    ("TX_PURPOSE", "00", "ORIGINAL", "Original transmission"),
    ("TX_PURPOSE", "01", "CANCELLATION", "Cancellation"),
    ("TX_PURPOSE", "05", "REPLACE", "Replacement"),
]

SPEC844_ROWS: list[tuple[str, ...]] = [
    ("G-001", "BGN01", "", "request.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("G-002", "BGN02", "", "request.request_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", "Chargeback request number."),
    ("G-003", "BGN03", "", "request.request_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("G-004", "REF02", "REF[CT]", "request.contract_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", "The contract the chargeback claims against."),
    ("G-005", "AMT02", "AMT[TT]", "request.total_debit", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "Total debit requested."),
    ("G-006", "", "", "request.record_type", "CONSTANT", "", "", "", "",
     "CHARGEBACK_REQUEST", "", "string", "", ""),
    ("G-007", "N102", "N1[MF]", "manufacturer.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("G-008", "N104", "N1[MF]", "manufacturer.id", "CONDITIONAL",
     "Map the manufacturer number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("G-009", "N102", "N1[DS]", "distributor.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("G-010", "N104", "N1[DS]", "distributor.id", "CONDITIONAL",
     "Map the distributor number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("G-011", "LIN03", "LIN", "lines[].ndc", "CONDITIONAL",
     "Map the NDC only when the product ID qualifier is ND.",
     "LIN02 = 'ND'", "SOURCE", "SKIP", "", "", "string", "len:10..11", ""),
    ("G-012", "QTY02", "LIN>QTY[38]", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Units the chargeback covers."),
    ("G-013", "QTY03", "LIN>QTY[38]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("G-014", "CTP03", "LIN>CTP[WS]", "lines[].wac", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2",
     "Acquisition (WAC) price; class-of-trade WS is the synthetic discriminator."),
    ("G-015", "CTP03", "LIN>CTP[CT]", "lines[].contract_price", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "Contract price; class-of-trade CT."),
    ("G-016", "AMT02", "LIN>AMT[1]", "lines[].debit_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2",
     "Line debit. (wac - contract) x qty arithmetic is a backlog check."),
    ("G-017", "LIN", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_844_ROWS = [*_PHARMA_PURPOSE, *_WH_UOM]

# Chargeback math: (5.85-4.10)x40 = 70.00; (12.30-9.80)x100 = 250.00; total 320.00
BASELINE_844 = [
    "BGN*00*CBR20260815001*20260815",
    "REF*CT*CTR-2026-0142",
    "AMT*TT*320.00",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "N1*DS*GRANITE PHARMA DISTRIBUTION*92*D208",
    "LIN**ND*00777310502",
    "QTY*38*40*EA",
    "CTP*WS**5.85",
    "CTP*CT**4.10",
    "AMT*1*70.00",
    "LIN**ND*00777310617",
    "QTY*38*100*EA",
    "CTP*WS**12.30",
    "CTP*CT**9.80",
    "AMT*1*250.00",
]

# Defects: the request total claims 350.00 against lines summing 320.00,
# the contract number REF is missing entirely (while the naive output
# still carries a stale one), unknown UOM, invalid request date, and an
# unreferenced PER contact.
DEFECTS_844 = [
    "BGN*00*CBR20260815002*20260845",
    "AMT*TT*350.00",
    "PER*IC*JANE DOE",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "N1*DS*GRANITE PHARMA DISTRIBUTION*92*D208",
    "LIN**ND*00777310502",
    "QTY*38*40*ZZ",
    "CTP*WS**5.85",
    "CTP*CT**4.10",
    "AMT*1*70.00",
    "LIN**ND*00777310617",
    "QTY*38*100*EA",
    "CTP*WS**12.30",
    "CTP*CT**9.80",
    "AMT*1*250.00",
]


def _cbk_line(ndc: str, qty: int, wac: float, contract: float, debit: float) -> dict:
    return {"ndc": ndc, "qty": qty, "uom": "EACH", "wac": wac,
            "contract_price": contract, "debit_amount": debit}


BASELINE_844_OUTPUT: dict = {
    "request": {"purpose": "ORIGINAL", "request_number": "CBR20260815001",
                "request_date": "2026-08-15", "contract_number": "CTR-2026-0142",
                "total_debit": 320.0, "record_type": "CHARGEBACK_REQUEST"},
    "manufacturer": {"name": "HELVETICA PHARMA", "id": "M501"},
    "distributor": {"name": "GRANITE PHARMA DISTRIBUTION", "id": "D208"},
    "lines": [
        _cbk_line("00777310502", 40, 5.85, 4.10, 70.0),
        _cbk_line("00777310617", 100, 12.30, 9.80, 250.0),
    ],
    "summary": {"line_count": 2},
}

DEFECTS_844_OUTPUT: dict = json.loads(json.dumps(BASELINE_844_OUTPUT))
DEFECTS_844_OUTPUT["request"].update(
    {"request_number": "CBR20260815002", "request_date": "2026-08-45",
     "total_debit": 350.0}
)  # contract_number kept although the source never sent one -> unexpected
DEFECTS_844_OUTPUT["lines"][0]["uom"] = "ZZ"

SPEC845_ROWS: list[tuple[str, ...]] = [
    ("H-001", "BGN01", "", "auth.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("H-002", "BGN02", "", "auth.authorization_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("H-003", "BGN03", "", "auth.authorization_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("H-004", "REF02", "REF[CT]", "auth.contract_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("H-005", "DTM02", "DTM[007]", "auth.effective_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "Start of the authorization window."),
    ("H-006", "DTM02", "DTM[036]", "auth.expiration_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "End of the authorization window."),
    ("H-007", "", "", "auth.record_type", "CONSTANT", "", "", "", "",
     "PRICE_AUTH", "", "string", "", ""),
    ("H-008", "N102", "N1[MF]", "manufacturer.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("H-009", "N104", "N1[MF]", "manufacturer.id", "CONDITIONAL",
     "Map the manufacturer number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("H-010", "N102", "N1[DS]", "distributor.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("H-011", "N104", "N1[DS]", "distributor.id", "CONDITIONAL",
     "Map the distributor number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("H-012", "LIN03", "LIN", "lines[].ndc", "CONDITIONAL",
     "Map the NDC only when the product ID qualifier is ND.",
     "LIN02 = 'ND'", "SOURCE", "SKIP", "", "", "string", "len:10..11", ""),
    ("H-013", "CTP03", "LIN>CTP[CT]", "lines[].authorized_price", "DIRECT",
     "", "", "", "", "", "", "decimal", "places:2", ""),
    ("H-014", "LIN", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_845_ROWS = [*_PHARMA_PURPOSE]

BASELINE_845 = [
    "BGN*00*PA20260701088*20260701",
    "REF*CT*CTR-2026-0142",
    "DTM*007*20260701",
    "DTM*036*20261231",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "N1*DS*GRANITE PHARMA DISTRIBUTION*92*D208",
    "LIN**ND*00777310502",
    "CTP*CT**4.10",
    "LIN**ND*00777310617",
    "CTP*CT**9.80",
]

# Defects: the authorization window is inverted (effective 2027-01-01,
# expires 2026-06-30 — the expired-authorization classic), invalid
# authorization date, a line missing its contract price, and an
# unreferenced PER contact.
DEFECTS_845 = [
    "BGN*00*PA20260701089*20260745",
    "REF*CT*CTR-2026-0142",
    "DTM*007*20270101",
    "DTM*036*20260630",
    "PER*IC*JANE DOE",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "N1*DS*GRANITE PHARMA DISTRIBUTION*92*D208",
    "LIN**ND*00777310502",
    "CTP*CT**4.10",
    "LIN**ND*00777310617",
]

BASELINE_845_OUTPUT: dict = {
    "auth": {"purpose": "ORIGINAL", "authorization_number": "PA20260701088",
             "authorization_date": "2026-07-01", "contract_number": "CTR-2026-0142",
             "effective_date": "2026-07-01", "expiration_date": "2026-12-31",
             "record_type": "PRICE_AUTH"},
    "manufacturer": {"name": "HELVETICA PHARMA", "id": "M501"},
    "distributor": {"name": "GRANITE PHARMA DISTRIBUTION", "id": "D208"},
    "lines": [
        {"ndc": "00777310502", "authorized_price": 4.10},
        {"ndc": "00777310617", "authorized_price": 9.80},
    ],
    "summary": {"line_count": 2},
}

DEFECTS_845_OUTPUT: dict = {
    "auth": {"purpose": "ORIGINAL", "authorization_number": "PA20260701089",
             "authorization_date": "2026-07-45", "contract_number": "CTR-2026-0142",
             "effective_date": "2027-01-01", "expiration_date": "2026-06-30",
             "record_type": "PRICE_AUTH"},
    "manufacturer": {"name": "HELVETICA PHARMA", "id": "M501"},
    "distributor": {"name": "GRANITE PHARMA DISTRIBUTION", "id": "D208"},
    "lines": [
        {"ndc": "00777310502", "authorized_price": 4.10},
        {"ndc": "00777310617"},
    ],
    "summary": {"line_count": 2},
}

SPEC849_ROWS: list[tuple[str, ...]] = [
    ("E-001", "BGN01", "", "response.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("E-002", "BGN02", "", "response.response_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("E-003", "BGN03", "", "response.response_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("E-004", "REF02", "REF[CT]", "response.contract_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("E-005", "REF02", "REF[TN]", "response.original_request_number", "DIRECT",
     "", "", "", "", "", "", "string", "len:1..30",
     "Links back to the 844; cross-transaction pairing is a future feature."),
    ("E-006", "AMT02", "AMT[TT]", "response.total_approved", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", ""),
    ("E-007", "", "", "response.record_type", "CONSTANT", "", "", "", "",
     "CHARGEBACK_RESPONSE", "", "string", "", ""),
    ("E-008", "N102", "N1[MF]", "manufacturer.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("E-009", "N104", "N1[MF]", "manufacturer.id", "CONDITIONAL",
     "Map the manufacturer number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("E-010", "LIN03", "LIN", "lines[].ndc", "CONDITIONAL",
     "Map the NDC only when the product ID qualifier is ND.",
     "LIN02 = 'ND'", "SOURCE", "SKIP", "", "", "string", "len:10..11", ""),
    ("E-011", "QTY02", "LIN>QTY[38]", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("E-012", "QTY03", "LIN>QTY[38]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("E-013", "AMT02", "LIN>AMT[1]", "lines[].approved_amount", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", ""),
    ("E-014", "LQ02", "LIN>LQ[RS]", "lines[].line_status", "CODE_LIST", "", "", "", "",
     "", "RESP_STATUS", "string", "", "The LQ whose list qualifier is RS."),
    ("E-015", "LIN", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_849_ROWS = [
    *_PHARMA_PURPOSE,
    *_WH_UOM,
    ("RESP_STATUS", "AP", "APPROVED", "Approved in full"),
    ("RESP_STATUS", "PA", "PARTIAL_APPROVAL", "Partially approved"),
    ("RESP_STATUS", "DE", "DENIED", "Denied"),
    ("RESP_STATUS", "PR", "PENDING_REVIEW", "Pending review"),
]

BASELINE_849 = [
    "BGN*00*CBR20260815001R*20260822",
    "REF*CT*CTR-2026-0142",
    "REF*TN*CBR20260815001",
    "AMT*TT*270.00",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "LIN**ND*00777310502",
    "QTY*38*40*EA",
    "AMT*1*70.00",
    "LQ*RS*AP",
    "LIN**ND*00777310617",
    "QTY*38*100*EA",
    "AMT*1*200.00",
    "LQ*RS*PA",
]

# Defects: unknown response status ZZ, the approved total (300.00)
# disagrees with the line sum (270.00), invalid response date, and an
# unreferenced PER contact.
DEFECTS_849 = [
    "BGN*00*CBR20260815002R*20260845",
    "REF*CT*CTR-2026-0142",
    "REF*TN*CBR20260815002",
    "AMT*TT*300.00",
    "PER*IC*JANE DOE",
    "N1*MF*HELVETICA PHARMA*92*M501",
    "LIN**ND*00777310502",
    "QTY*38*40*EA",
    "AMT*1*70.00",
    "LQ*RS*ZZ",
    "LIN**ND*00777310617",
    "QTY*38*100*EA",
    "AMT*1*200.00",
    "LQ*RS*PA",
]

BASELINE_849_OUTPUT: dict = {
    "response": {"purpose": "ORIGINAL", "response_number": "CBR20260815001R",
                 "response_date": "2026-08-22", "contract_number": "CTR-2026-0142",
                 "original_request_number": "CBR20260815001",
                 "total_approved": 270.0, "record_type": "CHARGEBACK_RESPONSE"},
    "manufacturer": {"name": "HELVETICA PHARMA", "id": "M501"},
    "lines": [
        {"ndc": "00777310502", "qty": 40, "uom": "EACH",
         "approved_amount": 70.0, "line_status": "APPROVED"},
        {"ndc": "00777310617", "qty": 100, "uom": "EACH",
         "approved_amount": 200.0, "line_status": "PARTIAL_APPROVAL"},
    ],
    "summary": {"line_count": 2},
}

DEFECTS_849_OUTPUT: dict = json.loads(json.dumps(BASELINE_849_OUTPUT))
DEFECTS_849_OUTPUT["response"].update(
    {"response_number": "CBR20260815002R", "response_date": "2026-08-45",
     "original_request_number": "CBR20260815002", "total_approved": 300.0}
)
DEFECTS_849_OUTPUT["lines"][0]["line_status"] = "ZZ"

SPEC854_ROWS: list[tuple[str, ...]] = [
    ("F-001", "BGN01", "", "discrepancy.purpose", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE", "string", "", ""),
    ("F-002", "BGN02", "", "discrepancy.report_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("F-003", "BGN03", "", "discrepancy.report_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", ""),
    ("F-004", "REF02", "REF[BM]", "discrepancy.bol_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..30", ""),
    ("F-005", "", "", "discrepancy.record_type", "CONSTANT", "", "", "", "",
     "DISCREPANCY", "", "string", "", ""),
    ("F-006", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("F-007", "N104", "N1[ST]", "ship_to.id", "CONDITIONAL",
     "Map the location number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("F-008", "N102", "N1[CA]", "carrier.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", ""),
    ("F-009", "N104", "N1[CA]", "carrier.scac", "CONDITIONAL",
     "Map the SCAC only when the carrier ID qualifier is 2.",
     "N103 = '2'", "SOURCE", "SKIP", "", "", "string", "len:2..4", ""),
    ("F-010", "LIN03", "LIN", "lines[].upc", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "LIN02 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14", ""),
    ("F-011", "QTY02", "LIN>QTY[38]", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Discrepancy quantity."),
    ("F-012", "QTY03", "LIN>QTY[38]", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", ""),
    ("F-013", "LQ02", "LIN>LQ[DR]", "lines[].reason", "CODE_LIST", "", "", "", "",
     "", "DISCREPANCY", "string", "", "The LQ whose list qualifier is DR."),
    ("F-014", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("F-015", "LIN", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_854_ROWS = [
    *_PHARMA_PURPOSE,
    *_WH_UOM,
    ("DISCREPANCY", "SH", "SHORTAGE", "Quantity short"),
    ("DISCREPANCY", "OV", "OVERAGE", "Quantity over"),
    ("DISCREPANCY", "DA", "DAMAGED", "Damaged"),
    ("DISCREPANCY", "WI", "WRONG_ITEM", "Wrong item"),
    ("DISCREPANCY", "NO", "NOT_ORDERED", "Not ordered"),
    ("DISCREPANCY", "LT", "LATE_DELIVERY", "Late delivery"),
]

BASELINE_854 = [
    "BGN*00*DSC20260818004*20260818",
    "REF*BM*BOL8802214",
    "N1*ST*ALPINE OUTFITTERS DC 12*92*0012",
    "N1*CA*ROADWAY EXPRESS*2*RDWY",
    "LIN**UP*614141007349",
    "QTY*38*4*EA",
    "LQ*DR*SH",
    "LIN**UP*614141007350",
    "QTY*38*2*EA",
    "LQ*DR*DA",
    "CTT*2",
]

# Defects: unknown discrepancy reason XX, CTT count lie (4 vs 2 lines),
# invalid report date, and an unreferenced PER contact.
DEFECTS_854 = [
    "BGN*00*DSC20260818005*20260845",
    "REF*BM*BOL8802290",
    "PER*IC*JANE DOE",
    "N1*ST*ALPINE OUTFITTERS DC 12*92*0012",
    "N1*CA*ROADWAY EXPRESS*2*RDWY",
    "LIN**UP*614141007349",
    "QTY*38*4*EA",
    "LQ*DR*XX",
    "LIN**UP*614141007350",
    "QTY*38*2*EA",
    "LQ*DR*DA",
    "CTT*4",
]

BASELINE_854_OUTPUT: dict = {
    "discrepancy": {"purpose": "ORIGINAL", "report_number": "DSC20260818004",
                    "report_date": "2026-08-18", "bol_number": "BOL8802214",
                    "record_type": "DISCREPANCY"},
    "ship_to": {"name": "ALPINE OUTFITTERS DC 12", "id": "0012"},
    "carrier": {"name": "ROADWAY EXPRESS", "scac": "RDWY"},
    "lines": [
        {"upc": "614141007349", "qty": 4, "uom": "EACH", "reason": "SHORTAGE"},
        {"upc": "614141007350", "qty": 2, "uom": "EACH", "reason": "DAMAGED"},
    ],
    "summary": {"line_count": 2},
}

DEFECTS_854_OUTPUT: dict = json.loads(json.dumps(BASELINE_854_OUTPUT))
DEFECTS_854_OUTPUT["discrepancy"].update(
    {"report_number": "DSC20260818005", "report_date": "2026-08-45",
     "bol_number": "BOL8802290"}
)
DEFECTS_854_OUTPUT["lines"][0]["reason"] = "XX"
DEFECTS_854_OUTPUT["summary"]["line_count"] = 4

PHARMA_SETS = [
    ("844", "CF", SPEC844_ROWS, CODE_LIST_844_ROWS,
     "Synthetic chargeback request - reference example",
     [("844_baseline.edi", "51", BASELINE_844), ("844_defects.edi", "52", DEFECTS_844)],
     {"chargeback_baseline.json": BASELINE_844_OUTPUT,
      "chargeback_defects.json": DEFECTS_844_OUTPUT}),
    ("845", "PA", SPEC845_ROWS, CODE_LIST_845_ROWS,
     "Synthetic price authorization - reference example",
     [("845_baseline.edi", "53", BASELINE_845), ("845_defects.edi", "54", DEFECTS_845)],
     {"priceauth_baseline.json": BASELINE_845_OUTPUT,
      "priceauth_defects.json": DEFECTS_845_OUTPUT}),
    ("849", "CF", SPEC849_ROWS, CODE_LIST_849_ROWS,
     "Synthetic chargeback response - reference example",
     [("849_baseline.edi", "55", BASELINE_849), ("849_defects.edi", "56", DEFECTS_849)],
     {"cbresponse_baseline.json": BASELINE_849_OUTPUT,
      "cbresponse_defects.json": DEFECTS_849_OUTPUT}),
    ("854", "AB", SPEC854_ROWS, CODE_LIST_854_ROWS,
     "Synthetic delivery discrepancy - reference example",
     [("854_baseline.edi", "57", BASELINE_854), ("854_defects.edi", "58", DEFECTS_854)],
     {"discrepancy_baseline.json": BASELINE_854_OUTPUT,
      "discrepancy_defects.json": DEFECTS_854_OUTPUT}),
]


def generate_pharma_files() -> None:
    for set_code, gs_code, rows, code_lists, spec_name, sources, outputs in PHARMA_SETS:
        meta = {
            "Transaction Set": set_code,
            "X12 Version": "004010",
            "Spec Name": spec_name,
            "Author": "EDI MapCheck project",
            "Date": "2026-07-06",
        }
        generate_spec(
            EXAMPLES / "specs" / f"{set_code}_reference_spec.xlsx",
            rows, code_lists, meta,
        )
        for name, control, segments in sources:
            path = EXAMPLES / "source" / name
            path.write_text(build_interchange(segments, control, set_code, gs_code))
            print(f"wrote {path.relative_to(REPO_ROOT)}")
        for name, data in outputs.items():
            path = EXAMPLES / "output" / name
            path.write_text(json.dumps(data, indent=2) + "\n")
            print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# 997 Functional Acknowledgment
# --------------------------------------------------------------------------

SPEC997_ROWS: list[tuple[str, ...]] = [
    ("Z-001", "AK101", "", "ack.functional_group", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..2", "Functional ID of the acknowledged group."),
    ("Z-002", "AK102", "", "ack.group_control", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Group control number being acknowledged."),
    ("Z-003", "AK901", "", "ack.group_status", "CODE_LIST", "", "", "", "",
     "", "ACK_CODE", "string", "", ""),
    ("Z-004", "AK902", "", "ack.sets_included", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("Z-005", "AK903", "", "ack.sets_received", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("Z-006", "AK904", "", "ack.sets_accepted", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("Z-007", "", "", "ack.record_type", "CONSTANT", "", "", "", "",
     "FUNC_ACK", "", "string", "", ""),
    ("Z-008", "AK201", "AK2", "lines[].transaction_set", "DIRECT", "", "", "", "",
     "", "", "string", "len:3..3", ""),
    ("Z-009", "AK202", "AK2", "lines[].control_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:4..9", "String: control numbers keep leading zeros."),
    ("Z-010", "AK501", "AK2", "lines[].status", "CODE_LIST", "", "", "", "",
     "", "ACK_CODE", "string", "", ""),
    ("Z-011", "AK2", "", "summary.ack_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
]

CODE_LIST_997_ROWS = [
    ("ACK_CODE", "A", "ACCEPTED", "Accepted"),
    ("ACK_CODE", "E", "ACCEPTED_WITH_ERRORS", "Accepted with errors noted"),
    ("ACK_CODE", "P", "PARTIALLY_ACCEPTED", "Partially accepted"),
    ("ACK_CODE", "R", "REJECTED", "Rejected"),
]

SPEC997_META = {
    "Transaction Set": "997",
    "X12 Version": "004010",
    "Spec Name": "Synthetic functional acknowledgment - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-06",
}

BASELINE_997 = [
    "AK1*PO*1044",
    "AK2*850*0001",
    "AK5*A",
    "AK2*850*0002",
    "AK5*R",
    "AK9*P*2*2*1",
]

# Defects: unknown acknowledgment code Z, AK9 claims 3 accepted of 2
# received (impossible), and AK3/AK4 error detail no spec rule maps.
DEFECTS_997 = [
    "AK1*PO*1045",
    "AK2*850*0003",
    "AK3*REF*3**8",
    "AK4*2**1",
    "AK5*Z",
    "AK2*850*0004",
    "AK5*A",
    "AK9*E*2*2*3",
]

BASELINE_997_OUTPUT: dict = {
    "ack": {"functional_group": "PO", "group_control": 1044,
            "group_status": "PARTIALLY_ACCEPTED", "sets_included": 2,
            "sets_received": 2, "sets_accepted": 1, "record_type": "FUNC_ACK"},
    "lines": [
        {"transaction_set": "850", "control_number": "0001", "status": "ACCEPTED"},
        {"transaction_set": "850", "control_number": "0002", "status": "REJECTED"},
    ],
    "summary": {"ack_count": 2},
}

DEFECTS_997_OUTPUT: dict = {
    "ack": {"functional_group": "PO", "group_control": 1045,
            "group_status": "ACCEPTED_WITH_ERRORS", "sets_included": 2,
            "sets_received": 2, "sets_accepted": 3, "record_type": "FUNC_ACK"},
    "lines": [
        {"transaction_set": "850", "control_number": "0003", "status": "Z"},
        {"transaction_set": "850", "control_number": "0004", "status": "ACCEPTED"},
    ],
    "summary": {"ack_count": 2},
}


def generate_997_files() -> None:
    generate_spec(
        EXAMPLES / "specs" / "997_reference_spec.xlsx",
        SPEC997_ROWS, CODE_LIST_997_ROWS, SPEC997_META,
    )
    sources = {
        "997_baseline.edi": build_interchange(BASELINE_997, "61", "997", "FA"),
        "997_defects.edi": build_interchange(DEFECTS_997, "62", "997", "FA"),
    }
    for name, content in sources.items():
        path = EXAMPLES / "source" / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")
    outputs = {
        "funcack_baseline.json": BASELINE_997_OUTPUT,
        "funcack_defects.json": DEFECTS_997_OUTPUT,
    }
    for name, data in outputs.items():
        path = EXAMPLES / "output" / name
        path.write_text(json.dumps(data, indent=2) + "\n")
        print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# 850-to-ORDERS05 (SAP IDoc) scenario: one spec, outputs in both IDoc formats
# --------------------------------------------------------------------------

SPEC_ORDERS05_ROWS: list[tuple[str, ...]] = [
    ("I-001", "BEG01", "", "header.action", "CODE_LIST", "", "", "", "",
     "", "SAP_ACTION", "string", "", "IDoc action code (synthetic convention)."),
    ("I-002", "BEG02", "", "header.bsart", "CODE_LIST", "", "", "", "",
     "", "PO_TYPE_SAP", "string", "", "SAP document type."),
    ("I-003", "BEG03", "", "refs.001.belnr", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..35", "E1EDK02 qualifier 001 = customer PO number."),
    ("I-004", "BEG05", "", "refs.001.datum", "DIRECT", "", "", "", "",
     "", "", "date", "%Y%m%d", "SAP dates are YYYYMMDD."),
    ("I-005", "BEG05", "", "dates.012", "DIRECT", "", "", "", "",
     "", "", "date", "%Y%m%d", "E1EDK03 IDDAT 012 = document date."),
    ("I-006", "CUR02", "", "header.curcy", "CONDITIONAL",
     "Map the currency only when a buying-party currency code is sent.",
     "CUR01 = 'BY'", "SOURCE", "SKIP", "", "", "string", "len:3..3", ""),
    ("I-007", "DTM02", "DTM[002]", "dates.002", "DIRECT", "", "", "", "",
     "", "", "date", "%Y%m%d", "IDDAT 002 = requested delivery."),
    ("I-008", "", "", "org.012", "CONSTANT", "", "", "", "",
     "NB", "", "string", "", "Hardcoded order type (E1EDK14 qualifier 012)."),
    ("I-009", "N102", "N1[ST]", "partners.we.name1", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..35", "WE = ship-to party."),
    ("I-010", "N104", "N1[ST]", "partners.we.partn", "CONDITIONAL",
     "Map the store number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("I-011", "N301", "N1[ST]", "partners.we.stras", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("I-012", "N401", "N1[ST]", "partners.we.ort01", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("I-013", "N402", "N1[ST]", "partners.we.regio", "DIRECT", "", "", "", "",
     "", "", "string", "len:2..3", ""),
    ("I-014", "N403", "N1[ST]", "partners.we.pstlz", "DIRECT", "", "", "", "",
     "", "", "string", "len:5..9", ""),
    ("I-015", "N102", "N1[BT]", "partners.ag.name1", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..35", "AG = sold-to party."),
    ("I-016", "N104", "N1[BT]", "partners.ag.partn", "CONDITIONAL",
     "Map the sold-to number only when the ID qualifier is 92.",
     "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("I-017", "PO101", "PO1", "lines[].refs.001.zeile", "DIRECT", "", "", "", "",
     "", "", "integer", "", "E1EDP02 qualifier 001 = customer PO line."),
    ("I-018", "PO102", "PO1", "lines[].menge", "DIRECT", "", "", "", "",
     "", "", "decimal", "", "SAP quantities carry decimals (12.000)."),
    ("I-019", "PO103", "PO1", "lines[].menee", "CODE_LIST", "", "", "", "",
     "", "UOM_SAP", "string", "", "X12 UOM to SAP unit."),
    ("I-020", "PO104", "PO1", "lines[].vprei", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", ""),
    ("I-021", "PO107", "PO1", "lines[].ids.003.idtnr", "CONDITIONAL",
     "Map the UPC only when the product ID qualifier is UP.",
     "PO106 = 'UP'", "SOURCE", "SKIP", "", "", "string", "len:12..14",
     "E1EDP19 qualifier 003 = EAN/UPC."),
    ("I-022", "PID05", "PO1", "lines[].ids.003.ktext", "CONDITIONAL",
     "Map the free-form description when the PID is free-form.",
     "PID01 = 'F'", "SOURCE", "SKIP", "", "", "string", "len:1..70", ""),
    ("I-023", "CTT01", "", "summary.001", "DIRECT", "", "", "", "",
     "", "", "integer", "", "E1EDS01 SUMID 001 = item count."),
    ("I-024", "PO1", "", "summary.001", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", ""),
    ("I-025", "AMT02", "AMT[TT]", "summary.002", "DIRECT", "", "", "", "",
     "", "", "decimal", "places:2", "E1EDS01 SUMID 002 = net value."),
]

CODE_LIST_ORDERS05_ROWS = [
    ("SAP_ACTION", "00", "000", "Original (synthetic convention)"),
    ("SAP_ACTION", "01", "001", "Cancellation (synthetic convention)"),
    ("SAP_ACTION", "05", "002", "Replacement (synthetic convention)"),
    ("PO_TYPE_SAP", "SA", "NB", "Stand-alone -> standard order"),
    ("PO_TYPE_SAP", "NE", "ZNB", "New order -> custom order type"),
    ("UOM_SAP", "EA", "ST", "Each -> Stueck"),
    ("UOM_SAP", "CA", "KAR", "Case -> Karton"),
    ("UOM_SAP", "DZ", "DZN", "Dozen"),
]

SPEC_ORDERS05_META = {
    "Transaction Set": "850",
    "X12 Version": "004010",
    "Spec Name": "Synthetic 850-to-ORDERS05 map - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-07",
}

SAP_850 = [
    "BEG*00*SA*PO5500077**20260701",
    "CUR*BY*USD",
    "DTM*002*20260715",
    "N1*ST*ALPINE OUTFITTERS STORE 118*92*0118",
    "N3*4501 CASCADE AVE",
    "N4*BOULDER*CO*80301",
    "N1*BT*ALPINE OUTFITTERS CORPORATE*92*0001",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "PID*F****TRAIL MIX 12OZ",
    "PO1*2*6*CA*24**UP*614141007350",
    "PID*F****SPRING WATER 24PK",
    "PO1*3*5*DZ*12**UP*614141007351",
    "PID*F****GRANOLA BARS VARIETY",
    "CTT*3",
    "AMT*TT*306",
]

# One scenario structure drives BOTH IDoc writers: (segment, {FIELD: value}).
# POSEX and E1EDP20 appear in the files for realism; the adapter ignores
# them by design.


def _orders05_segments(
    po_number: str,
    po_date: str,
    currency: str | None,
    bsart: str | None,
    order_type: str | None,
    lines: list[dict],
    item_count: str,
    net_value: str,
    extra_org: tuple[str, str] | None = None,
) -> list[tuple[str, dict[str, str]]]:
    segments: list[tuple[str, dict[str, str]]] = []
    k01 = {"ACTION": "000"}
    if currency:
        k01["CURCY"] = currency
    if bsart:
        k01["BSART"] = bsart
    segments.append(("E1EDK01", k01))
    if order_type:
        segments.append(("E1EDK14", {"QUALF": "012", "ORGID": order_type}))
    if extra_org:
        segments.append(("E1EDK14", {"QUALF": extra_org[0], "ORGID": extra_org[1]}))
    segments.append(("E1EDK03", {"IDDAT": "012", "DATUM": po_date}))
    segments.append(("E1EDK03", {"IDDAT": "002", "DATUM": "20260715"}))
    segments.append(("E1EDK02", {"QUALF": "001", "BELNR": po_number, "DATUM": po_date}))
    segments.append(("E1EDKA1", {
        "PARVW": "WE", "PARTN": "0118", "NAME1": "ALPINE OUTFITTERS STORE 118",
        "STRAS": "4501 CASCADE AVE", "ORT01": "BOULDER", "PSTLZ": "80301",
        "REGIO": "CO",
    }))
    segments.append(("E1EDKA1", {
        "PARVW": "AG", "PARTN": "0001", "NAME1": "ALPINE OUTFITTERS CORPORATE",
    }))
    for index, line in enumerate(lines, start=1):
        segments.append(("E1EDP01", {
            "POSEX": f"{index * 10:06d}", "MENGE": line["menge"],
            "MENEE": line["menee"], "VPREI": line["vprei"],
        }))
        segments.append(("E1EDP20", {"WMENG": line["menge"], "EDATU": "20260715"}))
        segments.append(("E1EDP02", {"QUALF": "001", "ZEILE": line["zeile"]}))
        segments.append(("E1EDP19", {
            "QUALF": "003", "IDTNR": line["idtnr"], "KTEXT": line["ktext"],
        }))
    segments.append(("E1EDS01", {"SUMID": "001", "SUMME": item_count}))
    segments.append(("E1EDS01", {"SUMID": "002", "SUMME": net_value}))
    return segments


_ORDERS05_LINES = [
    {"zeile": "1", "menge": "12.000", "menee": "ST", "vprei": "8.50",
     "idtnr": "614141007349", "ktext": "TRAIL MIX 12OZ"},
    {"zeile": "2", "menge": "6.000", "menee": "KAR", "vprei": "24.00",
     "idtnr": "614141007350", "ktext": "SPRING WATER 24PK"},
    {"zeile": "3", "menge": "5.000", "menee": "DZN", "vprei": "12.00",
     "idtnr": "614141007351", "ktext": "GRANOLA BARS VARIETY"},
]

BASELINE_ORDERS05 = _orders05_segments(
    po_number="PO5500077", po_date="20260701", currency="USD", bsart="NB",
    order_type="NB", lines=_ORDERS05_LINES, item_count="3", net_value="306.00",
)

# Planted defects (identical in both formats):
#   1. PO number transposed                    -> value_mismatch   (I-003)
#   2. currency EUR though CUR01=BY sent USD   -> condition_logic  (I-006)
#   3. line 1 unit left untranslated (EA)      -> code_translation (I-019)
#   4. E1EDK14 012 hardcode missing            -> constant_default (I-008)
#   5. refs.001.datum as YYMMDD-ish, not SAP YYYYMMDD -> format    (I-004)
#      (kept 8 chars wide so the flat DATUM field carries it whole and
#      both formats stay byte-identical)
#   6. third line dropped                      -> count_mismatch + missing_output
#   7. stray E1EDK14 008 sales org             -> unmapped_target warning
_DEFECT_LINES = [
    {**_ORDERS05_LINES[0], "menee": "EA"},
    _ORDERS05_LINES[1],
]

DEFECTS_ORDERS05 = _orders05_segments(
    po_number="PO5500770", po_date="26-07-01", currency="EUR", bsart="NB",
    order_type=None, lines=_DEFECT_LINES, item_count="3", net_value="306.00",
    extra_org=("008", "1000"),
)
# the defective datum is ISO in refs but the document date must stay valid
for _seg, _fields in DEFECTS_ORDERS05:
    if _seg == "E1EDK03" and _fields.get("IDDAT") == "012":
        _fields["DATUM"] = "20260701"

#: Writer-side field layouts: the parser tables plus realism-only fields
#: (POSEX, E1EDP20) that the adapter deliberately ignores.
_WRITER_FIELDS = {
    "E1EDK01": {"ACTION": (0, 3), "CURCY": (4, 3), "BSART": (79, 4)},
    "E1EDK14": {"QUALF": (0, 3), "ORGID": (3, 35)},
    "E1EDK03": {"IDDAT": (0, 3), "DATUM": (3, 8)},
    "E1EDK02": {"QUALF": (0, 3), "BELNR": (3, 35), "DATUM": (44, 8)},
    "E1EDKA1": {"PARVW": (0, 3), "PARTN": (3, 17), "NAME1": (37, 35),
                "STRAS": (177, 35), "ORT01": (282, 35), "PSTLZ": (326, 9),
                "REGIO": (655, 3)},
    "E1EDP01": {"POSEX": (0, 6), "MENGE": (11, 15), "MENEE": (26, 3),
                "VPREI": (54, 15)},
    "E1EDP20": {"WMENG": (0, 15), "EDATU": (25, 8)},
    "E1EDP02": {"QUALF": (0, 3), "BELNR": (3, 35), "ZEILE": (38, 6)},
    "E1EDP19": {"QUALF": (0, 3), "IDTNR": (3, 35), "KTEXT": (38, 70)},
    "E1EDS01": {"SUMID": (0, 3), "SUMME": (3, 18)},
}


def _write_idoc_flat(segments: list[tuple[str, dict[str, str]]]) -> str:
    lines = ["EDI_DC40".ljust(30) + " " * 20 + "ORDERS05"]
    for seg_name, fields in segments:
        table = _WRITER_FIELDS[seg_name]
        end = max(off + ln for off, ln in table.values())
        sdata = [" "] * end
        for name, value in fields.items():
            offset, length = table[name]
            assert len(value) <= length, f"{seg_name}.{name} value {value!r} overflows its field"
            sdata[offset : offset + len(value)] = value
        lines.append((seg_name.ljust(30) + " " * 33 + "".join(sdata)).rstrip())
    return "\n".join(lines) + "\n"


def _write_idoc_xml(segments: list[tuple[str, dict[str, str]]]) -> str:
    import xml.etree.ElementTree as ET

    root = ET.Element("ORDERS05")
    idoc = ET.SubElement(root, "IDOC", BEGIN="1")
    control = ET.SubElement(idoc, "EDI_DC40", SEGMENT="1")
    ET.SubElement(control, "IDOCTYP").text = "ORDERS05"
    current_item = None
    for seg_name, fields in segments:
        parent = idoc
        if seg_name == "E1EDP01":
            current_item = None  # new item starts at IDOC level
        elif seg_name.startswith("E1EDP") and current_item is not None:
            parent = current_item
        element = ET.SubElement(parent, seg_name, SEGMENT="1")
        for name, value in fields.items():
            ET.SubElement(element, name).text = value
        if seg_name == "E1EDP01":
            current_item = element
    ET.indent(root)
    return ET.tostring(root, encoding="unicode") + "\n"


def generate_orders05_files() -> None:
    generate_spec(
        EXAMPLES / "specs" / "orders05_reference_spec.xlsx",
        SPEC_ORDERS05_ROWS, CODE_LIST_ORDERS05_ROWS, SPEC_ORDERS05_META,
    )
    source_path = EXAMPLES / "source" / "850_sap.edi"
    source_path.write_text(build_interchange(SAP_850, "71", "850", "PO"))
    print(f"wrote {source_path.relative_to(REPO_ROOT)}")
    for name, segments in (
        ("orders05_baseline", BASELINE_ORDERS05),
        ("orders05_defects", DEFECTS_ORDERS05),
    ):
        flat_path = EXAMPLES / "output" / f"{name}.txt"
        flat_path.write_text(_write_idoc_flat(segments))
        print(f"wrote {flat_path.relative_to(REPO_ROOT)}")
        xml_path = EXAMPLES / "output" / f"{name}.xml"
        xml_path.write_text(_write_idoc_xml(segments))
        print(f"wrote {xml_path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Outbound 855: internal POA response -> X12 855 (Direction: outbound)
# --------------------------------------------------------------------------

# Row ID, Source Field (canonical path), Loop Context (X12 target side),
# Target Field (X12 element), Rule Type, ... — see the template Instructions.
SPEC855_OUT_ROWS: list[tuple[str, ...]] = [
    ("O-001", "order.purpose", "", "BAK01", "CODE_LIST", "", "", "", "",
     "", "TX_PURPOSE_OUT", "string", "", "Internal purpose to X12 code."),
    ("O-002", "order.ack_type", "", "BAK02", "CODE_LIST", "", "", "", "",
     "", "ACK_TYPE_OUT", "string", "", ""),
    ("O-003", "order.po_number", "", "BAK03", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", "PO number being acknowledged."),
    ("O-004", "order.po_date", "", "BAK04", "DIRECT", "", "", "", "",
     "", "", "date", "%Y%m%d", "Internal ISO date must land as CCYYMMDD."),
    ("O-005", "order.vendor_order_number", "", "BAK08", "CONDITIONAL",
     "Send the vendor order number when the response carries one.",
     "EXISTS(order.vendor_order_number)", "SOURCE", "SKIP",
     "", "", "string", "", ""),
    ("O-006", "", "N1[SE]", "N103", "CONSTANT", "", "", "", "",
     "92", "", "string", "", "ID qualifier hardcoded by the map."),
    ("O-007", "vendor.name", "N1[SE]", "N102", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", "Selling party."),
    ("O-008", "vendor.id", "N1[SE]", "N104", "CONDITIONAL",
     "Send the vendor number when the response carries one.",
     "EXISTS(vendor.id)", "SOURCE", "SKIP", "", "", "string", "", ""),
    ("O-009", "lines[].line_no", "PO1", "PO101", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("O-010", "lines[].qty_ordered", "PO1", "PO102", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Ordered quantity echoed back."),
    ("O-011", "lines[].uom", "PO1", "PO103", "CODE_LIST", "", "", "", "",
     "", "UOM_OUT", "string", "", "Internal unit to X12 UOM."),
    ("O-012", "lines[].unit_price", "PO1", "PO104", "DIRECT", "", "", "", "",
     "", "", "decimal", "", "PO104 is X12 R: explicit decimal, no fixed places."),
    ("O-013", "", "PO1", "PO106", "CONSTANT", "", "", "", "",
     "UP", "", "string", "", "Product ID qualifier hardcoded per line."),
    ("O-014", "lines[].upc", "PO1", "PO107", "DIRECT", "", "", "", "",
     "", "", "string", "len:12..14", ""),
    ("O-015", "lines[].line_status", "PO1", "ACK01", "CODE_LIST", "", "", "", "",
     "", "ACK_STATUS_OUT", "string", "", "Internal status to X12 ACK code."),
    ("O-016", "lines[].qty_acknowledged", "PO1", "ACK02", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("O-017", "lines[].ack_uom", "PO1", "ACK03", "CODE_LIST", "", "", "", "",
     "", "UOM_OUT", "string", "", ""),
    ("O-018", "summary.line_count", "", "CTT01", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("O-019", "lines[]", "", "CTT01", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "",
     "CTT01 must also equal the actual number of source lines."),
]

CODE_LIST_855_OUT_ROWS: list[tuple[str, str, str, str]] = [
    ("TX_PURPOSE_OUT", "ORIGINAL", "00", "Original transmission"),
    ("TX_PURPOSE_OUT", "CANCELLATION", "01", "Cancellation"),
    ("TX_PURPOSE_OUT", "REPLACE", "05", "Replacement"),
    ("ACK_TYPE_OUT", "ACKNOWLEDGE", "AC", "Acknowledge, no detail"),
    ("ACK_TYPE_OUT", "ACK_WITH_DETAIL", "AD", "Acknowledge with detail, no change"),
    ("ACK_TYPE_OUT", "ACK_WITH_CHANGE", "AK", "Acknowledge with detail and change"),
    ("ACK_TYPE_OUT", "REJECT_WITH_DETAIL", "RD", "Reject with detail"),
    ("ACK_TYPE_OUT", "REJECTED", "RJ", "Rejected, no detail"),
    ("UOM_OUT", "EACH", "EA", "Each"),
    ("UOM_OUT", "CASE", "CA", "Case"),
    ("UOM_OUT", "DOZEN", "DZ", "Dozen"),
    ("ACK_STATUS_OUT", "ACCEPTED", "IA", "Item accepted"),
    ("ACK_STATUS_OUT", "BACKORDERED", "IB", "Item backordered"),
    ("ACK_STATUS_OUT", "QTY_CHANGED", "IQ", "Item accepted, quantity changed"),
    ("ACK_STATUS_OUT", "REJECTED", "IR", "Item rejected"),
    ("ACK_STATUS_OUT", "DATE_RESCHEDULED", "DR", "Item accepted, date rescheduled"),
]

SPEC855_OUT_META = {
    "Transaction Set": "855",
    "X12 Version": "004010",
    "Direction": "outbound",
    "Spec Name": "Synthetic POA response-to-855 map - outbound reference",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-07",
}

#: The internal POA response document — the translation's INPUT. One
#: source drives both the clean and the defective X12 output.
POA_RESPONSE: dict = {
    "order": {
        "purpose": "ORIGINAL",
        "ack_type": "ACK_WITH_DETAIL",
        "po_number": "PO4400021",
        "po_date": "2026-06-15",
        "vendor_order_number": "VN2088841",
    },
    "vendor": {"name": "SUMMIT WHOLESALE FOODS", "id": "7731"},
    "lines": [
        {"line_no": 1, "qty_ordered": 12, "uom": "EACH", "unit_price": 8.5,
         "upc": "614141007349", "line_status": "ACCEPTED",
         "qty_acknowledged": 12, "ack_uom": "EACH"},
        {"line_no": 2, "qty_ordered": 6, "uom": "CASE", "unit_price": 24,
         "upc": "614141007350", "line_status": "ACCEPTED",
         "qty_acknowledged": 6, "ack_uom": "CASE"},
        {"line_no": 3, "qty_ordered": 5, "uom": "DOZEN", "unit_price": 12,
         "upc": "614141007351", "line_status": "ACCEPTED",
         "qty_acknowledged": 5, "ack_uom": "DOZEN"},
    ],
    "summary": {"line_count": 3},
}

BASELINE_855_OUT = [
    "BAK*00*AD*PO4400021*20260615****VN2088841",
    "N1*SE*SUMMIT WHOLESALE FOODS*92*7731",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "ACK*IA*12*EA",
    "PO1*2*6*CA*24**UP*614141007350",
    "ACK*IA*6*CA",
    "PO1*3*5*DZ*12**UP*614141007351",
    "ACK*IA*5*DZ",
    "CTT*3",
]

# Deliberately defective outbound 855 (same internal source). Planted:
#   1. BAK03 PO number transposed                 -> value_mismatch   (O-003)
#   2. BAK04 date left in ISO, not CCYYMMDD       -> format           (O-004)
#   3. BAK08 carries the else-literal 'NONE'
#      though the source has a vendor order no.   -> condition_logic  (O-005)
#   4. N103 qualifier ZZ, map must hardcode 92    -> constant_default (O-006)
#   5. line 1 ACK01 left untranslated (ACCEPTED)  -> code_translation (O-015)
#   6. third line dropped from the X12            -> count_mismatch +
#      (CTT still says 3 -> recon warning too)       missing_output
#   7. stray header REF*TN nobody maps            -> unmapped_target
DEFECTS_855_OUT = [
    "BAK*00*AD*PO4400012*2026-06-15****NONE",
    "REF*TN*123456",
    "N1*SE*SUMMIT WHOLESALE FOODS*ZZ*7731",
    "PO1*1*12*EA*8.5**UP*614141007349",
    "ACK*ACCEPTED*12*EA",
    "PO1*2*6*CA*24**UP*614141007350",
    "ACK*IA*6*CA",
    "CTT*3",
]


def generate_outbound_855_files() -> None:
    generate_spec(
        EXAMPLES / "specs" / "855_outbound_reference_spec.xlsx",
        SPEC855_OUT_ROWS, CODE_LIST_855_OUT_ROWS, SPEC855_OUT_META,
    )
    source_path = EXAMPLES / "source" / "poa_response.json"
    source_path.write_text(json.dumps(POA_RESPONSE, indent=2) + "\n")
    print(f"wrote {source_path.relative_to(REPO_ROOT)}")
    outputs = {
        "855_ack_baseline.edi": build_interchange(BASELINE_855_OUT, "81", "855", "PR"),
        "855_ack_defects.edi": build_interchange(DEFECTS_855_OUT, "82", "855", "PR"),
    }
    for name, content in outputs.items():
        path = EXAMPLES / "output" / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")


# --------------------------------------------------------------------------
# Multi-transaction interchange: three 850s -> JSON array of three orders
# --------------------------------------------------------------------------

SPEC_MULTI_ROWS: list[tuple[str, ...]] = [
    ("X-001", "BEG03", "", "order.po_number", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..22", "Customer PO number; also the pairing key."),
    ("X-002", "BEG05", "", "order.po_date", "DIRECT", "", "", "", "",
     "", "", "date", "%Y-%m-%d", "PO date, CCYYMMDD in source."),
    ("X-003", "BEG02", "", "order.po_type", "CODE_LIST", "", "", "", "",
     "", "PO_TYPE", "string", "", "PO type code."),
    ("X-004", "N102", "N1[ST]", "ship_to.name", "DIRECT", "", "", "", "",
     "", "", "string", "len:1..60", "Ship-to party name."),
    ("X-005", "N401", "N1[ST]", "ship_to.city", "DIRECT", "", "", "", "",
     "", "", "string", "", ""),
    ("X-006", "PO101", "PO1", "lines[].line_no", "DIRECT", "", "", "", "",
     "", "", "integer", "", ""),
    ("X-007", "PO102", "PO1", "lines[].qty", "DIRECT", "", "", "", "",
     "", "", "integer", "", "Quantity ordered."),
    ("X-008", "PO103", "PO1", "lines[].uom", "CODE_LIST", "", "", "", "",
     "", "UOM", "string", "", "Unit of measure."),
    ("X-009", "CTT01", "", "summary.line_count", "DIRECT", "", "", "", "",
     "", "", "integer", "", "CTT line count."),
    ("X-010", "PO1", "", "summary.line_count", "LOOP_COUNT", "", "", "", "",
     "", "", "integer", "", "Actual PO1 loop count must also match."),
]

CODE_LIST_MULTI_ROWS: list[tuple[str, str, str, str]] = [
    ("PO_TYPE", "SA", "STANDALONE", "Stand-alone order"),
    ("PO_TYPE", "NE", "NEW_ORDER", "New order"),
    ("UOM", "EA", "EACH", "Each"),
    ("UOM", "CA", "CASE", "Case"),
    ("UOM", "DZ", "DOZEN", "Dozen"),
]

SPEC_MULTI_META = {
    "Transaction Set": "850",
    "X12 Version": "004010",
    "Pairing Key (source)": "BEG03",
    "Pairing Key (target)": "order.po_number",
    "Spec Name": "Synthetic multi-order interchange - reference example",
    "Author": "EDI MapCheck project",
    "Date": "2026-07-10",
}


def _iso(x12_date: str) -> str:
    """CCYYMMDD -> CCYY-MM-DD, matching the spec's %Y-%m-%d output format."""
    return f"{x12_date[:4]}-{x12_date[4:6]}-{x12_date[6:]}"


def _multi_850(po_number: str, po_date: str, city: str, lines: list[tuple]) -> list[str]:
    """One 850 transaction body — only elements the spec maps, so a clean
    run stays silent (no rule-7 unmapped-source noise)."""
    segs = [
        f"BEG**SA*{po_number}**{po_date}",  # BEG02 type, BEG03 PO, BEG05 date
        f"N1*ST*RIVERSIDE MARKET",           # N101 context qualifier, N102 name
        f"N4*{city}",                        # N401 city
    ]
    for line_no, qty, uom in lines:
        segs.append(f"PO1*{line_no}*{qty}*{uom}")
    segs.append(f"CTT*{len(lines)}")
    return segs


def _order_doc(po_number: str, po_date: str, city: str, lines: list[tuple]) -> dict:
    return {
        "order": {"po_number": po_number, "po_date": _iso(po_date), "po_type": "STANDALONE"},
        "ship_to": {"name": "RIVERSIDE MARKET", "city": city},
        "lines": [
            {"line_no": ln, "qty": qty, "uom": {"EA": "EACH", "CA": "CASE", "DZ": "DOZEN"}[u]}
            for ln, qty, u in lines
        ],
        "summary": {"line_count": len(lines)},
    }


# three orders, distinct PO numbers
_MULTI_ORDERS = [
    ("PO7700001", "20260701", "BOULDER", [(1, 12, "EA"), (2, 6, "CA")]),
    ("PO7700002", "20260702", "DENVER", [(1, 24, "EA")]),
    ("PO7700003", "20260703", "AURORA", [(1, 5, "DZ"), (2, 10, "EA"), (3, 3, "CA")]),
]

BASELINE_MULTI_SOURCE = [_multi_850(*o) for o in _MULTI_ORDERS]
BASELINE_MULTI_OUTPUT = [_order_doc(*o) for o in _MULTI_ORDERS]

# Defective interchange. Planted, across the set:
#   1. document PO7700002 has a wrong ship_to city  -> per-doc value_mismatch
#   2. source adds PO7700004 with no output document -> missing_output (file)
#   3. output adds PO7700009 with no source txn      -> unexpected_output (file)
#   4. source PO7700001 appears twice (dup key)      -> count_mismatch (file)
DEFECT_SOURCE_ORDERS = list(_MULTI_ORDERS) + [
    ("PO7700001", "20260701", "BOULDER", [(1, 12, "EA"), (2, 6, "CA")]),  # dup key
    ("PO7700004", "20260704", "LONGMONT", [(1, 8, "EA")]),  # no output doc
]
DEFECTS_MULTI_SOURCE = [_multi_850(*o) for o in DEFECT_SOURCE_ORDERS]

_DEFECT_DOC_2 = _order_doc(*_MULTI_ORDERS[1])
_DEFECT_DOC_2["ship_to"]["city"] = "COLORADO SPRINGS"  # wrong city, doc PO7700002
DEFECTS_MULTI_OUTPUT = [
    BASELINE_MULTI_OUTPUT[0],
    _DEFECT_DOC_2,
    BASELINE_MULTI_OUTPUT[2],
    _order_doc("PO7700009", "20260709", "GOLDEN", [(1, 1, "EA")]),  # no source txn
]


def generate_multi_interchange_files() -> None:
    generate_spec(
        EXAMPLES / "specs" / "850_multi_reference_spec.xlsx",
        SPEC_MULTI_ROWS, CODE_LIST_MULTI_ROWS, SPEC_MULTI_META,
    )
    sources = {
        "850_multi_baseline.edi": build_multi_interchange(
            BASELINE_MULTI_SOURCE, "91", "850", "PO"),
        "850_multi_defects.edi": build_multi_interchange(
            DEFECTS_MULTI_SOURCE, "92", "850", "PO"),
    }
    for name, content in sources.items():
        path = EXAMPLES / "source" / name
        path.write_text(content)
        print(f"wrote {path.relative_to(REPO_ROOT)}")
    outputs = {
        "orders_multi_baseline.json": BASELINE_MULTI_OUTPUT,
        "orders_multi_defects.json": DEFECTS_MULTI_OUTPUT,
    }
    for name, data in outputs.items():
        path = EXAMPLES / "output" / name
        path.write_text(json.dumps(data, indent=2) + "\n")
        print(f"wrote {path.relative_to(REPO_ROOT)}")


def main() -> None:
    generate_spec(
        EXAMPLES / "specs" / "850_reference_spec.xlsx",
        SPEC_ROWS, CODE_LIST_ROWS, SPEC_META,
    )
    generate_source_files()
    generate_output_files()
    generate_spec(
        EXAMPLES / "specs" / "856_reference_spec.xlsx",
        SPEC856_ROWS, CODE_LIST_856_ROWS, SPEC856_META,
    )
    generate_856_files()
    generate_spec(
        EXAMPLES / "specs" / "855_reference_spec.xlsx",
        SPEC855_ROWS, CODE_LIST_855_ROWS, SPEC855_META,
    )
    generate_spec(
        EXAMPLES / "specs" / "810_reference_spec.xlsx",
        SPEC810_ROWS, CODE_LIST_810_ROWS, SPEC810_META,
    )
    generate_order_cycle_files()
    generate_warehouse_files()
    generate_inventory_files()
    generate_pharma_files()
    generate_997_files()
    generate_orders05_files()
    generate_outbound_855_files()
    generate_multi_interchange_files()


if __name__ == "__main__":
    main()
