"""draft-spec (Design 012): determinism, golden content, crosswalk
validation, TODO parity, and the full round trip.

Golden tests compare parsed sheet content, not raw bytes — xlsx embeds
save timestamps, so byte equality is meaningless while content equality
is the actual determinism contract.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from mapcheck.engine.results import Status
from mapcheck.engine.validator import validate_files
from mapcheck.output.idoc import default_output_registry
from mapcheck.spec.draft import (
    Crosswalk,
    CrosswalkError,
    assemble_draft,
    bundled_crosswalks,
    generate_draft,
    load_crosswalk,
    load_crosswalks,
    walk_source_elements,
    walk_target_paths,
)
from mapcheck.spec.parser import load_spec
from mapcheck.transactions.registry import default_registry

REPO = Path(__file__).resolve().parent.parent
EXAMPLES = REPO / "examples"

DEFINITION_850 = default_registry.get("850")
ORDERS05 = default_output_registry.get("idoc-orders05")
STARTER = bundled_crosswalks("850", "orders05")


def _draft(tmp_path: Path, name: str = "draft.xlsx", fill_unmapped: bool = False):
    out = tmp_path / name
    result = generate_draft(
        DEFINITION_850, ORDERS05, list(STARTER), out,
        include_optional_todos=fill_unmapped,
    )
    return result, out


def _sheet(path: Path, title: str, width: int) -> list[tuple]:
    ws = load_workbook(path)[title]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        rows.append(tuple(v if v is not None else "" for v in row[:width]))
    return rows


#: The golden 850 -> orders05 draft, parsed. Columns: Row ID, Source Field,
#: Loop Context, Target Field, Rule Type, Condition (Text), Condition
#: (Coded), Then, Else, Default Value, Code List Ref, Data Type, Format,
#: Notes.
GOLDEN_MAPPING = [
    ("D-001", "BEG01", "", "header.action", "CODE_LIST", "", "", "", "", "", "SAP_ACTION", "string", "", "IDoc action code from the transaction set purpose."),
    ("D-002", "CUR02", "", "header.curcy", "CONDITIONAL", "", "CUR01 = 'BY'", "SOURCE", "SKIP", "", "", "string", "len:3..3", "Currency only when a buying-party currency code is sent."),
    ("D-003", "BEG02", "", "header.bsart", "CODE_LIST", "", "", "", "", "", "PO_TYPE_SAP", "string", "", "SAP document type from the PO type code."),
    ("D-004", "BEG05", "", "dates.012", "DIRECT", "", "", "", "", "", "", "date", "%Y%m%d", "E1EDK03 IDDAT 012 = document date. SAP dates are YYYYMMDD."),
    ("D-005", "DTM02", "DTM[002]", "dates.002", "DIRECT", "", "", "", "", "", "", "date", "%Y%m%d", "E1EDK03 IDDAT 002 = requested delivery date."),
    ("D-006", "BEG03", "", "refs.001.belnr", "DIRECT", "", "", "", "", "", "", "string", "len:1..35", "E1EDK02 QUALF 001 = customer PO number."),
    ("D-007", "BEG05", "", "refs.001.datum", "DIRECT", "", "", "", "", "", "", "date", "%Y%m%d", "E1EDK02 QUALF 001 = customer PO date."),
    ("D-008", "N102", "N1[BT]", "partners.ag.name1", "DIRECT", "", "", "", "", "", "", "string", "len:1..35", "AG = sold-to party, from the bill-to N1 loop."),
    ("D-009", "N104", "N1[BT]", "partners.ag.partn", "CONDITIONAL", "", "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", "Sold-to number only when the ID qualifier is 92 (assigned by buyer)."),
    ("D-010", "N102", "N1[ST]", "partners.we.name1", "DIRECT", "", "", "", "", "", "", "string", "len:1..35", "WE = ship-to party."),
    ("D-011", "N104", "N1[ST]", "partners.we.partn", "CONDITIONAL", "", "N103 = '92'", "SOURCE", "SKIP", "", "", "string", "", "Ship-to number only when the ID qualifier is 92 (assigned by buyer)."),
    ("D-012", "N301", "N1[ST]", "partners.we.stras", "DIRECT", "", "", "", "", "", "", "string", "", "Ship-to street address."),
    ("D-013", "N401", "N1[ST]", "partners.we.ort01", "DIRECT", "", "", "", "", "", "", "string", "", "Ship-to city."),
    ("D-014", "N402", "N1[ST]", "partners.we.regio", "DIRECT", "", "", "", "", "", "", "string", "len:2..3", "Ship-to region (state/province as sent)."),
    ("D-015", "N403", "N1[ST]", "partners.we.pstlz", "DIRECT", "", "", "", "", "", "", "string", "len:5..9", "Ship-to postal code."),
    ("D-016", "PO101", "PO1", "lines[].refs.001.zeile", "DIRECT", "", "", "", "", "", "", "integer", "", "E1EDP02 QUALF 001 = customer PO line number."),
    ("D-017", "PO102", "PO1", "lines[].menge", "DIRECT", "", "", "", "", "", "", "decimal", "", "Order quantity; SAP quantities carry decimals."),
    ("D-018", "PO103", "PO1", "lines[].menee", "CODE_LIST", "", "", "", "", "", "UOM_SAP", "string", "", "X12 unit of measure to SAP unit."),
    ("D-019", "PO104", "PO1", "lines[].vprei", "DIRECT", "", "", "", "", "", "", "decimal", "places:2", "Unit price."),
    ("D-020", "", "", "lines[].ids.003.idtnr", "TODO", "", "", "", "", "", "", "string", "", "TODO: EAN/UPC (E1EDP19 QUALF 003)"),
    ("D-021", "", "", "lines[].ids.003.ktext", "TODO", "", "", "", "", "", "", "string", "", "TODO: Item description (E1EDP19 KTEXT)"),
    ("D-022", "CTT01", "", "summary.001", "DIRECT", "", "", "", "", "", "", "integer", "", "E1EDS01 SUMID 001 = item count as transmitted."),
    ("D-023", "AMT02", "AMT[TT]", "summary.002", "DIRECT", "", "", "", "", "", "", "decimal", "places:2", "E1EDS01 SUMID 002 = net order value."),
]

GOLDEN_UNMAPPED = [
    ("BEG04", "", "Release Number"),
    ("REF01", "", "Reference Identification Qualifier"),
    ("REF02", "", "Reference Identification"),
    ("PER01", "", "Contact Function Code"),
    ("PER02", "", "Contact Name"),
    ("PER03", "", "Communication Qualifier"),
    ("PER04", "", "Communication Number"),
    ("SAC01", "", "Allowance or Charge Indicator"),
    ("SAC02", "", "Allowance/Charge Code"),
    ("SAC03", "", "Agency Qualifier"),
    ("SAC04", "", "Agency Code"),
    ("SAC05", "", "Amount"),
    ("PID01", "", "Item Description Type"),
    ("PID02", "", "Product Characteristic Code"),
    ("PID03", "", "Agency Qualifier"),
    ("PID04", "", "Product Description Code"),
    ("PID05", "", "Description"),
    ("N302", "N1", "Address Line 2"),
    ("PO105", "PO1", "Basis of Unit Price Code"),
    ("PO106", "PO1", "Product ID Qualifier"),
    ("PO107", "PO1", "Product ID"),
    ("PO108", "PO1", "Product ID Qualifier 2"),
    ("PO109", "PO1", "Product ID 2"),
    ("CTT02", "", "Hash Total"),
]


class TestGoldenDraft:
    def test_mapping_sheet_matches_golden(self, tmp_path: Path):
        _, out = _draft(tmp_path)
        assert _sheet(out, "Mapping", 14) == GOLDEN_MAPPING

    def test_unmapped_source_sheet_matches_golden(self, tmp_path: Path):
        _, out = _draft(tmp_path)
        assert _sheet(out, "Unmapped Source", 3) == GOLDEN_UNMAPPED

    def test_meta_records_provenance_and_prefill(self, tmp_path: Path):
        result, out = _draft(tmp_path)
        meta = dict(
            row[:2]
            for row in load_workbook(out)["Meta"].iter_rows(min_row=2, values_only=True)
            if row[0]
        )
        assert meta["Transaction Set"] == "850"
        assert meta["Target Format"] == "idoc-orders05"
        assert meta["Crosswalks"] == "850_orders05.yaml"
        assert meta["Prefill"] == "0.90"
        assert meta["Tool Version"]
        assert result.prefill == pytest.approx(0.90)
        assert (result.filled_required, result.total_required) == (18, 20)

    def test_prefill_meets_the_bar(self, tmp_path: Path):
        result, _ = _draft(tmp_path)
        assert result.prefill >= 0.70

    def test_fill_unmapped_adds_optional_todos(self, tmp_path: Path):
        result, _ = _draft(tmp_path, fill_unmapped=True)
        todo_targets = {row.target for row in result.rows if not row.filled}
        assert "org.012" in todo_targets  # optional, deliberately uncrosswalked
        # prefill counts required targets only — unchanged by optional TODOs
        assert result.prefill == pytest.approx(0.90)


class TestDeterminism:
    def test_two_runs_produce_identical_parsed_content(self, tmp_path: Path):
        _, first = _draft(tmp_path, "one.xlsx")
        _, second = _draft(tmp_path, "two.xlsx")
        for title, width in (("Mapping", 14), ("CodeLists", 4), ("Unmapped Source", 3)):
            assert _sheet(first, title, width) == _sheet(second, title, width)
        meta = lambda p: [  # noqa: E731
            row[:2]
            for row in load_workbook(p)["Meta"].iter_rows(min_row=2, values_only=True)
        ]
        assert meta(first) == meta(second)


class TestCrosswalkValidation:
    def _load(self, tmp_path: Path, text: str) -> Crosswalk:
        path = tmp_path / "xw.yaml"
        path.write_text(text, encoding="utf-8")
        return load_crosswalk(path)

    def test_missing_target_names_file_and_entry(self, tmp_path: Path):
        with pytest.raises(CrosswalkError) as excinfo:
            self._load(tmp_path, "rules:\n  - {source: BEG03, rule: DIRECT}\n")
        assert "xw.yaml: entry [0]: 'target' is required" in str(excinfo.value)

    def test_unknown_rule_names_file_and_entry(self, tmp_path: Path):
        with pytest.raises(CrosswalkError) as excinfo:
            self._load(
                tmp_path,
                "rules:\n"
                "  - {source: BEG03, target: refs.001.belnr, rule: DIRECT}\n"
                "  - {source: BEG05, target: refs.001.datum, rule: COPY}\n",
            )
        assert "entry [1]: unknown rule 'COPY'" in str(excinfo.value)

    def test_code_list_rule_requires_a_resolvable_list(self, tmp_path: Path):
        path = tmp_path / "xw.yaml"
        path.write_text(
            "rules:\n  - {source: PO103, target: 'lines[].menee', rule: CODE_LIST, "
            "code_list: NOPE}\n",
            encoding="utf-8",
        )
        with pytest.raises(CrosswalkError) as excinfo:
            load_crosswalks([path])
        assert "references unknown code list 'NOPE'" in str(excinfo.value)

    def test_bad_condition_and_unknown_field_are_reported(self, tmp_path: Path):
        with pytest.raises(CrosswalkError) as excinfo:
            self._load(
                tmp_path,
                "rules:\n"
                "  - {source: N104, target: partners.we.partn, rule: CONDITIONAL, "
                "condition: 'N103 == 92', then: SOURCE}\n"
                "  - {sorce: BEG03, target: refs.001.belnr, rule: DIRECT}\n",
            )
        message = str(excinfo.value)
        assert "entry [0]: condition:" in message
        assert "entry [1]: unknown field(s) sorce" in message

    def test_duplicate_target_within_one_file_is_an_error(self, tmp_path: Path):
        with pytest.raises(CrosswalkError) as excinfo:
            self._load(
                tmp_path,
                "rules:\n"
                "  - {source: BEG03, target: refs.001.belnr, rule: DIRECT}\n"
                "  - {source: BEG04, target: refs.001.belnr, rule: DIRECT}\n",
            )
        assert "duplicate target 'refs.001.belnr'" in str(excinfo.value)

    def test_todo_is_not_a_crosswalk_rule(self, tmp_path: Path):
        with pytest.raises(CrosswalkError) as excinfo:
            self._load(
                tmp_path,
                "rules:\n  - {target: 'lines[].ids.003.idtnr', rule: TODO}\n",
            )
        assert "TODO is not a crosswalk pairing" in str(excinfo.value)


class TestCrosswalkMerge:
    def test_later_file_wins_per_target_path(self, tmp_path: Path):
        override = tmp_path / "override.yaml"
        override.write_text(
            "rules:\n"
            "  - source: REF02\n"
            "    target: refs.001.belnr\n"
            "    rule: DIRECT\n"
            "    context: REF[PO]\n"
            "    note: partner sends the PO number in REF*PO instead\n",
            encoding="utf-8",
        )
        merged = load_crosswalks([*STARTER, override])
        rule = merged.rules["refs.001.belnr"]
        assert rule.source == "REF02"
        assert rule.context == "REF[PO]"
        # untouched targets keep the starter's pairing
        assert merged.rules["lines[].menge"].source == "PO102"
        assert len(merged.files) == 2

    def test_draft_reflects_the_override(self, tmp_path: Path):
        override = tmp_path / "override.yaml"
        override.write_text(
            "rules:\n"
            "  - {source: REF02, target: refs.001.belnr, rule: DIRECT, context: 'REF[PO]'}\n",
            encoding="utf-8",
        )
        crosswalk = load_crosswalks([*STARTER, override])
        result = assemble_draft(DEFINITION_850, ORDERS05, crosswalk)
        row = next(r for r in result.rows if r.target == "refs.001.belnr")
        assert (row.source, row.context) == ("REF02", "REF[PO]")


class TestWalks:
    def test_source_walk_is_definition_ordered_and_labeled(self):
        elements = walk_source_elements(DEFINITION_850)
        fields = [e.field for e in elements]
        assert fields[0] == "BEG01"  # first header segment with a dictionary
        assert fields.index("BEG01") < fields.index("N101") < fields.index("PO101")
        po102 = next(e for e in elements if e.field == "PO102")
        assert po102.context == "PO1"
        assert po102.name == "Quantity Ordered"

    def test_target_walk_follows_definition_order(self):
        paths = [t.path for t in walk_target_paths(ORDERS05)]
        assert paths[0] == "header.action"
        assert paths.index("refs.001.belnr") < paths.index("lines[].menge") < paths.index("summary.001")


class TestTodoParity:
    """A TODO row loads as NOT TESTED with a load warning (Design 001
    BLANK parity), so an uncurated draft can never silently pass."""

    def test_draft_loads_with_todo_warnings(self, tmp_path: Path):
        _, out = _draft(tmp_path)
        spec = load_spec(out)
        todo_notes = [n for n in spec.load_notes if "Rule Type TODO" in n]
        assert len(todo_notes) == 2
        assert any("lines[].ids.003.idtnr" in n for n in todo_notes)

    def test_todo_rows_report_not_tested(self, tmp_path: Path):
        _, out = _draft(tmp_path)
        run = validate_files(
            str(out),
            str(EXAMPLES / "source" / "850_sap.edi"),
            str(EXAMPLES / "output" / "orders05_baseline.txt"),
        )
        todo_findings = [
            f for f in run.findings if f.row_id in ("D-020", "D-021")
        ]
        assert [f.status for f in todo_findings] == [Status.NOT_TESTED] * 2
        assert all("TODO" in f.message for f in todo_findings)
        # the draft warns (spec notes + unmapped sweeps): never a silent pass
        assert run.overall is Status.WARNING
        assert run.exit_code == 0


class TestRoundTrip:
    """Draft, complete the TODOs from the reference spec, run the baseline
    scenario, expect PASS."""

    def test_completed_draft_passes_the_baseline(self, tmp_path: Path):
        _, out = _draft(tmp_path, fill_unmapped=True)
        reference = {
            row[3]: row
            for row in load_workbook(EXAMPLES / "specs" / "orders05_reference_spec.xlsx")[
                "Mapping"
            ].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        wb = load_workbook(out)
        ws = wb["Mapping"]
        completed = []
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=5).value == "TODO":
                target = ws.cell(row=r, column=4).value
                ref_row = reference[target]
                for c in range(2, 15):  # keep the draft's Row ID
                    ws.cell(row=r, column=c, value=ref_row[c - 1])
                completed.append(target)
        assert completed == ["org.012", "lines[].ids.003.idtnr", "lines[].ids.003.ktext"]
        done = tmp_path / "draft_completed.xlsx"
        wb.save(done)

        run = validate_files(
            str(done),
            str(EXAMPLES / "source" / "850_sap.edi"),
            str(EXAMPLES / "output" / "orders05_baseline.txt"),
        )
        assert run.overall is Status.PASS
        assert run.counts[Status.FAIL] == 0
        assert run.counts[Status.WARNING] == 0


class TestCli:
    def test_draft_spec_verb_prints_prefill(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        out = tmp_path / "cli_draft.xlsx"
        code = main(
            [
                "draft-spec", "--transaction", "850", "--target", "orders05",
                "--output", str(out),
            ]
        )
        captured = capsys.readouterr()
        assert code == 0
        assert out.exists()
        assert "Prefill: 0.90 (18/20 required targets filled)" in captured.out

    def test_unknown_target_is_a_usage_error(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        code = main(
            [
                "draft-spec", "--transaction", "850", "--target", "nope",
                "--output", str(tmp_path / "x.xlsx"),
            ]
        )
        assert code == 2
        assert "unknown target format 'nope'" in capsys.readouterr().err
