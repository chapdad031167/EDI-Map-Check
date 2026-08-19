"""Security-hardening regression tests.

- H-1: spreadsheet formula injection is neutralized in the Excel export.
- M-1: an out-of-range decimal magnitude is rejected, not expanded.
- M-2: an XML DOCTYPE (entity-bomb vector) is refused.
- M-3: one failing check never aborts a whole batch run.
- L-2: the scrubber keys its cache on (value, strategy).
"""

from __future__ import annotations

from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from mapcheck.engine.formats import (
    NormalizationError,
    normalize_actual,
    normalize_source,
    normalize_x12,
)
from mapcheck.engine.results import Category, Finding, RunResult, Status
from mapcheck.output.adapter import OutputLoadError, parse_xml_safely
from mapcheck.report.excel import export_excel
from mapcheck.spec.formats import FormatSpec
from mapcheck.spec.model import DataType


# ---------------------------------------------------------------------------
# H-1 — Excel formula injection
# ---------------------------------------------------------------------------


class TestExcelFormulaInjection:
    def _run(self, actual: str) -> RunResult:
        return RunResult(
            spec_path="s", source_path="x", output_path="y", spec_name="t",
            findings=[Finding(
                status=Status.FAIL, message="m", row_id="M-1", target="order.x",
                expected="ok", actual=actual, category=Category.VALUE_MISMATCH,
            )],
            run_at=datetime(2026, 7, 14, tzinfo=timezone.utc),
        )

    @pytest.mark.parametrize("payload", [
        '=HYPERLINK("http://evil/","x")', "+1+1", "-2+3", "@SUM(A1)", "\t=1",
    ])
    def test_formula_leaders_neutralized(self, payload, tmp_path):
        path = export_excel(self._run(payload), tmp_path / "r.xlsx")
        ws = load_workbook(path).active
        # find the Actual column (6) on the first data row
        cell = ws.cell(row=2, column=6)
        assert cell.data_type != "f"          # not treated as a formula
        assert str(cell.value).startswith("'")  # apostrophe-guarded

    def test_ordinary_value_untouched(self, tmp_path):
        path = export_excel(self._run("PO4400021"), tmp_path / "r.xlsx")
        ws = load_workbook(path).active
        assert ws.cell(row=2, column=6).value == "PO4400021"


# ---------------------------------------------------------------------------
# M-1 — decimal expansion DoS
# ---------------------------------------------------------------------------


class TestDecimalBound:
    @pytest.mark.parametrize("bad", ["1E1000000", "1E-9999999", "9" * 5000])
    def test_oversized_rejected(self, bad):
        with pytest.raises(NormalizationError):
            normalize_source(bad, DataType.DECIMAL, FormatSpec())

    def test_normal_values_still_work(self):
        assert normalize_source("123456789012.50", DataType.DECIMAL, FormatSpec())
        val, _ = normalize_actual("25.00", DataType.DECIMAL, FormatSpec(), typed=False)
        assert str(val) == "25.00"
        x, _ = normalize_x12("306.00", DataType.DECIMAL, FormatSpec())
        assert str(x) == "306.00"

    def test_x12_target_side_rejects_oversized(self):
        with pytest.raises(NormalizationError):
            normalize_x12("1E1000000", DataType.DECIMAL, FormatSpec())


# ---------------------------------------------------------------------------
# M-2 — XML DOCTYPE / entity bomb
# ---------------------------------------------------------------------------


class TestXmlDoctype:
    def test_doctype_refused(self, tmp_path):
        bomb = (
            '<?xml version="1.0"?><!DOCTYPE lolz [<!ENTITY a "aa">]>'
            "<ORDERS05>&a;</ORDERS05>"
        )
        p = tmp_path / "b.xml"
        p.write_text(bomb, encoding="utf-8")
        with pytest.raises(OutputLoadError, match="DOCTYPE"):
            parse_xml_safely(p)

    def test_plain_xml_parses(self, tmp_path):
        p = tmp_path / "ok.xml"
        p.write_text("<ORDERS05><E1EDK01>x</E1EDK01></ORDERS05>", encoding="utf-8")
        root = parse_xml_safely(p)
        assert root.tag == "ORDERS05"

    def test_malformed_xml_is_load_error(self, tmp_path):
        p = tmp_path / "bad.xml"
        p.write_text("<ORDERS05><unclosed>", encoding="utf-8")
        with pytest.raises(OutputLoadError):
            parse_xml_safely(p)

    #: A DOCTYPE is invisible to a raw byte scan in any encoding that is
    #: not ASCII-compatible, but ElementTree honors the BOM and expands
    #: the entities regardless.
    _BOMB = (
        '<?xml version="1.0" encoding="{}"?>'
        '<!DOCTYPE lolz [<!ENTITY a "aa">]>'
        "<ORDERS05>&a;</ORDERS05>"
    )

    @pytest.mark.parametrize(
        "encoding,label",
        [
            ("utf-16", "UTF-16"),      # codec writes the BOM
            ("utf-16-le", "UTF-16"),   # no BOM: malformed, still refused
            ("utf-16-be", "UTF-16"),
            ("utf-32", "UTF-32"),
        ],
    )
    def test_doctype_refused_in_wide_encodings(self, tmp_path, encoding, label):
        p = tmp_path / "b.xml"
        p.write_bytes(self._BOMB.format(label).encode(encoding))
        with pytest.raises(OutputLoadError, match="DOCTYPE"):
            parse_xml_safely(p)

    def test_utf16_entity_is_not_expanded(self, tmp_path):
        """The guard has to win before ElementTree does the expanding."""
        p = tmp_path / "b.xml"
        p.write_bytes(self._BOMB.format("UTF-16").encode("utf-16"))
        try:
            root = parse_xml_safely(p)
        except OutputLoadError:
            return
        raise AssertionError(f"entity expanded to {root.text!r} instead of refusing")

    # UTF-32 is absent on purpose: expat, and so ElementTree, has never
    # read it, which is why a UTF-32 DOCTYPE is refused above for
    # completeness rather than because it could otherwise be expanded.
    @pytest.mark.parametrize("encoding", ["utf-8", "utf-8-sig", "utf-16"])
    def test_clean_xml_still_parses_in_every_encoding(self, tmp_path, encoding):
        """Over-rejecting would break legitimate IDoc XML — it must not."""
        declared = "UTF-8" if encoding.startswith("utf-8") else encoding.upper()
        p = tmp_path / "ok.xml"
        p.write_bytes(
            f'<?xml version="1.0" encoding="{declared}"?>'
            "<ORDERS05><E1EDK01>x</E1EDK01></ORDERS05>".encode(encoding)
        )
        assert parse_xml_safely(p).findtext("E1EDK01") == "x"


# ---------------------------------------------------------------------------
# M-3 — batch resilience
# ---------------------------------------------------------------------------


class TestBatchResilience:
    def test_unexpected_error_becomes_error_outcome(self, tmp_path, monkeypatch, examples_dir):
        from mapcheck import batch as batch_mod

        calls = {"n": 0}

        def boom(check, history):
            calls["n"] += 1
            if check.name == "explode":
                raise RuntimeError("unexpected internal error")
            return batch_mod.CheckOutcome(check.name, "850", "pass", {})

        monkeypatch.setattr(batch_mod, "_run_validate", boom)
        manifest = batch_mod.Manifest(checks=[
            batch_mod.Check("explode", "s", "x", "y"),
            batch_mod.Check("survivor", "s", "x", "y"),
        ])
        result = batch_mod.run_batch(manifest, str(tmp_path / "h.db"), record=False)
        # the RuntimeError did not abort the batch — both checks ran
        assert calls["n"] == 2
        assert result.exit_code == 2
        statuses = {o.name: o.status for o in result.outcomes}
        assert statuses == {"explode": "error", "survivor": "pass"}


# ---------------------------------------------------------------------------
# L-2 — scrub cache keyed by (value, strategy)
# ---------------------------------------------------------------------------


class TestScrubCacheKey:
    def test_same_value_different_strategy_differs(self):
        from mapcheck.scrub import Profile
        from mapcheck.scrub.scrubber import Scrubber

        s = Scrubber(Profile(name="t", rules=[], envelope=False), seed="x")
        pseudo = s._mask_value("SECRET-42", "id")
        redact = s._mask_value("SECRET-42", "redact")
        assert pseudo != redact
        assert redact == "XXXXXX-XX"  # alnum -> X, separator kept
        # and each strategy is still internally consistent
        assert s._mask_value("SECRET-42", "id") == pseudo
