"""Phase 0 tests: definition schema, loader, registry, and the plugin path.

The toy-transaction test is the architectural proof: a transaction set the
codebase has never heard of, defined purely as YAML, parses and validates
end to end with zero parser/engine changes.
"""

from __future__ import annotations

import textwrap
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mapcheck.cli import main
from mapcheck.engine import Category, Status, validate, validate_files
from mapcheck.output.adapter import load_output
from mapcheck.spec.parser import SpecLoadError, load_spec
from mapcheck.spec.template import create_template
from mapcheck.transactions import (
    DefinitionError,
    TransactionRegistry,
    default_registry,
    load_definition,
)
from mapcheck.transactions.registry import UnknownTransactionError
from mapcheck.x12 import X12ParseError, parse_transaction


class TestBundled850Definition:
    def test_bundled_850_loads(self):
        definition = default_registry.get("850")
        assert definition.name == "Purchase Order"
        assert definition.functional_group == "PO"
        assert [area.id for area in definition.areas] == ["header", "detail", "summary"]
        assert [loop.id for loop in definition.all_loops()] == ["N1", "PO1"]
        assert definition.output_pairing.loop == "PO1"
        assert definition.reconciliation[0].id == "ctt-line-count"

    def test_unknown_set_raises(self):
        with pytest.raises(UnknownTransactionError, match="ST01='811'"):
            default_registry.get("811")

    def test_registry_refuses_silent_shadowing(self):
        definition = default_registry.get("850")
        with pytest.raises(ValueError, match="already registered"):
            default_registry.register(definition)


TOY_YAML = """
transaction:
  set: "999"
  name: Toy Transaction (tests only)
  functional_group: ZZ
  version: "004010"

structure:
  - area: header
    segments: [XA, XB]
    loops:
      - id: LQ
        trigger: LQ
        qualifier: 1
        segments: [XR]
  - area: detail
    opens_area: [LD]
    loops:
      - id: LD
        trigger: LD
        segments: [XN]
        max_repeats: 2
  - area: summary
    opens_area: [XT]
    segments: [XT]

output_pairing:
  list_path: lines
  loop: LD

reconciliation:
  - id: xt-line-count
    description: XT01 must equal the LD loop count
    left: { value: XT01 }
    right: { count: LD }
    when: { exists: XT01 }
    severity: error
"""


def _write_toy_edi(path: Path, business: list[str], set_code: str = "999") -> Path:
    segments = [
        "ISA*00*          *00*          *ZZ*SND            *ZZ*RCV            "
        "*260701*0900*U*00401*000000042*0*T*>",
        "GS*ZZ*SND*RCV*20260701*0900*42*X*004010",
        f"ST*{set_code}*0042",
        *business,
        f"SE*{len(business) + 2}*0042",
        "GE*1*42",
        "IEA*1*000000042",
    ]
    path.write_text("~\n".join(segments) + "~\n")
    return path


@pytest.fixture()
def toy_registry(tmp_path: Path) -> TransactionRegistry:
    yaml_path = tmp_path / "999.yaml"
    yaml_path.write_text(textwrap.dedent(TOY_YAML))
    registry = TransactionRegistry()
    registry.register_file(str(yaml_path))
    return registry


class TestToyTransactionEndToEnd:
    """The plugin path: a brand-new set defined in YAML validates end to end."""

    def _toy_source(self, tmp_path: Path) -> Path:
        return _write_toy_edi(
            tmp_path / "toy.edi",
            [
                "XA*HELLO*20260701",
                "LQ*AA*Alpha",
                "XR*Q1",
                "LD*1*RED",
                "XN*first",
                "LD*2*BLUE",
                "XN*second",
                "XT*2",
            ],
        )

    def _toy_spec(self, tmp_path: Path) -> Path:
        spec_path = create_template(tmp_path / "toy_spec.xlsx")
        wb = load_workbook(spec_path)
        ws = wb["Mapping"]
        rows = [
            ("T-001", "XA01", "", "order.greeting", "DIRECT", "", "", "", "", "", "",
             "string", "", ""),
            ("T-002", "XA02", "", "order.doc_date", "DIRECT", "", "", "", "", "", "",
             "date", "%Y-%m-%d", ""),
            ("T-003", "LQ02", "LQ[AA]", "order.alpha_name", "DIRECT", "", "", "", "",
             "", "", "string", "", ""),
            ("T-004", "XR01", "LQ[AA]", "order.alpha_ref", "DIRECT", "", "", "", "",
             "", "", "string", "", ""),
            ("T-005", "LD01", "LD", "lines[].seq", "DIRECT", "", "", "", "", "", "",
             "integer", "", ""),
            ("T-006", "LD02", "LD", "lines[].color", "DIRECT", "", "", "", "", "", "",
             "string", "", ""),
            ("T-007", "XN01", "LD", "lines[].note", "DIRECT", "", "", "", "", "", "",
             "string", "", ""),
            ("T-008", "XT01", "", "summary.line_count", "DIRECT", "", "", "", "", "",
             "", "integer", "", ""),
            ("T-009", "LD", "", "summary.line_count", "LOOP_COUNT", "", "", "", "", "",
             "", "integer", "", ""),
        ]
        for r, row in enumerate(rows, start=2):
            for c, value in enumerate(row, start=1):
                if value != "":
                    ws.cell(row=r, column=c, value=value)
        ws_meta = wb["Meta"]
        ws_meta["B2"] = "999"
        wb.save(spec_path)
        return spec_path

    def _toy_output(self, tmp_path: Path) -> Path:
        out = tmp_path / "toy.json"
        out.write_text(
            """
            {
              "order": {"greeting": "HELLO", "doc_date": "2026-07-01",
                        "alpha_name": "Alpha", "alpha_ref": "Q1"},
              "lines": [
                {"seq": 1, "color": "RED", "note": "first"},
                {"seq": 2, "color": "BLUE", "note": "second"}
              ],
              "summary": {"line_count": 2}
            }
            """
        )
        return out

    def test_parses_with_structure(self, tmp_path, toy_registry):
        tx = parse_transaction(self._toy_source(tmp_path), registry=toy_registry)
        assert tx.definition.set_code == "999"
        assert [s.seg_id for s in tx.heading] == ["XA"]
        assert [loop.qualifier for loop in tx.loops("LQ")] == ["AA"]
        assert len(tx.loops("LD")) == 2
        assert [s.seg_id for s in tx.loops("LD")[0].segments] == ["LD", "XN"]
        assert tx.flat_scope().value("XT", 1) == "2"

    def test_validates_clean_end_to_end(self, tmp_path, toy_registry):
        tx = parse_transaction(self._toy_source(tmp_path), registry=toy_registry)
        spec = load_spec(self._toy_spec(tmp_path))
        output = load_output(self._toy_output(tmp_path))
        result = validate(spec, tx, output)
        assert result.overall is Status.PASS
        assert result.counts[Status.FAIL] == 0
        assert result.counts[Status.WARNING] == 0
        assert result.transaction_set == "999"
        # 6 header/summary rules + 3 per-line rules x 2 LD loop occurrences
        assert result.counts[Status.PASS] == 6 + 3 * 2

    def test_reconciliation_error_severity_fails(self, tmp_path, toy_registry):
        source = _write_toy_edi(
            tmp_path / "toy_bad.edi",
            ["XA*HELLO*20260701", "LD*1*RED", "XT*9"],
        )
        tx = parse_transaction(source, registry=toy_registry)
        spec = load_spec(self._toy_spec(tmp_path))
        out = tmp_path / "out.json"
        out.write_text(
            '{"order": {"greeting": "HELLO", "doc_date": "2026-07-01"},'
            ' "lines": [{"seq": 1, "color": "RED"}], "summary": {"line_count": 9}}'
        )
        result = validate(spec, tx, load_output(out))
        recon = [f for f in result.findings if f.source_ref == "recon:xt-line-count"]
        assert len(recon) == 1
        assert recon[0].status is Status.FAIL  # severity: error in the toy YAML
        assert recon[0].category is Category.COUNT_MISMATCH
        assert "XT01=9" in recon[0].expected

    def test_max_repeats_noted(self, tmp_path, toy_registry):
        source = _write_toy_edi(
            tmp_path / "toy_many.edi",
            ["XA*HI*20260701", "LD*1*A", "LD*2*B", "LD*3*C", "XT*3"],
        )
        tx = parse_transaction(source, registry=toy_registry)
        assert any("max_repeats" in note for note in tx.definition_notes)

    def test_unknown_segment_warns_but_parses(self, tmp_path, toy_registry):
        source = _write_toy_edi(
            tmp_path / "toy_extra.edi",
            ["XA*HI*20260701", "QQ*surprise", "LD*1*RED", "XN*ok", "XT*1"],
        )
        tx = parse_transaction(source, registry=toy_registry)
        assert any("segment QQ" in note for note in tx.definition_notes)
        # structure is intact around the surprise
        assert len(tx.loops("LD")) == 1
        assert [s.seg_id for s in tx.loops("LD")[0].segments] == ["LD", "XN"]


class TestAutoDetection:
    def test_unknown_st01_is_a_parse_error(self, tmp_path):
        source = _write_toy_edi(tmp_path / "mystery.edi", ["XA*HI"], set_code="811")
        with pytest.raises(X12ParseError, match="no transaction definition registered"):
            parse_transaction(source)

    def test_spec_source_mismatch(self, tmp_path, examples_dir):
        spec_path = create_template(tmp_path / "wrong.xlsx")
        wb = load_workbook(spec_path)
        wb["Mapping"]["A2"] = "X-001"
        wb["Mapping"]["B2"] = "BEG03"
        wb["Mapping"]["D2"] = "order.po_number"
        wb["Mapping"]["E2"] = "DIRECT"
        wb["Meta"]["B2"] = "855"
        wb.save(spec_path)
        with pytest.raises(SpecLoadError, match="spec is for transaction set 855"):
            validate_files(
                str(spec_path),
                str(examples_dir / "source" / "850_baseline.edi"),
                str(examples_dir / "output" / "po_baseline.json"),
            )


class TestReconciliationOn850:
    def test_defects_file_flags_ctt(self, examples_dir):
        result = validate_files(
            str(examples_dir / "specs" / "850_reference_spec.xlsx"),
            str(examples_dir / "source" / "850_defects.edi"),
            str(examples_dir / "output" / "po_from_bad_source.json"),
        )
        recon = [f for f in result.findings if f.source_ref == "recon:ctt-line-count"]
        assert len(recon) == 1
        assert recon[0].status is Status.WARNING  # default severity
        assert "CTT01=5" in recon[0].expected
        assert "count(PO1)=3" in recon[0].actual

    def test_baseline_stays_silent(self, examples_dir):
        result = validate_files(
            str(examples_dir / "specs" / "850_reference_spec.xlsx"),
            str(examples_dir / "source" / "850_baseline.edi"),
            str(examples_dir / "output" / "po_baseline.json"),
        )
        assert not [f for f in result.findings if f.source_ref.startswith("recon:")]
        assert result.transaction_set == "850"
        assert result.transaction_name == "Purchase Order"


class TestDefinitionLoaderErrors:
    def _load(self, tmp_path: Path, text: str):
        path = tmp_path / "bad.yaml"
        path.write_text(textwrap.dedent(text))
        return load_definition(path)

    def test_missing_transaction_block(self, tmp_path):
        with pytest.raises(DefinitionError, match="'transaction' block is required"):
            self._load(tmp_path, "structure: []\n")

    def test_missing_structure(self, tmp_path):
        with pytest.raises(DefinitionError, match="at least one area"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                """,
            )

    def test_later_area_needs_opener(self, tmp_path):
        with pytest.raises(DefinitionError, match="needs opens_area"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                structure:
                  - {area: header, segments: [XA]}
                  - {area: summary, segments: [XT]}
                """,
            )

    def test_bad_operand(self, tmp_path):
        with pytest.raises(DefinitionError, match="exactly one of 'value', 'count', or 'sum'"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                structure:
                  - {area: header, segments: [XA]}
                reconciliation:
                  - id: bad
                    left: {value: XA01, count: LD}
                    right: {count: LD}
                """,
            )

    def test_pairing_must_reference_declared_loop(self, tmp_path):
        with pytest.raises(DefinitionError, match="not a declared loop"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                structure:
                  - {area: header, segments: [XA]}
                output_pairing: {list_path: lines, loop: NOPE}
                """,
            )

    def test_hierarchical_requires_levels(self, tmp_path):
        with pytest.raises(DefinitionError, match="at least one level"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                structure:
                  - area: header
                    loops:
                      - id: HL
                        trigger: HL
                        hierarchical: {id_element: 1, parent_element: 2, level_element: 3}
                """,
            )

    def test_hierarchical_schema_parses(self, tmp_path):
        definition = self._load(
            tmp_path,
            """
            transaction: {set: "999", name: X, functional_group: ZZ}
            structure:
              - area: header
                loops:
                  - id: HL
                    trigger: HL
                    hierarchical: {id_element: 1, parent_element: 2, level_element: 3}
                    levels:
                      S: {children: [O], segments: [TD1]}
                      O: {children: [], segments: [PRF]}
            """,
        )
        loop = definition.loop("HL")
        assert loop.is_hierarchical
        assert loop.qualifier == 3  # qualifies on the level element
        assert loop.level("S").children == ("O",)

    def test_unknown_child_level(self, tmp_path):
        with pytest.raises(DefinitionError, match="unknown child level"):
            self._load(
                tmp_path,
                """
                transaction: {set: "999", name: X, functional_group: ZZ}
                structure:
                  - area: header
                    loops:
                      - id: HL
                        trigger: HL
                        hierarchical: {id_element: 1, parent_element: 2, level_element: 3}
                        levels:
                          S: {children: [Q]}
                """,
            )


class TestCliTransactions:
    def test_lists_registered_sets(self, capsys):
        assert main(["transactions"]) == 0
        out = capsys.readouterr().out
        assert "850" in out
        assert "Purchase Order" in out
        assert "N1, PO1" in out

    def test_validate_transaction_override(self, examples_dir, capsys):
        code = main([
            "validate",
            "--spec", str(examples_dir / "specs" / "850_reference_spec.xlsx"),
            "--source", str(examples_dir / "source" / "850_baseline.edi"),
            "--output", str(examples_dir / "output" / "po_baseline.json"),
            "--transaction", "850",
            "--no-history", "--no-color",
        ])
        assert code == 0
        assert "Transaction: 850 — Purchase Order" in capsys.readouterr().out

    def test_validate_unknown_transaction_override(self, examples_dir, capsys):
        code = main([
            "validate",
            "--spec", str(examples_dir / "specs" / "850_reference_spec.xlsx"),
            "--source", str(examples_dir / "source" / "850_baseline.edi"),
            "--output", str(examples_dir / "output" / "po_baseline.json"),
            "--transaction", "810",
            "--no-history",
        ])
        assert code == 2
        assert "no transaction definition registered" in capsys.readouterr().err
