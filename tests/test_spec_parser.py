"""Tests for the spec model, template generator, and Excel spec loader."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from mapcheck.spec import (
    DataType,
    LoopContext,
    OutcomeKind,
    RuleType,
    SpecLoadError,
    create_template,
    load_spec,
)
from mapcheck.spec.model import split_source_field
from mapcheck.spec.template import MAPPING_COLUMNS


class TestModelHelpers:
    def test_split_source_field(self):
        assert split_source_field("BEG03") == ("BEG", 3)
        assert split_source_field("PO102") == ("PO1", 2)
        assert split_source_field("N104") == ("N1", 4)

    @pytest.mark.parametrize("bad", ["BEG", "beg03", "BEG3", "TOOLONG01", "03"])
    def test_split_source_field_rejects(self, bad):
        with pytest.raises(ValueError):
            split_source_field(bad)

    def test_loop_context_parse(self):
        ctx = LoopContext.parse("N1[ST]")
        assert (ctx.segment_id, ctx.qualifier) == ("N1", "ST")
        assert LoopContext.parse("PO1").qualifier is None
        assert LoopContext.parse("DTM[002]").qualifier == "002"

    @pytest.mark.parametrize("bad", ["", "n1[st]", "N1[ST", "N1[]", "TOOLONG"])
    def test_loop_context_rejects(self, bad):
        with pytest.raises(ValueError):
            LoopContext.parse(bad)


class TestTemplate:
    def test_create_template_sheets(self, tmp_path: Path):
        path = create_template(tmp_path / "blank.xlsx")
        wb = load_workbook(path)
        assert set(wb.sheetnames) >= {"Mapping", "CodeLists", "Meta", "Instructions"}
        headers = [c.value for c in wb["Mapping"][1]][: len(MAPPING_COLUMNS)]
        assert headers == [h for h, _ in MAPPING_COLUMNS]


class TestReferenceSpec:
    """The generated reference spec is itself a fixture for the whole suite."""

    def test_loads_clean(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        assert spec.transaction_set == "850"
        assert len(spec.rules) == 30
        assert set(spec.code_lists) == {"TX_PURPOSE", "PO_TYPE", "UOM"}

    def test_covers_every_rule_type(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        assert {r.rule_type for r in spec.rules} == set(RuleType)

    def test_conditional_rule_parsed(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        (rule,) = [r for r in spec.rules if r.row_id == "M-003"]
        assert rule.rule_type is RuleType.CONDITIONAL
        assert rule.condition is not None
        assert rule.condition.predicates[0].field == "BEG02"
        assert rule.then_outcome.kind is OutcomeKind.LITERAL
        assert rule.then_outcome.literal == "Y"
        assert rule.else_outcome.literal == "N"

    def test_else_defaults_to_skip(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        (rule,) = [r for r in spec.rules if r.row_id == "M-009"]
        assert rule.else_outcome.kind is OutcomeKind.SKIP

    def test_per_line_rules(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        per_line = [r for r in spec.rules if r.is_per_line]
        assert {r.loop_context.segment_id for r in per_line} == {"PO1"}
        assert all(r.target_field.startswith("lines[]") for r in per_line)

    def test_code_list_translation(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        assert spec.code_lists["UOM"].translate("EA") == "EACH"
        assert spec.code_lists["UOM"].translate("ZZ") is None

    def test_data_types(self, reference_spec_path: Path):
        spec = load_spec(reference_spec_path)
        by_id = {r.row_id: r for r in spec.rules}
        assert by_id["M-005"].data_type is DataType.DATE
        assert by_id["M-013"].format == "implied:2;places:2"
        assert by_id["M-023"].data_type is DataType.INTEGER


def _mutate_spec(reference_spec_path: Path, tmp_path: Path, mutate) -> Path:
    """Copy the reference spec, apply ``mutate(workbook)``, save, return path."""
    wb = load_workbook(reference_spec_path)
    mutate(wb)
    out = tmp_path / "mutated.xlsx"
    wb.save(out)
    return out


class TestLoadErrors:
    def _expect_error(self, path: Path, fragment: str):
        with pytest.raises(SpecLoadError) as exc_info:
            load_spec(path)
        assert any(fragment in e for e in exc_info.value.errors), exc_info.value.errors

    def test_missing_file(self, tmp_path: Path):
        self._expect_error(tmp_path / "nope.xlsx", "not found")

    def test_missing_sheet(self, reference_spec_path, tmp_path):
        path = _mutate_spec(reference_spec_path, tmp_path, lambda wb: wb.remove(wb["CodeLists"]))
        self._expect_error(path, "missing required sheet 'CodeLists'")

    def test_header_tampered(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["B1"] = "Src"

        self._expect_error(_mutate_spec(reference_spec_path, tmp_path, mutate), "header mismatch")

    def test_unknown_rule_type(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["E2"] = "MAGIC"

        self._expect_error(_mutate_spec(reference_spec_path, tmp_path, mutate), "unknown rule type")

    def test_duplicate_row_id(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["A3"] = wb["Mapping"]["A2"].value

        self._expect_error(_mutate_spec(reference_spec_path, tmp_path, mutate), "duplicate Row ID")

    def test_bad_source_field(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["B4"] = "BEG"  # M-003 row

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate), "invalid source field"
        )

    def test_bad_loop_context(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["C8"] = "REF[DP"

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate), "invalid loop context"
        )

    def test_missing_code_list(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["K2"] = "NO_SUCH_LIST"

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate), "not found on the CodeLists sheet"
        )

    def test_constant_requires_default(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["J13"] = None  # M-012 CONSTANT row

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate), "CONSTANT rule requires"
        )

    def test_conditional_requires_condition(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["G4"] = None  # M-003 conditional row

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate),
            "CONDITIONAL rule requires a Condition",
        )

    def test_condition_on_direct_rule_rejected(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["G5"] = "BEG02 = 'SA'"  # M-004 DIRECT row

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate),
            "only allowed on CONDITIONAL rules",
        )

    def test_per_line_target_requires_loop_context(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["C23"] = None  # M-022 lines[].line_no

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate),
            "requires a repeating Loop Context",
        )

    def test_repeating_context_requires_per_line_target(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["D23"] = "order.first_line_no"

        self._expect_error(
            _mutate_spec(reference_spec_path, tmp_path, mutate),
            "requires a per-line target",
        )

    def test_all_errors_reported_together(self, reference_spec_path, tmp_path):
        def mutate(wb):
            wb["Mapping"]["E2"] = "MAGIC"
            wb["Mapping"]["B4"] = "BEG"
            wb["Mapping"]["A3"] = None  # missing row id

        path = _mutate_spec(reference_spec_path, tmp_path, mutate)
        with pytest.raises(SpecLoadError) as exc_info:
            load_spec(path)
        assert len(exc_info.value.errors) >= 3
