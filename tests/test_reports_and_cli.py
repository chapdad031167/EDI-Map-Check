"""Tests for terminal rendering, Excel export, SQLite history, and the CLI."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from mapcheck.cli import main
from mapcheck.engine import Status, validate_files
from mapcheck.report.excel import export_excel
from mapcheck.report.history import RunHistory
from mapcheck.report.terminal import render_report


@pytest.fixture(scope="module")
def defective_run(examples_dir: Path):
    return validate_files(
        str(examples_dir / "specs" / "850_reference_spec.xlsx"),
        str(examples_dir / "source" / "850_baseline.edi"),
        str(examples_dir / "output" / "po_baseline_defects.json"),
    )


class TestTerminalReport:
    def test_contains_summary_and_root_causes(self, defective_run):
        text = render_report(defective_run, color=False)
        assert "RESULT: FAIL" in text
        assert "Root causes:" in text
        assert "value_mismatch" in text
        assert "M-004" in text

    def test_verbose_includes_pass(self, defective_run):
        quiet = render_report(defective_run, color=False)
        verbose = render_report(defective_run, verbose=True, color=False)
        assert "PASS" not in quiet.split("Summary:")[0].split("DETAIL")[1]
        assert verbose.count("\nPASS") > 10

    def test_color_codes_only_when_enabled(self, defective_run):
        assert "\033[" not in render_report(defective_run, color=False)
        assert "\033[" in render_report(defective_run, color=True)


class TestExcelExport:
    def test_workbook_structure(self, defective_run, tmp_path: Path):
        path = export_excel(defective_run, tmp_path / "report.xlsx")
        wb = load_workbook(path)
        assert wb.sheetnames == ["Results", "Summary"]
        ws = wb["Results"]
        headers = [c.value for c in ws[1]]
        assert headers[:4] == ["Status", "Row ID", "Source", "Target"]
        assert ws.auto_filter.ref is not None
        # FAIL rows come first and are color-coded
        assert ws["A2"].value == "FAIL"
        assert ws["A2"].fill.fgColor.rgb.endswith("FFC7CE")
        statuses = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        assert statuses == {"FAIL", "WARNING", "PASS"}

    def test_summary_sheet(self, defective_run, tmp_path: Path):
        path = export_excel(defective_run, tmp_path / "report.xlsx")
        ws = load_workbook(path)["Summary"]
        cells = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        assert "Overall result" in cells
        assert "value_mismatch" in cells


class TestHistory:
    def test_record_and_read_back(self, defective_run, tmp_path: Path):
        db = tmp_path / "history.db"
        with RunHistory(db) as history:
            run_id = history.record(defective_run)
            runs = history.recent_runs()
            assert len(runs) == 1
            assert runs[0]["id"] == run_id
            assert runs[0]["result"] == "FAIL"
            assert runs[0]["failed"] == defective_run.counts[Status.FAIL]
            findings = history.findings_for(run_id)
            assert len(findings) == len(defective_run.findings)
            assert findings[0]["status"] == "FAIL"

    def test_multiple_runs_newest_first(self, defective_run, tmp_path: Path):
        db = tmp_path / "history.db"
        with RunHistory(db) as history:
            first = history.record(defective_run)
            second = history.record(defective_run)
            runs = history.recent_runs()
            assert [r["id"] for r in runs] == [second, first]


class TestCli:
    def _validate_args(self, examples_dir: Path, output_name: str, *extra: str) -> list[str]:
        return [
            "validate",
            "--spec", str(examples_dir / "specs" / "850_reference_spec.xlsx"),
            "--source", str(examples_dir / "source" / "850_baseline.edi"),
            "--output", str(examples_dir / "output" / output_name),
            "--no-color",
            *extra,
        ]

    def test_validate_pass_exit_zero(self, examples_dir, tmp_path, capsys, monkeypatch):
        monkeypatch.chdir(tmp_path)
        code = main(self._validate_args(examples_dir, "po_baseline.json"))
        captured = capsys.readouterr()
        assert code == 0
        assert "RESULT: PASS" in captured.out
        assert (tmp_path / "mapcheck_history.db").exists()

    def test_validate_fail_exit_one(self, examples_dir, tmp_path, capsys, monkeypatch):
        monkeypatch.chdir(tmp_path)
        code = main(self._validate_args(examples_dir, "po_baseline_defects.json", "--no-history"))
        captured = capsys.readouterr()
        assert code == 1
        assert "RESULT: FAIL" in captured.out
        assert not (tmp_path / "mapcheck_history.db").exists()

    def test_validate_with_export(self, examples_dir, tmp_path, capsys, monkeypatch):
        monkeypatch.chdir(tmp_path)
        report = tmp_path / "report.xlsx"
        code = main(
            self._validate_args(
                examples_dir, "po_baseline_defects.json", "--no-history",
                "--export-xlsx", str(report),
            )
        )
        assert code == 1
        assert report.exists()

    def test_validate_bad_spec_path(self, examples_dir, tmp_path, capsys):
        code = main([
            "validate",
            "--spec", str(tmp_path / "nope.xlsx"),
            "--source", str(examples_dir / "source" / "850_baseline.edi"),
            "--output", str(examples_dir / "output" / "po_baseline.json"),
        ])
        captured = capsys.readouterr()
        assert code == 2
        assert "not found" in captured.err

    def test_init_spec(self, tmp_path, capsys):
        target = tmp_path / "new_spec.xlsx"
        assert main(["init-spec", str(target)]) == 0
        assert target.exists()
        assert main(["init-spec", str(target)]) == 2  # refuses to overwrite

    def test_history_command(self, examples_dir, tmp_path, capsys, monkeypatch):
        monkeypatch.chdir(tmp_path)
        main(self._validate_args(examples_dir, "po_baseline.json"))
        capsys.readouterr()
        assert main(["history"]) == 0
        captured = capsys.readouterr()
        assert "PASS" in captured.out
        assert "850_baseline.edi -> po_baseline.json" in captured.out

    def test_history_missing_db(self, tmp_path, capsys, monkeypatch):
        monkeypatch.chdir(tmp_path)
        assert main(["history"]) == 2
