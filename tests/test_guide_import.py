"""Guide import (Design 014): parser, profile, overlay, and guided drafts.

The synthetic fixture guide (tests/fixtures/guides/) is the golden input:
same guide in, same profile out — deterministically. Flag-never-guess is
tested as behavior: lines the grammar cannot hold land in ``review`` with
page context, never in the data. The PDF twin must extract to the exact
same profile as the .txt (parity), because both are one guide family.
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from mapcheck.guides.overlay import (
    PartnerRules,
    PartnerRulesError,
    emit_partner_rules,
)
from mapcheck.guides.parser import GuideParseError, parse_guide
from mapcheck.guides.profile import GuideProfile, GuideProfileError
from mapcheck.output.idoc import default_output_registry
from mapcheck.spec.draft import (
    Crosswalk,
    assemble_draft,
    bundled_crosswalks,
    generate_draft,
    load_crosswalks,
)
from mapcheck.transactions.registry import default_registry
from mapcheck.transactions.schema import RequiredElementDef, RequiredSegmentDef

REPO = Path(__file__).resolve().parent.parent
GUIDES = REPO / "tests" / "fixtures" / "guides"
GUIDE_TXT = GUIDES / "acme_pharma_850_guide.txt"
GUIDE_PDF = GUIDES / "acme_pharma_850_guide.pdf"

DEFINITION_850 = default_registry.get("850")
ORDERS05 = default_output_registry.get("idoc-orders05")


def _profile() -> GuideProfile:
    return parse_guide(GUIDE_TXT, transaction="850", partner="acme_pharma")


# --------------------------------------------------------------------------
# Family detection and extraction
# --------------------------------------------------------------------------


class TestFamilyDetection:
    def test_non_guide_text_is_rejected_naming_fingerprints(self, tmp_path: Path):
        other = tmp_path / "notes.txt"
        other.write_text("Meeting notes\nDiscussed the 850 project.\n", encoding="utf-8")
        with pytest.raises(GuideParseError) as err:
            parse_guide(other, transaction="850", partner="x")
        message = str(err.value)
        assert "not a recognized implementation-guide layout" in message
        assert "segment header block" in message
        assert "Element Summary" in message

    def test_unsupported_suffix_is_rejected(self, tmp_path: Path):
        doc = tmp_path / "guide.docx"
        doc.write_bytes(b"PK\x03\x04")
        with pytest.raises(GuideParseError, match="unsupported guide format"):
            parse_guide(doc, transaction="850", partner="x")

    def test_missing_file(self):
        with pytest.raises(GuideParseError, match="not found"):
            parse_guide(GUIDES / "no_such_guide.txt", transaction="850", partner="x")


# --------------------------------------------------------------------------
# Golden profile from the synthetic fixture
# --------------------------------------------------------------------------


class TestGoldenProfile:
    def test_segment_inventory(self):
        profile = _profile()
        assert [seg.id for seg in profile.segments] == [
            "BEG", "CUR", "REF", "PER", "DTM", "N1", "N3", "N4",
            "PO1", "PID", "SAC", "CTT",
        ]
        assert profile.transaction == "850"
        assert profile.partner == "acme_pharma"
        assert profile.source == GUIDE_TXT.name

    def test_usages_and_loops(self):
        profile = _profile()
        usage = {seg.id: seg.usage for seg in profile.segments}
        assert usage["BEG"] == "must_use"
        assert usage["CUR"] == "used"
        assert usage["SAC"] == "not_used"
        assert profile.segment("N3").loop == "N1"
        assert profile.segment("PO1").loop == "PO1"
        assert profile.segment("BEG").loop == ""  # printed as N/A

    def test_element_rows(self):
        beg = _profile().segment("BEG")
        assert [el.ref for el in beg.elements] == [
            "BEG01", "BEG02", "BEG03", "BEG04", "BEG05",
        ]
        beg04 = beg.element("BEG04")
        assert beg04.usage == "not_used"
        assert beg04.req == "O"
        assert beg04.type == "AN"
        assert (beg04.min, beg04.max) == (1, 30)

    def test_code_subsets_attach_to_their_element(self):
        profile = _profile()
        beg02 = profile.segment("BEG").element("BEG02")
        assert [(c.code, c.name) for c in beg02.codes] == [("NE", "New Order")]
        po1 = profile.segment("PO1")
        assert [c.code for c in po1.element("PO106").codes] == ["VN"]
        assert [c.code for c in po1.element("PO108").codes] == ["UP"]
        assert po1.element("PO107").codes == ()

    def test_notes_attach_to_element_and_flow_with_label_stripped(self):
        profile = _profile()
        beg05 = profile.segment("BEG").element("BEG05")
        assert beg05.notes == (
            "This is the date assigned by Acme Pharma to the purchase order.",
        )
        po109 = profile.segment("PO1").element("PO109")
        assert po109.notes == (
            "Every PO1 line must carry a UP qualifier pair with the UPC.",
        )

    def test_full_coverage_and_empty_review(self):
        profile = _profile()
        assert profile.review == []
        assert profile.parse_coverage == 1.0
        assert profile.facts_detected == 64

    def test_deterministic(self):
        assert _profile().to_dict() == _profile().to_dict()


class TestPdfParity:
    def test_pdf_twin_extracts_identically(self):
        pytest.importorskip("pdfplumber")
        txt = _profile().to_dict()
        pdf = parse_guide(GUIDE_PDF, transaction="850", partner="acme_pharma").to_dict()
        txt.pop("source")
        pdf.pop("source")
        assert txt == pdf


# --------------------------------------------------------------------------
# Split-form headers: how real PDFs of this family extract (two-column
# header boxes flatten to a "Pos: … Max: …" line, then the name line, the
# name's wrapped tail landing on the Loop line)
# --------------------------------------------------------------------------

SPLIT_FORM_GUIDE = """Fixture Guide
Purchase Order - 850

Pos: 020 Max: 1
BEG Beginning Segment for
Heading - Mandatory
Purchase Order Loop: N/A Elements: 2
User Option (Usage): Must use

Element Summary:
BEG01 353 Transaction Set Purpose Code M ID 2/2 Must use
BEG03 324 Purchase Order Number M AN 1/22 Must use

Pos: 240 Max: 1
TD5 Carrier Details (Routing
Heading - Mandatory
Sequence/Transit Time) Loop: N/A Elements: 1
User Option (Usage): Must use

Element Summary:
TD501 133 Routing Sequence Code O ID 1/2 Used

Pos: 310 Max: 1
N1 Name
Heading - Mandatory
Loop: N1 Elements: 2
User Option (Usage): Must use

Element Summary:
N101 98 Entity Identifier Code M ID 2/3 Must use
Code List Summary (Total Codes: 1312, Included: 1)
Code Name
ST Ship To
N102 93 Name O AN 1/60 Must use

Pos: 310 Max: 1
N1 Name
Heading - Mandatory
Loop: N1 Elements: 2
User Option (Usage): Must use

Element Summary:
N101 98 Entity Identifier Code M ID 2/3 Must use
Code List Summary (Total Codes: 1312, Included: 1)
Code Name
BT Bill To
N102 93 Name O AN 1/60 Must use
"""


class TestSplitFormHeaders:
    def _parse(self, tmp_path: Path, text: str) -> GuideProfile:
        guide = tmp_path / "split.txt"
        guide.write_text(text, encoding="utf-8")
        return parse_guide(guide, transaction="850", partner="x")

    def test_split_header_block_parses(self, tmp_path: Path):
        profile = self._parse(tmp_path, SPLIT_FORM_GUIDE)
        assert profile.review == []
        assert profile.parse_coverage == 1.0
        beg = profile.segment("BEG")
        assert beg.pos == "020"
        assert beg.max_use == "1"
        assert beg.usage == "must_use"

    def test_wrapped_name_reassembles_from_loop_line(self, tmp_path: Path):
        profile = self._parse(tmp_path, SPLIT_FORM_GUIDE)
        assert profile.segment("BEG").name == "Beginning Segment for Purchase Order"
        assert (
            profile.segment("TD5").name
            == "Carrier Details (Routing Sequence/Transit Time)"
        )
        assert profile.segment("N1").name == "Name"  # no tail, nothing appended

    def test_one_block_per_loop_occurrence(self, tmp_path: Path):
        profile = self._parse(tmp_path, SPLIT_FORM_GUIDE)
        n1_blocks = [seg for seg in profile.segments if seg.id == "N1"]
        assert len(n1_blocks) == 2
        assert [b.element("N101").codes[0].code for b in n1_blocks] == ["ST", "BT"]

    def test_occurrence_blocks_emit_distinct_qualified_rules(self, tmp_path: Path):
        profile = self._parse(tmp_path, SPLIT_FORM_GUIDE)
        rules = emit_partner_rules(profile)
        n1_rules = [(r.segment, r.qualifier) for r in rules.required_segments if r.segment == "N1"]
        assert n1_rules == [("N1", "ST"), ("N1", "BT")]

    def test_identical_duplicate_blocks_emit_once(self, tmp_path: Path):
        doubled = SPLIT_FORM_GUIDE + SPLIT_FORM_GUIDE
        profile = self._parse(tmp_path, doubled)
        rules = emit_partner_rules(profile)
        pairs = [(r.segment, r.qualifier) for r in rules.required_segments]
        assert len(pairs) == len(set(pairs))
        # scoped rules for different qualifiers may share (segment, element);
        # the identity of an element rule includes its qualifier
        elements = [
            (r.segment, r.element, r.qualifier) for r in rules.required_elements
        ]
        assert len(elements) == len(set(elements))

    def test_pos_line_without_name_line_is_review(self, tmp_path: Path):
        profile = self._parse(
            tmp_path,
            "Pos: 020 Max: 1\n12/5/2023 page header\n" + SPLIT_FORM_GUIDE,
        )
        assert any(
            "'Pos: 020 Max: 1' was not followed by a segment name" in note
            for note in profile.review
        )
        assert profile.parse_coverage < 1.0
        assert profile.segment("BEG") is not None  # the real block still parses

    def test_loop_prefix_on_one_line_header_is_review(self, tmp_path: Path):
        guide = tmp_path / "g.txt"
        guide.write_text(
            "BEG Beginning Segment for Purchase Order Pos: 020 Max: 1\n"
            "Stray tail Loop: N/A Elements: 2\n"
            "User Option (Usage): Must use\n\n"
            "Element Summary:\n"
            "BEG01 353 Transaction Set Purpose Code M ID 2/2 Must use\n",
            encoding="utf-8",
        )
        profile = parse_guide(guide, transaction="850", partner="x")
        assert profile.segment("BEG").name == "Beginning Segment for Purchase Order"
        assert any("unexpected text before 'Loop:'" in note for note in profile.review)


class TestOverstrikeDedupe:
    def test_bold_overstrike_text_extracts_once(self, tmp_path: Path):
        """Real guides print bold lines as overstruck duplicate glyphs;
        extraction must collapse them or notes come out doubled."""
        pytest.importorskip("pdfplumber")
        fpdf = pytest.importorskip("fpdf")
        from mapcheck.guides.parser import extract_lines

        pdf = fpdf.FPDF()
        pdf.add_page()
        pdf.set_font("Courier", size=10)
        pdf.text(10, 20, "Plain line stays intact")
        # simulate bold-by-overprint: same string twice, offset well inside
        # dedupe tolerance
        pdf.text(10, 30, "Insight Note: required on every order.")
        pdf.text(10.05, 30, "Insight Note: required on every order.")
        path = tmp_path / "overstrike.pdf"
        pdf.output(str(path))

        lines = [line for _, line in extract_lines(path)]
        assert "Plain line stays intact" in lines
        assert "Insight Note: required on every order." in lines
        assert not any("IInnssiigghhtt" in line for line in lines)


# --------------------------------------------------------------------------
# Flag-never-guess: uncertain lines go to review, never into the data
# --------------------------------------------------------------------------


def _guide_text(*blocks: str) -> str:
    head = (
        "Fixture Guide\n\n"
        "BEG Beginning Segment for Purchase Order Pos: 020 Max: 1\n"
        "User Option (Usage): Must use\n\n"
        "Element Summary:\n"
        "BEG01 353 Transaction Set Purpose Code M ID 2/2 Must use\n"
    )
    return head + "".join(blocks)


class TestFlagNeverGuess:
    def test_mangled_element_row_is_review_not_data(self, tmp_path: Path):
        guide = tmp_path / "g.txt"
        guide.write_text(
            _guide_text("BEG02 92 Purchase Order Type Code M ID 2/2\n"),
            encoding="utf-8",
        )
        profile = parse_guide(guide, transaction="850", partner="x")
        beg = profile.segment("BEG")
        assert beg.element("BEG02") is None
        assert len(profile.review) == 1
        assert "BEG02" in profile.review[0]
        assert "page 1" in profile.review[0]
        assert profile.parse_coverage < 1.0

    def test_unrecognized_segment_usage_is_review(self, tmp_path: Path):
        guide = tmp_path / "g.txt"
        guide.write_text(
            _guide_text(
                "\nCUR Currency Pos: 040 Max: 1\n"
                "User Option (Usage): Sometimes\n\n"
                "Element Summary:\n"
                "CUR01 98 Entity Identifier Code M ID 2/3 Used\n"
            ),
            encoding="utf-8",
        )
        profile = parse_guide(guide, transaction="850", partner="x")
        assert profile.segment("CUR").usage == ""
        assert any("Sometimes" in note for note in profile.review)

    def test_orphan_code_row_is_review(self, tmp_path: Path):
        guide = tmp_path / "g.txt"
        guide.write_text(
            _guide_text(
                "\nCUR Currency Pos: 040 Max: 1\n"
                "User Option (Usage): Used\n"
                "Code List Summary (Total Codes: 3, Included: 1)\n"
                "Code Name\n"
                "XX Mystery\n"
                "\nElement Summary:\n"
                "CUR01 98 Entity Identifier Code M ID 2/3 Used\n"
            ),
            encoding="utf-8",
        )
        profile = parse_guide(guide, transaction="850", partner="x")
        assert any("code row with no open element" in note for note in profile.review)
        assert all(not el.codes for el in profile.segment("CUR").elements)


# --------------------------------------------------------------------------
# Profile serialization
# --------------------------------------------------------------------------


class TestProfileRoundTrip:
    def test_save_load_round_trip(self, tmp_path: Path):
        profile = _profile()
        path = profile.save(tmp_path / "acme.yaml")
        again = GuideProfile.load(path)
        assert again.to_dict() == profile.to_dict()
        assert again.parse_coverage == profile.parse_coverage

    def test_load_rejects_non_profile(self, tmp_path: Path):
        path = tmp_path / "other.yaml"
        path.write_text("just: text\n", encoding="utf-8")
        with pytest.raises(GuideProfileError, match="missing 'transaction'"):
            GuideProfile.load(path)

    def test_load_names_every_problem(self, tmp_path: Path):
        path = tmp_path / "bad.yaml"
        path.write_text(
            yaml.safe_dump(
                {
                    "transaction": "850",
                    "segments": [
                        {"name": "no id"},
                        {"id": "BEG", "usage": "sometimes"},
                        {"id": "REF", "elements": [{"ref": "REF01", "usage": "meh"}]},
                    ],
                }
            ),
            encoding="utf-8",
        )
        with pytest.raises(GuideProfileError) as err:
            GuideProfile.load(path)
        message = str(err.value)
        assert "3 problem(s)" in message
        assert "segments[0]" in message
        assert "segments[1]" in message
        assert "segments[2].elements[0]" in message


# --------------------------------------------------------------------------
# Overlay emission and apply
# --------------------------------------------------------------------------


class TestOverlayEmission:
    def test_qualified_segments_pin_single_code(self):
        rules = emit_partner_rules(_profile())
        by_segment = {(r.segment, r.qualifier): r for r in rules.required_segments}
        assert ("REF", "IA") in by_segment
        assert ("DTM", "002") in by_segment
        assert by_segment[("DTM", "002")].name == "Requested Delivery"
        assert by_segment[("DTM", "002")].origin == "acme_pharma"
        # BEG/N1/PO1/CTT are unqualified requirements
        assert ("BEG", None) in by_segment
        assert ("N1", None) in by_segment  # N101 carries no code subset

    def test_element_rules_from_unqualified_segments(self):
        rules = emit_partner_rules(_profile())
        refs = {(r.segment, r.element) for r in rules.required_elements}
        assert ("CTT", 2) in refs
        assert ("BEG", 4) not in refs  # Not used, never required
        # pinned pair slots become pair rules, not positional requirements
        assert ("PO1", 6) not in refs and ("PO1", 7) not in refs
        assert ("PO1", 8) not in refs and ("PO1", 9) not in refs
        assert ("PO1", 2) in refs  # non-pair elements stay positional

    def test_qualified_segments_get_scoped_element_rules(self):
        """Design 015: 'REF02 required in REF*IA occurrences' is a rule
        now, not a review note."""
        rules = emit_partner_rules(_profile())
        scoped = {
            (r.segment, r.element, r.qualifier)
            for r in rules.required_elements
            if r.qualifier is not None
        }
        assert ("REF", 2, "IA") in scoped
        assert ("DTM", 2, "002") in scoped
        assert rules.review == []  # nothing left the schema cannot express

    def test_pinned_pair_slots_become_pair_rules(self):
        rules = emit_partner_rules(_profile())
        pairs = {(r.segment, r.code): r for r in rules.required_pairs}
        assert set(pairs) == {("PO1", "VN"), ("PO1", "UP")}
        up = pairs[("PO1", "UP")]
        assert up.name == "UPC Consumer Package Code"
        assert (up.first_element, up.last_element, up.step) == (6, 24, 2)
        assert up.origin == "acme_pharma"

    def test_multi_code_qualifier_emits_a_code_set(self, tmp_path: Path):
        """Design 018: a REF block listing {IA, DP} emits a qualifiers set
        ('a REF whose REF01 is IA or DP'), not an unqualified rule."""
        guide = tmp_path / "g.txt"
        guide.write_text(
            "REF Reference Identification Pos: 050 Max: 1\n"
            "User Option (Usage): Must use\n\n"
            "Element Summary:\n"
            "REF01 128 Reference Identification Qualifier M ID 2/3 Must use\n"
            "Code List Summary (Total Codes: 320, Included: 2)\n"
            "Code Name\n"
            "IA Internal Vendor Number\n"
            "DP Department Number\n"
            "REF02 127 Reference Identification M AN 1/30 Must use\n",
            encoding="utf-8",
        )
        rules = emit_partner_rules(parse_guide(guide, transaction="850", partner="x"))
        assert len(rules.required_segments) == 1
        seg = rules.required_segments[0]
        assert seg.segment == "REF"
        assert seg.qualifier is None
        assert seg.qualifiers == ("IA", "DP")
        assert not any("multiple qualifier codes" in note for note in rules.review)
        # REF02 becomes a code-set-scoped element rule, not unqualified
        ref02 = [r for r in rules.required_elements if r.element == 2]
        assert len(ref02) == 1
        assert ref02[0].qualifiers == ("IA", "DP")

    def test_apply_dedups_against_unconditional_base_only(self):
        rules = emit_partner_rules(_profile())
        base = DEFINITION_850
        base_pairs = {(r.segment, r.element, r.when_present) for r in base.required_elements}
        assert ("BEG", 3, None) in base_pairs  # unconditional base rule
        assert ("PO1", 2, 3) in base_pairs  # conditional base rule
        merged = rules.apply(base)
        merged_partner = [
            (r.segment, r.element) for r in merged.required_elements if r.origin
        ]
        assert ("BEG", 3) not in merged_partner  # base already enforces it
        assert ("PO1", 2) in merged_partner  # conditional base does not
        assert len(merged.required_segments) == len(base.required_segments) + 6
        assert {(r.segment, r.code) for r in merged.required_pairs} == {
            ("PO1", "VN"), ("PO1", "UP"),
        }

    def test_apply_rejects_wrong_transaction(self):
        rules = emit_partner_rules(_profile())
        other = default_registry.get("810")
        with pytest.raises(PartnerRulesError, match="850.*810"):
            rules.apply(other)


class TestOverlayRoundTrip:
    def test_save_load_round_trip(self, tmp_path: Path):
        rules = emit_partner_rules(_profile())
        path = rules.save(tmp_path / "acme_rules.yaml")
        again = PartnerRules.load(path)
        assert again.to_dict() == rules.to_dict()
        assert all(r.origin == "acme_pharma" for r in again.required_segments)

    def test_load_names_every_problem(self, tmp_path: Path):
        path = tmp_path / "bad.yaml"
        path.write_text(
            yaml.safe_dump(
                {
                    "transaction": "850",
                    "partner": "x",
                    "required_segments": [{"qualifier": "IA"}],
                    "required_elements": [{"segment": "REF", "element": 0}],
                }
            ),
            encoding="utf-8",
        )
        with pytest.raises(PartnerRulesError) as err:
            PartnerRules.load(path)
        message = str(err.value)
        assert "required_segments[0]" in message
        assert "required_elements[0]" in message

    def test_committed_example_overlay_loads(self):
        rules = PartnerRules.load(REPO / "examples" / "partner_rules" / "acme_pharma_850.yaml")
        assert rules.partner == "acme_pharma"
        assert len(rules.required_segments) == 6
        profile = GuideProfile.load(
            REPO / "examples" / "partner_rules" / "acme_pharma_850_profile.yaml"
        )
        assert profile.parse_coverage == 1.0


# --------------------------------------------------------------------------
# Required-segment enforcement (definition-level, no spec involved)
# --------------------------------------------------------------------------


class TestRequiredSegmentLoader:
    def test_definition_yaml_round_trip(self, tmp_path: Path):
        from mapcheck.transactions.loader import parse_definition

        data = {
            "transaction": {"set": "850", "name": "PO", "functional_group": "PO"},
            "structure": [{"area": "heading", "segments": ["BEG"]}],
            "required_segments": [
                {"segment": "dtm", "qualifier": "002", "name": "Requested Delivery"},
                {"segment": "PO1"},
            ],
        }
        definition = parse_definition(data, source="inline")
        assert definition.required_segments == (
            RequiredSegmentDef(segment="DTM", qualifier="002", name="Requested Delivery"),
            RequiredSegmentDef(segment="PO1"),
        )

    def test_loader_rejects_bad_entries(self):
        from mapcheck.transactions.loader import DefinitionError, parse_definition

        data = {
            "transaction": {"set": "850", "name": "PO", "functional_group": "PO"},
            "structure": [{"area": "heading", "segments": ["BEG"]}],
            "required_segments": [{"qualifier": "IA"}, {"segment": "REF", "qualifier_element": 0}],
        }
        with pytest.raises(DefinitionError) as err:
            parse_definition(data, source="inline")
        message = str(err.value)
        assert "required_segments[0]" in message
        assert "required_segments[1].qualifier_element" in message


# --------------------------------------------------------------------------
# The tabular family (Design 017): Gentran-style "Data Element Summary"
# --------------------------------------------------------------------------

TABULAR_TXT = GUIDES / "bluewater_medical_850_guide.txt"
TABULAR_PDF = GUIDES / "bluewater_medical_850_guide.pdf"


def _tabular_profile() -> GuideProfile:
    return parse_guide(TABULAR_TXT, transaction="850", partner="bluewater_medical")


class TestFamilyRouting:
    def test_each_fixture_routes_to_its_family(self):
        assert _profile().family == "templated"
        assert _tabular_profile().family == "tabular"

    def test_rejection_names_both_families(self, tmp_path: Path):
        other = tmp_path / "notes.txt"
        other.write_text("Meeting notes\nDiscussed the 850 project.\n", encoding="utf-8")
        with pytest.raises(GuideParseError) as err:
            parse_guide(other, transaction="850", partner="x")
        message = str(err.value)
        assert "templated (Element Summary layout)" in message
        assert "tabular (Data Element Summary layout)" in message
        assert "Data Element Summary' marker" in message

    def test_matching_both_families_is_a_defined_error(self, tmp_path: Path):
        both = tmp_path / "both.txt"
        both.write_text(
            SPLIT_FORM_GUIDE
            + "\nSegment: Impossible Mix\nUsage: Mandatory\nData Element Summary\n"
            "ZZ01 999 Mixed Row M ID 2/2 M\n",
            encoding="utf-8",
        )
        with pytest.raises(GuideParseError, match="two guide layout families"):
            parse_guide(both, transaction="850", partner="x")

    def test_family_round_trips_through_yaml(self, tmp_path: Path):
        profile = _tabular_profile()
        path = profile.save(tmp_path / "b.yaml")
        assert GuideProfile.load(path).family == "tabular"


class TestTabularGolden:
    def test_segment_inventory_and_coverage(self):
        profile = _tabular_profile()
        assert [seg.id for seg in profile.segments] == [
            "BEG", "REF", "DTM", "N1", "N3", "PO1", "CTT",
        ]
        assert profile.parse_coverage == 1.0
        assert profile.review == []

    def test_user_attribute_overrides_base(self):
        """The maintainer-confirmed rule: User column wins when present."""
        profile = _tabular_profile()
        # base C, User M — the partner made a conditional segment mandatory
        assert profile.segment("REF").usage == "must_use"
        # base O, User M on an element
        po101 = profile.segment("PO1").element("PO101")
        assert (po101.req, po101.usage) == ("M", "must_use")

    def test_absent_user_attribute_falls_back_to_base(self):
        profile = _tabular_profile()
        beg05 = profile.segment("BEG").element("BEG05")
        assert (beg05.req, beg05.usage) == ("M", "must_use")  # base M
        ref02 = profile.segment("REF").element("REF02")
        assert (ref02.req, ref02.usage) == ("C", "used")  # base C
        n302 = profile.segment("N3").element("N302")
        assert (n302.req, n302.usage) == ("O", "used")  # base O

    def test_wrapped_code_names_reassemble(self):
        profile = _tabular_profile()
        n103 = profile.segment("N1").element("N103")
        assert [c.name for c in n103.codes] == [
            "D-U-N-S+4, D-U-N-S Number with Four Character Suffix"
        ]
        po106 = profile.segment("PO1").element("PO106")
        assert ("UK", "U.P.C. Shipping Container Code (1-2-5-5-1)") in [
            (c.code, c.name) for c in po106.codes
        ]

    def test_codes_only_under_id_elements(self):
        profile = _tabular_profile()
        for seg in profile.segments:
            for el in seg.elements:
                if el.type != "ID":
                    assert el.codes == ()

    def test_notes_kept_examples_filtered(self):
        profile = _tabular_profile()
        assert profile.segment("REF").notes == (
            "The Bluewater vendor reference is required in the heading of every order.",
        )
        assert profile.segment("PO1").notes == (
            "Every order line must carry the NDC in the ND qualifier pair.",
        )
        assert not any("Example" in n for s in profile.segments for n in s.notes)
        assert not any("|" in n for s in profile.segments for n in s.notes)

    def test_page_furniture_never_becomes_codes(self):
        """Running headers and dated footers are stripped before parsing."""
        profile = _tabular_profile()
        all_codes = {
            c.code for s in profile.segments for e in s.elements for c in e.codes
        }
        assert "850" not in all_codes  # the footer line's first token
        assert not any("Bluewater" in c.name for s in profile.segments
                       for e in s.elements for c in e.codes)

    def test_loop_and_level_metadata(self):
        profile = _tabular_profile()
        assert profile.segment("N1").loop == "N1"
        assert profile.segment("PO1").loop == "PO1"
        assert profile.segment("BEG").loop == ""
        assert profile.segment("DTM").max_use == "10"

    def test_pdf_twin_extracts_identically(self):
        pytest.importorskip("pdfplumber")
        txt = _tabular_profile().to_dict()
        pdf = parse_guide(TABULAR_PDF, transaction="850", partner="bluewater_medical").to_dict()
        txt.pop("source")
        pdf.pop("source")
        assert txt == pdf

    def test_deterministic(self):
        assert _tabular_profile().to_dict() == _tabular_profile().to_dict()


class TestTabularCrossCheck:
    def _parse_variant(self, tmp_path: Path, transform) -> GuideProfile:
        text = TABULAR_TXT.read_text(encoding="utf-8")
        variant = tmp_path / "variant.txt"
        variant.write_text(transform(text), encoding="utf-8")
        return parse_guide(variant, transaction="850", partner="x")

    def test_clean_guide_has_no_cross_check_review(self):
        assert _tabular_profile().review == []

    def test_usage_disagreement_is_review(self, tmp_path: Path):
        profile = self._parse_variant(
            tmp_path,
            lambda t: t.replace(
                "3 150 DTM Requested Delivery O M 10",
                "3 150 DTM Requested Delivery O 10",
            ),
        )
        assert any(
            "DTM at position 150" in note and "index table says used" in note
            for note in profile.review
        )

    def test_index_row_without_detail_page_is_review(self, tmp_path: Path):
        profile = self._parse_variant(
            tmp_path,
            lambda t: t.replace(
                "Summary:\nPage Pos. Seg. Base User Loop",
                "Summary:\nPage Pos. Seg. Base User Loop\n"
                "5 020 AMT Total Order Amount M M 1",
            ),
        )
        assert any(
            "index table lists AMT at position 020" in note
            and "no detail page defines it" in note
            for note in profile.review
        )

    def test_detail_without_index_row_is_review(self, tmp_path: Path):
        profile = self._parse_variant(
            tmp_path,
            lambda t: t.replace("2 050 REF Vendor Reference C M >1\n", ""),
        )
        assert any(
            "detail pages define REF at position 050" in note
            for note in profile.review
        )

    def test_no_index_at_all_skips_the_cross_check(self, tmp_path: Path):
        text = TABULAR_TXT.read_text(encoding="utf-8")
        # drop the whole front page (the index) — layout variant, not defect
        body = "\f".join(text.split("\f")[1:])
        variant = tmp_path / "noindex.txt"
        variant.write_text(body, encoding="utf-8")
        profile = parse_guide(variant, transaction="850", partner="x")
        assert profile.review == []
        assert len(profile.segments) == 7


class TestTabularOverlay:
    def test_same_rule_shapes_as_the_templated_family(self):
        """The whole point of one GuideProfile: emission is family-blind."""
        rules = emit_partner_rules(_tabular_profile())
        segments = {(r.segment, r.qualifier) for r in rules.required_segments}
        assert ("DTM", "002") in segments  # single-code qualified family
        assert ("REF", "VR") in segments  # partner-mandatory REF, qualified
        assert ("BEG", None) in segments and ("PO1", None) in segments
        # PO106 is multi-code: positional fallback plus a review note
        assert any("PO106" in note and "multiple codes" in note for note in rules.review)
        refs = {(r.segment, r.element) for r in rules.required_elements}
        assert ("PO1", 6) in refs  # multi-code slot stays positional
        assert ("PO1", 1) in refs  # user-mandatory PO101

    def test_scoped_element_rules_from_tabular_profiles(self):
        rules = emit_partner_rules(_tabular_profile())
        scoped = {
            (r.segment, r.element, r.qualifier)
            for r in rules.required_elements
            if r.qualifier is not None
        }
        assert ("DTM", 2, "002") in scoped


# --------------------------------------------------------------------------
# Qualifier-scoped enforcement (Design 015): the truth table as tests
# --------------------------------------------------------------------------

AUDIT_FILE_4 = REPO / "tests" / "fixtures" / "audit_kit" / "850_04_partner_rules.edi"


def _presence_findings(edi_text: str, tmp_path: Path):
    """Parse an 850, apply the acme overlay, run the presence checks."""
    from mapcheck.engine.results import Category, RunResult
    from mapcheck.engine.validator import (
        _check_required_elements,
        _check_required_pairs,
        _check_required_segments,
    )
    from mapcheck.x12.parser import parse_transaction

    path = tmp_path / "variant.edi"
    path.write_text(edi_text, encoding="utf-8")
    tx = parse_transaction(str(path))
    tx.definition = emit_partner_rules(_profile()).apply(tx.definition)
    result = RunResult(spec_path="", source_path=str(path), output_path="")
    _check_required_elements(tx, result, Category.SOURCE_DATA, "")
    _check_required_segments(tx, result, Category.SOURCE_DATA, "")
    _check_required_pairs(tx, result, Category.SOURCE_DATA, "")
    return result.findings


class TestQualifierScopedEnforcement:
    """Pair rules are code-keyed, scoped element rules occurrence-keyed —
    the exact semantics the positional v1 could not express."""

    def _base(self) -> str:
        return AUDIT_FILE_4.read_text(encoding="utf-8")

    def test_file4_fails_on_pairs_not_positions(self, tmp_path: Path):
        findings = _presence_findings(self._base(), tmp_path)
        messages = sorted(f.message for f in findings)
        assert len(messages) == 4
        assert sum("no UP (UPC Consumer Package Code) qualifier pair" in m for m in messages) == 2
        assert any("required segment DTM*002" in m for m in messages)
        assert any("required segment REF*IA" in m for m in messages)

    def test_swapped_pair_order_is_not_a_failure(self, tmp_path: Path):
        """UP in the first slot pair satisfies 'must carry UP' — the
        positional false failure is gone."""
        edi = self._base().replace(
            "PO1*1*24*EA*18.75**VN*SAF-00110~",
            "PO1*1*24*EA*18.75**UP*036000291452*VN*SAF-00110~",
        ).replace(
            "PO1*2*12*CA*96.00**VN*SAF-00244~",
            "PO1*2*12*CA*96.00**UP*036000291469*VN*SAF-00244~",
        )
        findings = _presence_findings(edi, tmp_path)
        assert not any("qualifier pair" in f.message for f in findings)

    def test_filled_slots_without_the_code_still_fail(self, tmp_path: Path):
        """VN+BP fills both slot pairs but carries no UP — the defect the
        positional check missed."""
        edi = self._base().replace(
            "PO1*1*24*EA*18.75**VN*SAF-00110~",
            "PO1*1*24*EA*18.75**VN*SAF-00110*BP*88213~",
        )
        findings = _presence_findings(edi, tmp_path)
        line1 = [f for f in findings if f.source_ref == "PO1 #1"]
        assert len(line1) == 1
        assert "no UP (UPC Consumer Package Code) qualifier pair" in line1[0].message

    def test_pair_qualifier_with_empty_value_fails(self, tmp_path: Path):
        edi = self._base().replace(
            "PO1*1*24*EA*18.75**VN*SAF-00110~",
            "PO1*1*24*EA*18.75**VN*SAF-00110*UP*~",
        )
        findings = _presence_findings(edi, tmp_path)
        assert any(
            f.source_ref == "PO1 #1" and "qualifier pair" in f.message
            for f in findings
        )

    def test_scoped_element_fires_only_in_its_occurrence(self, tmp_path: Path):
        """REF*IA with an empty REF02 fails; the REF*DP occurrence with a
        filled 02 is untouched; presence of REF*IA itself passes."""
        edi = self._base().replace("REF*DP*054~", "REF*DP*054~REF*IA~")
        findings = _presence_findings(edi, tmp_path)
        scoped = [f for f in findings if "REF02" in f.message]
        assert len(scoped) == 1
        assert scoped[0].message == (
            "source data invalid: REF02 (Reference Identification) in REF*IA "
            "is required by partner acme_pharma but is empty"
        )
        assert not any("required segment REF*IA" in f.message for f in findings)

    def test_scoped_element_is_vacuous_when_occurrence_absent(self, tmp_path: Path):
        """No REF*IA at all → the presence rule reports it; the scoped
        element rule stays silent instead of double-reporting."""
        findings = _presence_findings(self._base(), tmp_path)
        assert not any("REF02" in f.message for f in findings)
        assert any("required segment REF*IA" in f.message for f in findings)


class TestQualifierRuleLoading:
    def test_definition_yaml_round_trip(self):
        from mapcheck.transactions.loader import parse_definition
        from mapcheck.transactions.schema import RequiredPairDef

        data = {
            "transaction": {"set": "850", "name": "PO", "functional_group": "PO"},
            "structure": [{"area": "heading", "segments": ["BEG"]}],
            "required_elements": [
                {"segment": "ref", "element": 2, "qualifier": "IA"},
            ],
            "required_pairs": [
                {"segment": "po1", "code": "UP", "name": "UPC"},
                {"segment": "LIN", "code": "BP", "first_element": 2,
                 "last_element": 30, "step": 2},
            ],
        }
        definition = parse_definition(data, source="inline")
        assert definition.required_elements[0].qualifier == "IA"
        assert definition.required_elements[0].qualifier_element == 1
        assert definition.required_pairs == (
            RequiredPairDef(segment="PO1", code="UP", name="UPC"),
            RequiredPairDef(segment="LIN", code="BP", first_element=2,
                            last_element=30, step=2),
        )

    def test_loader_rejects_bad_pair_entries(self):
        from mapcheck.transactions.loader import DefinitionError, parse_definition

        data = {
            "transaction": {"set": "850", "name": "PO", "functional_group": "PO"},
            "structure": [{"area": "heading", "segments": ["BEG"]}],
            "required_pairs": [
                {"segment": "PO1"},
                {"segment": "PO1", "code": "UP", "first_element": 0},
                {"segment": "PO1", "code": "UP", "last_element": 2},
            ],
        }
        with pytest.raises(DefinitionError) as err:
            parse_definition(data, source="inline")
        message = str(err.value)
        assert "required_pairs[0]" in message
        assert "required_pairs[1].first_element" in message
        assert "required_pairs[2].last_element" in message

    def test_overlay_load_validates_pairs(self, tmp_path: Path):
        path = tmp_path / "bad.yaml"
        path.write_text(
            yaml.safe_dump(
                {
                    "transaction": "850",
                    "partner": "x",
                    "required_pairs": [
                        {"code": "UP"},
                        {"segment": "PO1", "code": "UP", "step": 0},
                    ],
                }
            ),
            encoding="utf-8",
        )
        with pytest.raises(PartnerRulesError) as err:
            PartnerRules.load(path)
        message = str(err.value)
        assert "required_pairs[0]" in message
        assert "required_pairs[1]" in message


# --------------------------------------------------------------------------
# Design 018: loop-scoped placement and multi-code qualifiers
# --------------------------------------------------------------------------


def _apply_findings(edi_text: str, rules: "PartnerRules", tmp_path: Path):
    from mapcheck.engine.results import Category, RunResult
    from mapcheck.engine.validator import (
        _check_required_elements,
        _check_required_pairs,
        _check_required_segments,
    )
    from mapcheck.x12.parser import parse_transaction

    path = tmp_path / "v.edi"
    path.write_text(edi_text, encoding="utf-8")
    tx = parse_transaction(str(path))
    tx.definition = rules.apply(tx.definition)
    result = RunResult(spec_path="", source_path=str(path), output_path="")
    _check_required_elements(tx, result, Category.SOURCE_DATA, "")
    _check_required_segments(tx, result, Category.SOURCE_DATA, "")
    _check_required_pairs(tx, result, Category.SOURCE_DATA, "")
    return result.findings


#: A valid 850 with a heading PER but no PER inside the N1 loop, and a
#: REF*DP in the header — the shapes the two caveats need.
_EDI_850 = (
    "ISA*00*          *00*          *ZZ*SENDER         *ZZ*RECEIVER       "
    "*260810*0915*U*00401*000000101*0*T*>~"
    "GS*PO*SENDER*RECEIVER*20260810*0915*101*X*004010~"
    "ST*850*0001~"
    "BEG*00*NE*PO-1*20260810~"
    "REF*DP*054~"
    "PER*BD*JANE BUYER*TE*6145550142~"
    "N1*BY*BUYER CORP*92*BUY001~"
    "N1*ST*SHIP DEST*92*ST01~"
    "N3*1 DISTRIBUTION DR~"
    "N4*COLUMBUS*OH*43228~"
    "PO1*1*24*EA*18.75**VN*SAF-1~"
    "CTT*1*24~"
    "SE*11*0001~GE*1*101~IEA*1*000000101~"
)


class TestLoopScopedPlacement:
    def _rules(self, **kw):
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredSegmentDef

        return PartnerRules(
            partner="acme", transaction="850",
            required_segments=[RequiredSegmentDef(segment="PER", origin="acme", **kw)],
        )

    def test_global_rule_satisfied_by_heading_occurrence(self, tmp_path: Path):
        findings = _apply_findings(_EDI_850, self._rules(), tmp_path)
        assert not findings  # a heading PER satisfies an unscoped PER rule

    def test_loop_scoped_rule_ignores_the_heading_occurrence(self, tmp_path: Path):
        findings = _apply_findings(_EDI_850, self._rules(loop="N1"), tmp_path)
        assert len(findings) == 1
        assert findings[0].source_ref == "PER"
        assert "within the N1 loop" in findings[0].message

    def test_loop_scoped_rule_satisfied_by_in_loop_occurrence(self, tmp_path: Path):
        # add a PER inside the ship-to N1 loop
        edi = _EDI_850.replace(
            "N4*COLUMBUS*OH*43228~",
            "N4*COLUMBUS*OH*43228~PER*IC*DESK*TE*6145550000~",
        )
        findings = _apply_findings(edi, self._rules(loop="N1"), tmp_path)
        assert not findings

    def test_qualified_loop_context_matches_only_that_occurrence(self, tmp_path: Path):
        # PER inside the BY-qualified N1 only; a rule scoped to N1[ST] fails
        edi = _EDI_850.replace(
            "N1*BY*BUYER CORP*92*BUY001~",
            "N1*BY*BUYER CORP*92*BUY001~PER*IC*BUYDESK*TE*6145551111~",
        )
        assert not _apply_findings(edi, self._rules(loop="N1[BY]"), tmp_path)
        fail = _apply_findings(edi, self._rules(loop="N1[ST]"), tmp_path)
        assert len(fail) == 1 and "within the N1[ST] loop" in fail[0].message

    def test_bare_loop_matches_numbered_occurrence_labels(self, tmp_path: Path):
        """The line loop labels occurrences 'PO1 #1', not 'PO1[...]'; a bare
        'PO1' loop scope must match those via the ' #' branch of
        _label_in_loop, and must not match a different loop."""
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredSegmentDef

        def rule(loop):
            return PartnerRules(
                partner="acme", transaction="850",
                required_segments=[
                    RequiredSegmentDef(segment="PO1", loop=loop, origin="acme")
                ],
            )

        # the PO1 line occurrence is labeled "PO1 #1" — a bare "PO1" scope finds it
        assert not _apply_findings(_EDI_850, rule("PO1"), tmp_path)
        # scoping the same rule to the N1 loop finds no PO1 there → fails
        fail = _apply_findings(_EDI_850, rule("N1"), tmp_path)
        assert len(fail) == 1 and "within the N1 loop" in fail[0].message


class TestLoopScopedElements:
    """The element-level loop filter and its message path, end to end."""

    def _el_rules(self, **kw):
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredElementDef

        return PartnerRules(
            partner="acme", transaction="850",
            required_elements=[RequiredElementDef(segment="PER", element=2, origin="acme", **kw)],
        )

    def test_loop_scoped_element_ignores_heading_occurrence(self, tmp_path: Path):
        # the file's only PER (heading) has an empty PER02; a global rule
        # fails it, but that is not the point — scope it to N1 and the
        # heading occurrence is out of scope, so no finding.
        edi = _EDI_850.replace(
            "PER*BD*JANE BUYER*TE*6145550142~", "PER*BD**TE*6145550142~"
        )
        assert not _apply_findings(edi, self._el_rules(loop="N1"), tmp_path)

    def test_loop_scoped_element_fires_inside_the_loop(self, tmp_path: Path):
        edi = _EDI_850.replace(
            "N4*COLUMBUS*OH*43228~",
            "N4*COLUMBUS*OH*43228~PER*IC**TE*6145550000~",  # PER02 empty in-loop
        )
        findings = _apply_findings(edi, self._el_rules(loop="N1"), tmp_path)
        assert len(findings) == 1
        assert "within the N1 loop" in findings[0].message

    def test_loop_and_qualifier_compose_on_a_segment(self, tmp_path: Path):
        """A rule carrying BOTH a loop scope and a qualifier: PER*IC within
        N1 — satisfied only by an IC-qualified PER inside the N1 loop."""
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredSegmentDef

        rules = PartnerRules(
            partner="acme", transaction="850",
            required_segments=[
                RequiredSegmentDef(segment="PER", qualifier="IC", loop="N1", origin="acme")
            ],
        )
        # heading PER is BD, not IC, and not in N1 → fails
        fail = _apply_findings(_EDI_850, rules, tmp_path)
        assert len(fail) == 1
        assert fail[0].source_ref == "PER*IC"
        assert "within the N1 loop" in fail[0].message
        # an IC-qualified PER inside the N1 loop satisfies it
        edi = _EDI_850.replace(
            "N4*COLUMBUS*OH*43228~",
            "N4*COLUMBUS*OH*43228~PER*IC*DESK*TE*6145550000~",
        )
        assert not _apply_findings(edi, rules, tmp_path)


class TestMultiCodeEnforcement:
    def _rules(self, qualifiers):
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredSegmentDef

        return PartnerRules(
            partner="acme", transaction="850",
            required_segments=[
                RequiredSegmentDef(segment="REF", qualifiers=qualifiers, origin="acme")
            ],
        )

    def test_either_code_satisfies_the_set(self, tmp_path: Path):
        # the file carries REF*DP; a {IA, DP} rule is satisfied
        assert not _apply_findings(_EDI_850, self._rules(("IA", "DP")), tmp_path)

    def test_no_code_in_the_set_fails_naming_the_set(self, tmp_path: Path):
        findings = _apply_findings(_EDI_850, self._rules(("IA", "ZZ")), tmp_path)
        assert len(findings) == 1
        assert findings[0].source_ref == "REF*{IA|ZZ}"
        assert "REF*{IA|ZZ}" in findings[0].message

    def test_scoped_element_over_a_code_set(self, tmp_path: Path):
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredElementDef

        # REF03 required in REF*{IA|DP}; the file's REF*DP has no REF03
        rules = PartnerRules(
            partner="acme", transaction="850",
            required_elements=[
                RequiredElementDef(
                    segment="REF", element=3, qualifiers=("IA", "DP"), origin="acme"
                )
            ],
        )
        findings = _apply_findings(_EDI_850, rules, tmp_path)
        assert len(findings) == 1
        assert "in REF*{IA|DP}" in findings[0].message


class TestDesign018Loader:
    def test_overlay_round_trip_loop_and_qualifiers(self, tmp_path: Path):
        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import (
            RequiredElementDef,
            RequiredSegmentDef,
        )

        rules = PartnerRules(
            partner="acme", transaction="850",
            required_segments=[
                RequiredSegmentDef(segment="PER", loop="N1", origin="acme"),
                RequiredSegmentDef(
                    segment="REF", qualifiers=("IA", "DP"), origin="acme"
                ),
            ],
            required_elements=[
                RequiredElementDef(
                    segment="REF", element=2, qualifiers=("IA", "DP"),
                    loop="", origin="acme",
                )
            ],
        )
        again = PartnerRules.load(rules.save(tmp_path / "r.yaml"))
        assert again.to_dict() == rules.to_dict()
        assert again.required_segments[0].loop == "N1"
        assert again.required_segments[1].qualifiers == ("IA", "DP")

    def test_qualifier_and_qualifiers_together_is_an_error(self, tmp_path: Path):
        from mapcheck.guides.overlay import PartnerRules, PartnerRulesError

        path = tmp_path / "bad.yaml"
        path.write_text(
            yaml.safe_dump(
                {
                    "transaction": "850", "partner": "x",
                    "required_segments": [
                        {"segment": "REF", "qualifier": "IA", "qualifiers": ["IA", "DP"]}
                    ],
                }
            ),
            encoding="utf-8",
        )
        with pytest.raises(PartnerRulesError, match="not both"):
            PartnerRules.load(path)

    def test_definition_loader_parses_both(self):
        from mapcheck.transactions.loader import parse_definition

        data = {
            "transaction": {"set": "850", "name": "PO", "functional_group": "PO"},
            "structure": [{"area": "heading", "segments": ["BEG"]}],
            "required_segments": [
                {"segment": "per", "loop": "N1"},
                {"segment": "ref", "qualifiers": ["IA", "DP"]},
            ],
        }
        definition = parse_definition(data, source="inline")
        assert definition.required_segments[0].loop == "N1"
        assert definition.required_segments[1].qualifiers == ("IA", "DP")


class TestDesign018Apply:
    def test_qualifiers_dedup_is_order_insensitive(self):
        """apply() treats {IA,DP} and {DP,IA} as the same rule (frozenset
        key), so a reordered base rule suppresses the overlay duplicate."""
        import dataclasses

        from mapcheck.guides.overlay import PartnerRules
        from mapcheck.transactions.schema import RequiredSegmentDef

        base = dataclasses.replace(
            DEFINITION_850,
            required_segments=(RequiredSegmentDef(segment="REF", qualifiers=("IA", "DP")),),
        )
        overlay = PartnerRules(
            partner="x", transaction="850",
            required_segments=[
                RequiredSegmentDef(segment="REF", qualifiers=("DP", "IA"), origin="x")
            ],
        )
        merged = overlay.apply(base)
        ref_rules = [r for r in merged.required_segments if r.segment == "REF"]
        assert len(ref_rules) == 1  # the reordered duplicate was deduped


class TestDesign018Emission:
    def test_loop_member_segment_gets_a_loop_scope(self):
        """A must_use segment inside a loop it does not trigger is
        loop-scoped; the loop trigger is not."""
        profile = _tabular_profile()  # Bluewater: N3 is must_use? check N1/PO1
        rules = emit_partner_rules(profile)
        by_seg = {}
        for r in rules.required_segments:
            by_seg.setdefault(r.segment, []).append(r)
        # N1 is the trigger of loop N1 → not loop-scoped
        assert all(r.loop == "" for r in by_seg.get("N1", []))
        # PO1 is the trigger of loop PO1 → not loop-scoped
        assert all(r.loop == "" for r in by_seg.get("PO1", []))

    def test_loop_member_scope_from_a_synthetic_member(self, tmp_path: Path):
        guide = tmp_path / "g.txt"
        guide.write_text(
            "N1 Name Pos: 300 Max: 1\n"
            "Loop: N1\n"
            "User Option (Usage): Must use\n\n"
            "Element Summary:\n"
            "N101 98 Entity Identifier Code M ID 2/3 Must use\n\n"
            "N3 Address Information Pos: 330 Max: 1\n"
            "Loop: N1\n"
            "User Option (Usage): Must use\n\n"
            "Element Summary:\n"
            "N301 166 Address Information M AN 1/55 Must use\n",
            encoding="utf-8",
        )
        rules = emit_partner_rules(parse_guide(guide, transaction="850", partner="x"))
        by_seg = {r.segment: r for r in rules.required_segments}
        assert by_seg["N1"].loop == ""  # trigger
        assert by_seg["N3"].loop == "N1"  # member scoped to its loop
        n301 = [r for r in rules.required_elements if r.segment == "N3"]
        assert n301 and all(r.loop == "N1" for r in n301)


# --------------------------------------------------------------------------
# Guided drafts (consumer one)
# --------------------------------------------------------------------------


def _starter() -> Crosswalk:
    return load_crosswalks(bundled_crosswalks("850", "orders05"))


class TestGuidedDraft:
    def test_partner_notes_flow_to_filled_rows(self):
        result = assemble_draft(DEFINITION_850, ORDERS05, _starter(), guide=_profile())
        beg05_rows = [r for r in result.rows if r.source == "BEG05"]
        assert beg05_rows
        assert all(
            "Partner note: This is the date assigned by Acme Pharma" in r.notes
            for r in beg05_rows
        )

    def test_not_used_elements_leave_unmapped_source(self):
        plain = assemble_draft(DEFINITION_850, ORDERS05, _starter())
        guided = assemble_draft(DEFINITION_850, ORDERS05, _starter(), guide=_profile())
        plain_fields = {el.field for el in plain.unmapped_source}
        guided_fields = {el.field for el in guided.unmapped_source}
        assert "SAC01" in plain_fields and "SAC01" not in guided_fields
        assert "PO105" in plain_fields and "PO105" not in guided_fields
        # elements the guide does not mention stay
        assert "SAC03" in guided_fields

    def test_must_use_markers_on_unmapped_source(self):
        result = assemble_draft(DEFINITION_850, ORDERS05, _starter(), guide=_profile())
        usage = {el.field: el.partner_usage for el in result.unmapped_source}
        assert usage["PO108"] == "Must use"
        assert usage["PO109"] == "Must use"
        assert usage["CTT02"] == "Must use"
        assert usage["PER01"] == "Used"
        assert usage["SAC03"] == ""  # not in the guide

    def test_code_lists_filter_to_partner_subset(self):
        result = assemble_draft(DEFINITION_850, ORDERS05, _starter(), guide=_profile())
        assert [e.source for e in result.code_lists["SAP_ACTION"]] == ["00"]
        assert [e.source for e in result.code_lists["PO_TYPE_SAP"]] == ["NE"]
        # PO103 carries no code subset in the guide: UOM_SAP is untouched
        assert [e.source for e in result.code_lists["UOM_SAP"]] == ["EA", "CA", "DZ"]
        assert result.guide_review == ()

    def test_partner_code_without_translation_is_review(self, tmp_path: Path):
        crosswalk_file = tmp_path / "xw.yaml"
        crosswalk_file.write_text(
            yaml.safe_dump(
                {
                    "rules": [
                        {
                            "source": "BEG02",
                            "target": "header.bsart",
                            "rule": "CODE_LIST",
                            "code_list": "PO_TYPE",
                        }
                    ],
                    "code_lists": {
                        "PO_TYPE": [{"source": "SA", "target": "NB"}],
                    },
                }
            ),
            encoding="utf-8",
        )
        crosswalk = load_crosswalks([crosswalk_file])
        result = assemble_draft(DEFINITION_850, ORDERS05, crosswalk, guide=_profile())
        assert result.code_lists["PO_TYPE"] == ()  # SA is not a partner code
        assert any(
            "partner code NE (New Order) has no crosswalk translation" in note
            for note in result.guide_review
        )

    def test_mapped_but_not_used_element_is_review(self, tmp_path: Path):
        crosswalk_file = tmp_path / "xw.yaml"
        crosswalk_file.write_text(
            yaml.safe_dump(
                {
                    "rules": [
                        {"source": "BEG04", "target": "org.012", "rule": "DIRECT"}
                    ]
                }
            ),
            encoding="utf-8",
        )
        crosswalk = load_crosswalks([crosswalk_file])
        result = assemble_draft(DEFINITION_850, ORDERS05, crosswalk, guide=_profile())
        assert any(
            "BEG04 is mapped to org.012 but the guide marks it Not used" in note
            for note in result.guide_review
        )

    def test_unguided_draft_is_unchanged(self):
        result = assemble_draft(DEFINITION_850, ORDERS05, _starter())
        assert result.guided is False
        assert result.guide_review == ()
        assert all(el.partner_usage == "" for el in result.unmapped_source)

    def test_guided_workbook_content(self, tmp_path: Path):
        out = tmp_path / "draft.xlsx"
        generate_draft(
            DEFINITION_850, ORDERS05, bundled_crosswalks("850", "orders05"), out,
            guide=_profile(),
        )
        wb = load_workbook(out)
        ws = wb["Unmapped Source"]
        headers = [c.value for c in ws[1]]
        assert headers[:4] == ["Source Field", "Loop Context", "Element Name", "Partner Usage"]
        rows = {row[0]: row[3] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
        assert rows["PO108"] == "Must use"
        meta = {
            row[0]: row[1]
            for row in wb["Meta"].iter_rows(values_only=True)
            if row and row[0]
        }
        assert meta["Guide"] == "acme_pharma — acme_pharma_850_guide.txt"

    def test_guided_draft_deterministic(self, tmp_path: Path):
        results = [
            assemble_draft(DEFINITION_850, ORDERS05, _starter(), guide=_profile())
            for _ in range(2)
        ]
        assert results[0] == results[1]


# --------------------------------------------------------------------------
# CLI verbs
# --------------------------------------------------------------------------


class TestCli:
    def test_import_guide_writes_profile_and_overlay(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        profile_path = tmp_path / "acme.yaml"
        overlay_path = tmp_path / "acme_rules.yaml"
        code = main(
            [
                "import-guide", str(GUIDE_TXT),
                "--transaction", "850", "--partner", "acme_pharma",
                "--profile", str(profile_path), "--overlay", str(overlay_path),
            ]
        )
        assert code == 0
        out = capsys.readouterr().out
        assert "parse coverage 1.00" in out
        assert GuideProfile.load(profile_path).partner == "acme_pharma"
        assert PartnerRules.load(overlay_path).transaction == "850"

    def test_import_guide_refuses_overwrite(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        existing = tmp_path / "acme.yaml"
        existing.write_text("x", encoding="utf-8")
        code = main(
            [
                "import-guide", str(GUIDE_TXT),
                "--transaction", "850", "--partner", "acme_pharma",
                "--profile", str(existing),
            ]
        )
        assert code == 2
        assert "already exists" in capsys.readouterr().err

    def test_import_guide_notes_unregistered_set(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        code = main(
            [
                "import-guide", str(GUIDE_TXT),
                "--transaction", "832", "--partner", "x",
                "--profile", str(tmp_path / "p.yaml"),
            ]
        )
        assert code == 0
        assert "no registered definition" in capsys.readouterr().err

    def test_draft_spec_accepts_profile_and_raw_guide(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        profile_path = tmp_path / "acme.yaml"
        _profile().save(profile_path)
        for name, guide_arg in (
            ("from_profile.xlsx", str(profile_path)),
            ("from_raw.xlsx", str(GUIDE_TXT)),
        ):
            code = main(
                [
                    "draft-spec", "--transaction", "850", "--target", "orders05",
                    "--guide", guide_arg, "--output", str(tmp_path / name),
                ]
            )
            assert code == 0
        out = capsys.readouterr().out
        assert "flavored by partner acme_pharma" in out  # profile path
        assert "flavored by the partner guide" in out  # raw path, unnamed

    def test_draft_spec_rejects_wrong_transaction_profile(self, tmp_path: Path, capsys):
        from mapcheck.cli import main

        profile = _profile()
        profile.transaction = "810"
        profile_path = profile.save(tmp_path / "other.yaml")
        code = main(
            [
                "draft-spec", "--transaction", "850", "--target", "orders05",
                "--guide", str(profile_path), "--output", str(tmp_path / "d.xlsx"),
            ]
        )
        assert code == 2
        assert "transaction set 810" in capsys.readouterr().err
